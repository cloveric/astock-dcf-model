# -*- coding: utf-8 -*-
"""
模型验收脚本 (通用版): LibreOffice headless 重算 + 逐项校验

用法:
  python verify_model.py --code 300476                 # 找 out/ 下xlsx+addr, 重算并校验
  python verify_model.py out/300476_胜宏科技_估值模型.xlsx
  python verify_model.py <xlsx> --no-recalc            # 直接读已重算文件
校验项: Checks布尔总闸 / BS配平 / 现金=最低现金 / FIN迭代残差 / CF-BS现金勾稽 /
        WACC采用值=计算值 / 年中折现t序 / 基准情景=主模型 / 敏感性矩阵中心=DCF /
        黄金值断言(基准标的DCF每股复现, 跨LO版本浮点容差 + config指纹门, 见GOLDEN_DCF_PS)。
"""
import argparse
import hashlib
import json
import os
import re
import shutil
import statistics
import subprocess
import sys
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# B3 黄金值: 基准标的构建后DCF每股应复现
# - 断言用容差而非精确==: abs(ps-golden) <= max(1e-6, abs(golden)*1e-9) (跨LibreOffice版本浮点容差)
# - config指纹门: 冻结value时同时冻结 configs/<code>.yaml 的sha256;
#   默认配置指纹漂移会FAIL；仅 --accept-golden-config-change 可显式标为WAIVED。
GOLDEN_DCF_PS = {
    '300476': {
        'value': 340.415208186145,
        'config_sha256': '590d14283ce2f28665b2b0bf356054b4332a23da2dd624f6a00cc2ec8c7baea0',
    },
}


def _num(v):
    """E2 守卫: data_only 读回的标量, 仅真数值(int/float, 排除bool)→float;
    None / 错误串('#DIV/0!'等) / 其他类型 → None, 由调用方记FAIL, 不抛TypeError。"""
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        return None
    return float(v)


def fmtv(v, spec='.2f'):
    """E2 守卫式格式化: 数值按spec格式化; 非数值(None/错误串)原样repr输出, 不崩溃"""
    n = _num(v)
    return format(n, spec) if n is not None else repr(v)


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, 'rb') as fh:
        for chunk in iter(lambda: fh.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _golden_value(code, cfg_path):
    """Legacy lookup helper retained for API compatibility.

    Production verification uses :func:`_golden_control`, where default-config
    fingerprint drift is a failure rather than this historical lookup's ``None``.
    """
    g = GOLDEN_DCF_PS.get(str(code))
    if not g:
        return None
    cur = _sha256_file(cfg_path) if (cfg_path and os.path.exists(cfg_path)) else None
    if cur != g.get('config_sha256'):
        print('  [INFO] 配置已变化, 跳过黄金值断言')
        return None
    return g['value']


def _control_result(name, status, detail='', **metrics):
    """Return one JSON-serializable verifier control verdict."""
    result = {'name': name, 'status': status, 'detail': detail}
    if metrics:
        result['metrics'] = metrics
    return result


def _golden_control(code, actual_value, meta, repo_root=None,
                    accept_config_change=False):
    """Evaluate the frozen DCF value only for a proven default config build.

    New addr metadata makes provenance unambiguous.  Legacy addr files are
    deliberately marked REVIEW instead of being compared to an unrelated
    repository config or silently skipping a changed default config.
    """
    code = str(code or '')
    golden = GOLDEN_DCF_PS.get(code)
    if not golden:
        return _control_result('黄金值控制', 'NOT_APPLICABLE', '该代码未冻结黄金值')

    meta = meta or {}
    root = os.path.abspath(repo_root or os.path.dirname(os.path.abspath(__file__)))
    default_path = os.path.join(root, 'configs', f'{code}.yaml')
    config_path = (meta.get('config_path') or meta.get('config_source') or
                   meta.get('source_config'))
    config_scope = meta.get('config_path_scope')
    fingerprint = (meta.get('config_sha256') or meta.get('config_fingerprint') or
                   meta.get('source_fingerprint'))
    is_default = meta.get('is_default_config')

    resolved_config_path = None
    if config_path:
        if os.path.isabs(config_path):
            resolved_config_path = os.path.realpath(config_path)  # legacy addr compatibility
        elif config_scope == 'project':
            resolved_config_path = os.path.realpath(os.path.join(root, config_path))
        elif config_scope != 'external':
            resolved_config_path = os.path.realpath(os.path.abspath(config_path))

    if is_default is None and resolved_config_path:
        is_default = resolved_config_path == os.path.realpath(default_path)
    if is_default is False:
        return _control_result(
            '黄金值控制', 'NOT_APPLICABLE', '自定义配置不与默认配置黄金值比较',
            config_path=config_path or '',
        )
    if is_default is None:
        if fingerprint == golden.get('config_sha256'):
            # Backwards-compatible bridge for metadata that has a fingerprint
            # but predates the explicit is_default_config field.
            is_default = True
        else:
            return _control_result(
                '黄金值控制', 'REVIEW',
                'addr缺少默认/自定义配置来源, 无法安全套用黄金值',
                observed_sha256=fingerprint or '',
            )

    if resolved_config_path and os.path.exists(resolved_config_path):
        source_fingerprint = _sha256_file(resolved_config_path)
        if fingerprint and fingerprint != source_fingerprint:
            return _control_result(
                '黄金值控制', 'FAIL',
                'addr配置指纹与当前来源文件不一致; 产物/addr/config可能未配套',
                observed_sha256=fingerprint, source_sha256=source_fingerprint,
            )
        if not fingerprint:
            fingerprint = source_fingerprint
    if not fingerprint and os.path.exists(default_path):
        fingerprint = _sha256_file(default_path)
    if not fingerprint:
        return _control_result('黄金值控制', 'FAIL', '默认配置构建缺少可验证指纹')

    frozen_hash = golden.get('config_sha256')
    if fingerprint != frozen_hash:
        status = 'WAIVED' if accept_config_change else 'FAIL'
        detail = ('已显式接受默认配置变化; 需重审并更新黄金值'
                  if accept_config_change else
                  '默认配置指纹已变化; 禁止静默跳过黄金值断言')
        return _control_result(
            '黄金值控制', status, detail,
            observed_sha256=fingerprint, frozen_sha256=frozen_hash,
        )

    actual = _num(actual_value)
    expected = _num(golden.get('value'))
    if actual is None or expected is None:
        return _control_result(
            '黄金值: DCF每股复现', 'FAIL',
            f'黄金值或DCF值非数值: actual={actual_value!r}, expected={golden.get("value")!r}',
        )
    tolerance = max(1e-6, abs(expected) * 1e-9)
    passed = abs(actual - expected) <= tolerance
    return _control_result(
        '黄金值: DCF每股复现', 'PASS' if passed else 'FAIL',
        f'{actual!r} vs 基准{expected!r}, 容差{tolerance:.3g}',
        actual=actual, expected=expected, tolerance=tolerance,
    )


def _wacc_terminal_control(wacc, terminal_g, min_spread=0.005):
    wacc_n, growth_n = _num(wacc), _num(terminal_g)
    if wacc_n is None or growth_n is None:
        return _control_result(
            'WACC/永续增长率有效性', 'FAIL',
            f'非数值: WACC={wacc!r}, g={terminal_g!r}',
        )
    spread = wacc_n - growth_n
    passed = wacc_n > 0 and spread >= min_spread
    return _control_result(
        'WACC>g且安全边际充足', 'PASS' if passed else 'FAIL',
        f'WACC={wacc_n:.4%}, g={growth_n:.4%}, spread={spread:.2%}, 最低={min_spread:.2%}',
        wacc=wacc_n, terminal_g=growth_n, spread=spread, min_spread=min_spread,
    )


def _scenario_control(bear, base, bull, engine=None):
    values = [_num(v) for v in (bear, base, bull)]
    if any(v is None for v in values):
        return _control_result(
            '三情景单调性/同引擎', 'FAIL',
            f'情景值非数值: bear={bear!r}, base={base!r}, bull={bull!r}',
        )
    monotonic = values[0] <= values[1] <= values[2]
    if not monotonic:
        return _control_result(
            '三情景单调性/同引擎', 'FAIL',
            f'不满足 bear<=base<=bull: {values}',
            bear=values[0], base=values[1], bull=values[2], engine=engine or '',
        )
    if engine and engine != 'simplified_same_engine_v1':
        return _control_result(
            '三情景单调性/同引擎', 'REVIEW',
            f'数值单调, 但addr声明的情景引擎不是同引擎口径: {engine}',
            bear=values[0], base=values[1], bull=values[2], engine=engine,
        )
    return _control_result(
        '三情景单调性/同引擎', 'PASS',
        'bear<=base<=bull; ' + (f'engine={engine}' if engine else 'legacy addr按单调性验收'),
        bear=values[0], base=values[1], bull=values[2], engine=engine or '',
    )


def _relative_median_control(
    median_value,
    pricing_values,
    name='相对估值计价中位数',
):
    if not pricing_values:
        return _control_result(name, 'NOT_APPLICABLE', '无适用可比行')
    values = [_num(v) for v in pricing_values]
    median_n = _num(median_value)
    if median_n is None or any(v is None for v in values):
        return _control_result(
            name, 'FAIL',
            f'中位数或计价行含非数值: median={median_value!r}, rows={pricing_values!r}',
        )
    expected = float(statistics.median(values))
    tolerance = max(1e-9, abs(expected) * 1e-9)
    passed = abs(median_n - expected) <= tolerance
    return _control_result(
        name, 'PASS' if passed else 'FAIL',
        f'addr中位数={median_n:.6g}, 计价行中位数={expected:.6g}',
        actual=median_n, expected=expected, pricing_rows=len(values),
    )


def _relative_median_ref(addr):
    """Use the pricing median when it exists, otherwise the configured PE fallback."""
    if addr.get('relative_price_rows'):
        return addr.get('rel_median_cell') or addr.get('rel_med_pe')
    return addr.get('rel_med_pe') or addr.get('rel_median_cell')


def _gross_margin_control(values):
    numbers = [_num(v) for v in values]
    if not numbers or any(v is None for v in numbers):
        return _control_result('毛利率硬边界', 'FAIL', f'毛利率含非数值: {values!r}')
    # Gross margin may be negative for pre-scale or distressed businesses, but
    # outside [-100%, 100%] is an impossible/hard-bound data or formula error.
    passed = all(-1.0 <= v <= 1.0 for v in numbers)
    return _control_result(
        '毛利率硬边界', 'PASS' if passed else 'FAIL',
        f'范围=[{min(numbers):.2%}, {max(numbers):.2%}], 要求[-100%, 100%]',
        minimum=min(numbers), maximum=max(numbers),
    )


def _historical_plug_control(plugs, abs_tol=0.5, rel_tol=1e-5):
    """Reapply the builder's exact rounding-only threshold to addr metadata."""
    plugs = plugs or []
    malformed = []
    excessive = []
    for plug in plugs:
        if not isinstance(plug, dict):
            malformed.append(plug)
            continue
        diff, assets = _num(plug.get('diff')), _num(plug.get('assets'))
        if diff is None or assets is None or assets <= 0:
            malformed.append(plug)
            continue
        threshold = max(abs_tol, abs(assets) * rel_tol)
        if abs(diff) > threshold:
            excessive.append((plug, abs(diff), threshold, abs(diff) / assets))
    if malformed:
        return _control_result('历史配平plug重大性', 'FAIL', f'plug元数据不完整: {malformed!r}')
    if excessive:
        years = ', '.join(str(p.get('year', '?')) for p, *_ in excessive)
        max_ratio = max(ratio for *_, ratio in excessive)
        return _control_result(
            '历史配平plug重大性', 'FAIL',
            f'{years} 超过舍入阈值 max({abs_tol:g}, 总资产×{rel_tol:g}); '
            f'最大占总资产{max_ratio:.4%}',
            excessive_count=len(excessive), max_ratio=max_ratio,
            absolute_tolerance=abs_tol, relative_tolerance=rel_tol,
        )
    return _control_result(
        '历史配平plug重大性', 'PASS',
        f'{len(plugs)}项均满足舍入阈值 max({abs_tol:g}, 总资产×{rel_tol:g})',
        plug_count=len(plugs), absolute_tolerance=abs_tol, relative_tolerance=rel_tol,
    )


def _per_share_unit(meta):
    meta = meta or {}
    explicit = meta.get('per_share_unit') or meta.get('currency_label')
    if explicit:
        return str(explicit)
    code = str(meta.get('currency_code') or '').upper()
    if code == 'USD':
        return '美元/股'
    if code == 'HKD':
        return '港元/股'
    if code in ('CNY', 'RMB'):
        return '元/股'
    note = str(meta.get('currency_note') or meta.get('unit') or '')
    if '美元' in note:
        return '美元/股'
    if '港元' in note:
        return '港元/股'
    return '元/股'


def _fcfe_divergence_control(difference, waived=False, waiver_reason=None):
    diff = _num(difference)
    if diff is None:
        return _control_result('FCFE/FCFF差异复核', 'FAIL', f'差异非数值: {difference!r}')
    if abs(diff) < 0.30:
        return _control_result('FCFE/FCFF差异复核', 'PASS', f'差异={diff:+.1%}')
    reason = str(waiver_reason or '').strip()
    if waived or reason:
        waiver_detail = ('CLI显式豁免' if waived else f'配置复核豁免: {reason}')
        return _control_result(
            'FCFE/FCFF差异复核', 'WAIVED',
            f'差异={diff:+.1%}; {waiver_detail}', difference=diff,
            waiver_reason=reason,
        )
    return _control_result(
        'FCFE/FCFF差异复核', 'REVIEW',
        f'差异={diff:+.1%}达到30%; 需复核融资/杠杆路径或显式豁免', difference=diff,
    )


def _summarize_controls(controls):
    statuses = {item.get('status') for item in controls}
    if 'FAIL' in statuses:
        return {'verdict': 'FAIL', 'exit_code': 1, 'label': '存在FAIL项'}
    if 'REVIEW' in statuses:
        return {'verdict': 'REVIEW', 'exit_code': 2, 'label': '需要人工复核'}
    return {'verdict': 'PASS', 'exit_code': 0, 'label': 'ALL PASS'}


def _verification_payload(path, meta, controls):
    summary = _summarize_controls(controls)
    return {
        'schema_version': 1,
        'file': os.path.basename(os.fspath(path)),
        'model': dict(meta or {}),
        'verdict': summary['verdict'],
        'exit_code': summary['exit_code'],
        'label': summary['label'],
        'controls': controls,
    }


def _value_at_ref(workbook, ref, default_sheet=None):
    """Read an addr entry in Sheet!A1 form (or A1 with a default sheet)."""
    if ref is None:
        return None
    parts = str(ref).replace('$', '').split('!', 1)
    if len(parts) == 2:
        sheet, cell = parts
    elif default_sheet:
        sheet, cell = default_sheet, parts[0]
    else:
        return None
    return workbook[sheet][cell].value


def lo_recalc(src, workdir):
    """LibreOffice headless 重算 (v3模型无循环引用, 一次convert即得计算值)

    E1 加固: 独立UserInstallation profile(避免桌面LibreOffice占用默认profile导致
    --convert-to 静默no-op) + 转换前删旧输出 + 转换后断言产出了新文件。"""
    os.makedirs(workdir, exist_ok=True)
    work = os.path.join(workdir, os.path.basename(src))
    shutil.copy(src, work)
    outdir = os.path.join(workdir, 'out')
    os.makedirs(outdir, exist_ok=True)
    out_path = os.path.join(outdir, os.path.basename(src))
    if os.path.exists(out_path):  # E1: 删除已存在的目标输出, 防止转换no-op时误读陈旧文件假PASS
        os.remove(out_path)
    t0 = time.time()
    profile = 'file:///tmp/lo_verify_profile_{0}'.format(os.getpid())
    subprocess.run(['soffice', '--headless', '-env:UserInstallation=' + profile,
                    '--convert-to', 'xlsx', '--outdir', outdir, work],
                   check=True, capture_output=True, timeout=300)
    # E1: 输出必须存在且 mtime >= 转换开始时间 (留2s裕量容忍粗粒度文件系统mtime)
    if not os.path.exists(out_path) or os.path.getmtime(out_path) < t0 - 2:
        raise RuntimeError('LibreOffice 转换未产出新文件')
    return out_path


def main():
    ap = argparse.ArgumentParser(description='A股估值模型验收 (LO重算+Checks)')
    ap.add_argument('xlsx', nargs='?', help='模型xlsx路径 (或用--code)')
    ap.add_argument('--code', help='读取 out/<code>_*_估值模型.xlsx')
    ap.add_argument('--addr', help='addr json 路径 (默认随xlsx同名)')
    ap.add_argument('--no-recalc', action='store_true', help='不做LO重算, 直接读文件')
    ap.add_argument('--workdir', default=None, help='重算工作目录 (默认 .cache/lo_<code>)')
    ap.add_argument(
        '--accept-golden-config-change', action='store_true',
        help='显式接受默认配置指纹变化并将黄金值控制标为WAIVED (CI不得默认使用)',
    )
    ap.add_argument(
        '--waive-fcfe-divergence', action='store_true',
        help='显式豁免FCFE/FCFF每股差异>=30%%的人工复核闸门',
    )
    ap.add_argument(
        '--json-summary', metavar='PATH',
        help='将结构化验收摘要写入PATH; PATH为-时同时输出到stdout',
    )
    args = ap.parse_args()

    path = args.xlsx
    if not path:
        if not args.code:
            ap.error('需提供 xlsx 路径或 --code')
        base = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'out')
        cands = [f for f in os.listdir(base) if f.startswith(args.code) and f.endswith('.xlsx')]
        if not cands:
            ap.error(f'out/ 下未找到 {args.code} 的xlsx')
        # E8: 多个候选时按mtime取最新, 不再按字典序取旧文件
        cands.sort(key=lambda f: os.path.getmtime(os.path.join(base, f)), reverse=True)
        path = os.path.join(base, cands[0])
        print(f'[INFO] --code 匹配 {len(cands)} 个候选, 选用最新: {cands[0]}')
    addr_path = args.addr or os.path.splitext(path)[0] + '.addr.json'
    addr = json.load(open(addr_path, encoding='utf-8'))
    meta = addr.get('meta', {})
    hist_years = meta.get('hist_years')
    fcst_years = meta.get('fcst_years')
    if not hist_years or not fcst_years:
        # E7: 不再用写死年份兜底, 元数据缺失直接报错
        raise RuntimeError('addr.json 缺少年份元数据, 请用配套 build 重新生成')
    YRS = list(hist_years) + list(fcst_years)
    FCST = list(fcst_years)
    COLS = {y: get_column_letter(2 + i) for i, y in enumerate(YRS)}

    if args.no_recalc:
        recalced = path
    else:
        workdir = args.workdir or os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                               '.cache', f"lo_{meta.get('code', 'x')}")
        recalced = lo_recalc(path, workdir)
    WB = load_workbook(recalced, data_only=True)
    unit_meta = dict(meta)
    if not any(unit_meta.get(k) for k in ('per_share_unit', 'currency_code', 'currency_note', 'unit')):
        # Legacy addr bridge: infer the financial-statement currency from the
        # generated sheet subtitle rather than hard-coding RMB labels for HK.
        unit_meta['currency_note'] = WB['Scenarios']['A2'].value
    per_share_unit = _per_share_unit(unit_meta)

    def V(sheet, cell):
        return WB[sheet][cell].value

    def series(sheet, row, years=FCST):
        return [V(sheet, f'{COLS[y]}{row}') for y in years]

    def fmt(v, nd=1):
        if v is None:
            return 'None'
        if isinstance(v, str):
            return v
        return f'{v:,.{nd}f}'

    controls = []

    def record(result):
        controls.append(result)
        detail = result.get('detail', '')
        print(f"  [{result.get('status')}] {result.get('name')} {detail}")
        return result

    def check(name, cond, detail=''):
        return record(_control_result(name, 'PASS' if cond else 'FAIL', detail))

    def numguard(name, **kv):
        """E2: check前数值守卫. 任一值非数值→该check直接记FAIL并打印实际值, 返回None;
        全为数值→返回float字典, 调用方再做真正的条件判断。"""
        bad = {k: v for k, v in kv.items() if _num(v) is None}
        if bad:
            check(name, False, '非数值: ' + ', '.join('{0}={1!r}'.format(k, v) for k, v in bad.items()))
            return None
        return {k: _num(v) for k, v in kv.items()}

    print('=' * 70)
    print(f"文件: {os.path.basename(recalced)}")
    print(f"标的: {meta.get('name')}({meta.get('code_full')})  基准日: {meta.get('valuation_date')}")
    print('=' * 70)

    print('\n[Checks 总闸]')
    sumv = V('Checks', addr['checks_sum'].split('!')[1])
    boolv = V('Checks', addr['checks_bool'].split('!')[1])
    print('  汇总单元格:', sumv, '| 布尔:', boolv)
    check('Checks全部通过', boolv is True or (isinstance(boolv, (int, float)) and boolv == 1), f'bool={boolv}')

    print('\n[FIN 债务与融资调度]')
    diffs = series('BS', addr['bs_chk_row'], YRS)
    print('  BS配平差额 全期:', [f'{d:.6f}' if isinstance(d, (int, float)) else str(d) for d in diffs])
    bs_maxdiff = max((abs(d) for d in diffs if isinstance(d, (int, float))), default=0)
    check('① BS差额全期|diff|<0.01', all(isinstance(d, (int, float)) and abs(d) < 0.01 for d in diffs),
          f'max|diff|={bs_maxdiff:.2e}')

    cash = series('BS', addr['bs_cash_row'])
    minc = series('FIN', addr['fin_mincash_row'])
    check('② 全期现金=最低现金', all(isinstance(c, (int, float)) and isinstance(m, (int, float))
                                   and abs(c - m) < 0.01 for c, m in zip(cash, minc)),
          f'{FCST[-1]}E cash={fmt(cash[-1])} vs mincash={fmt(minc[-1])}')
    intexp = series('FIN', addr['fin_intexp_row'])
    print('  利息支出:', [fmt(v) for v in intexp])
    dtot = series('FIN', addr['fin_dtot1_row'])
    check('④ 债务余额>=0', all(isinstance(v, (int, float)) and v >= -0.01 for v in dtot))

    rmax = series('FIN', addr['fin_resid_max_row'])
    print('  残差上限:', [f'{v:.6f}' if isinstance(v, (int, float)) else str(v) for v in rmax])
    resid_max = max((v for v in rmax if isinstance(v, (int, float))), default=0)
    check('⑤ FIN迭代残差<0.01', all(isinstance(v, (int, float)) and v < 0.01 for v in rmax),
          f'max={resid_max:.2e}')
    cfc = series('CF', addr['cf_cash1_row'])
    # B1: 先算maxdiff变量再进f-string，保持Python 3.10+兼容。
    cf_pairs = [(a, b) for a, b in zip(cfc, cash)
                if isinstance(a, (int, float)) and isinstance(b, (int, float))]
    cf_maxdiff = max((abs(a - b) for a, b in cf_pairs), default=0)
    check('⑥ CF期末现金=BS货币资金',
          len(cf_pairs) == len(cfc) and all(abs(a - b) < 0.01 for a, b in cf_pairs),
          f'maxdiff={cf_maxdiff:.2e}')

    print('\n[WACC]')
    wcalc = V('Assumptions', addr['wacc_calc'].split('!')[1])
    wused = V('Assumptions', addr['wacc_used'].split('!')[1])
    ke = V('Assumptions', addr['ke'].split('!')[1])
    bu = V('Assumptions', addr['beta_u'].split('!')[1])
    bl = V('Assumptions', addr['beta_l'].split('!')[1])
    wd = V('Assumptions', addr['wd'].split('!')[1])
    print(f"  βu中位={fmtv(bu, '.4f')}  βl(relever)={fmtv(bl, '.4f')}  Wd={fmtv(wd, '.4%')}  Ke={fmtv(ke, '.4%')}")
    print(f"  WACC计算值={fmtv(wcalc, '.4%')}  采用值={fmtv(wused, '.4%')}")
    w = numguard('采用值=计算值(override留空)', wacc_calc=wcalc, wacc_used=wused)
    if w:
        check('采用值=计算值(override留空)', abs(w['wacc_calc'] - w['wacc_used']) < 1e-9)
    term_ref = (addr.get('term_g') or addr.get('terminal_g') or
                (addr.get('named_ranges') or {}).get('TERM_G'))
    terminal_g = _value_at_ref(WB, term_ref)
    record(_wacc_terminal_control(wused, terminal_g, min_spread=0.005))

    print('\n[DCF]')
    ev = V('DCF', addr['dcf_ev'].split('!')[1]); eq = V('DCF', addr['dcf_eq'].split('!')[1])
    ps_raw = V('DCF', addr['dcf_ps'].split('!')[1]); up = V('DCF', addr['dcf_upside'].split('!')[1])
    print(f"  EV={fmt(ev)}  股权价值={fmt(eq)}  每股价值={fmtv(ps_raw, '.4f')}{per_share_unit}  隐含涨跌幅={fmtv(up, '+.1%')}")
    ps = _num(ps_raw)
    if ps is None:
        check('DCF每股为有效数值', False, f'实际值={ps_raw!r}')
    record(_golden_control(
        meta.get('code'), ps_raw, meta,
        repo_root=os.path.dirname(os.path.abspath(__file__)),
        accept_config_change=args.accept_golden_config_change,
    ))
    # E3: t序所在行从 addr['dcf_tidx'] (形如 "DCF!C11") 解析, 不再硬编码
    tidx_row = 11
    tidx_addr = addr.get('dcf_tidx')
    if tidx_addr:
        m = re.search(r'(\d+)\s*$', str(tidx_addr).split('!')[-1].replace('$', ''))
        if m:
            tidx_row = int(m.group(1))
        else:
            print(f'[WARN] dcf_tidx 无法解析行号({tidx_addr!r}), 按行11回退')
    else:
        print('[WARN] addr 缺 dcf_tidx, 按行11回退')
    tidx = [V('DCF', f'{COLS[y]}{tidx_row}') for y in FCST]
    check('年中折现t=0.5..4.5', all(isinstance(v, (int, float)) and abs(v - (0.5 + i)) < 1e-9
                                  for i, v in enumerate(tidx)), str(tidx))

    if addr.get('fcfe_ps'):
        print('\n[FCFE 双视图]')
        fps = V('FCFE', addr['fcfe_ps'].split('!')[1])
        feq = V('FCFE', addr['fcfe_eq'].split('!')[1])
        fdiff = V('FCFE', addr['fcfe_diff'].split('!')[1])
        print(f"  FCFE股权价值={fmt(feq)}  FCFE每股={fmtv(fps, '.4f')}{per_share_unit}  "
              f"vs FCFF {fmtv(ps_raw, '.4f')}{per_share_unit} ({fmtv(fdiff, '+.1%')})")
        g = numguard('FCFE每股>0', fcfe_ps=fps)
        if g:
            check('FCFE每股>0', g['fcfe_ps'] > 0,
                  '{0:.2f}{1}'.format(g['fcfe_ps'], per_share_unit))
        fcfe_waiver = (meta.get('fcfe_divergence_waiver') or
                       addr.get('fcfe_divergence_waiver'))
        record(_fcfe_divergence_control(
            fdiff, waived=args.waive_fcfe_divergence,
            waiver_reason=fcfe_waiver,
        ))
    if addr.get('named_ranges'):
        print(f"  named ranges: {len(addr['named_ranges'])}个 ({', '.join(list(addr['named_ranges'])[:6])}...)")

    print('\n[三情景]')
    scenario_ps = {}
    scenario_rel = {}
    for k in ['bear', 'base', 'bull']:
        scenario_ps[k] = _value_at_ref(WB, addr.get(f'scen_{k}_ps'))
        scenario_rel[k] = _value_at_ref(WB, addr.get(f'scen_{k}_rel'))
        print(f"  {k}: DCF每股={fmtv(scenario_ps[k], '.2f')}{per_share_unit}  "
              f"相对估值={fmtv(scenario_rel[k], '.2f')}{per_share_unit}")
    scenario_engine = meta.get('scenario_engine')
    record(_scenario_control(
        scenario_ps['bear'], scenario_ps['base'], scenario_ps['bull'], scenario_engine,
    ))
    rel_scenario_result = _scenario_control(
        scenario_rel['bear'], scenario_rel['base'], scenario_rel['bull'], scenario_engine,
    )
    rel_scenario_result['name'] = '相对估值三情景单调性/同引擎'
    record(rel_scenario_result)

    main_scenario_ref = addr.get('scen_main_ps') or addr.get('scen_base_ps')
    sb = numguard('主模型桥接情景DCF=主模型DCF',
                  scen_main=_value_at_ref(WB, main_scenario_ref), dcf_ps=ps_raw)
    if sb:
        check('主模型桥接情景DCF=主模型DCF', abs(sb['scen_main'] - sb['dcf_ps']) < 0.01)

    lo = V('Relative_Val', addr['rel_lo'].split('!')[1]); hi = V('Relative_Val', addr['rel_hi'].split('!')[1])
    median_ref = _relative_median_ref(addr)
    medpe = _value_at_ref(WB, median_ref)
    pricing_refs = addr.get('relative_price_rows')
    pricing_values = ([_value_at_ref(WB, ref) for ref in pricing_refs]
                      if pricing_refs is not None else [])
    record(_relative_median_control(medpe, pricing_values))
    if 'relative_f1_rows' in addr:
        f1_refs = addr.get('relative_f1_rows') or []
        f1_values = [_value_at_ref(WB, ref) for ref in f1_refs]
        f1_median = _value_at_ref(WB, addr.get('rel_f1_median_cell'))
        record(_relative_median_control(
            f1_median, f1_values, name='相对估值FY2中位数',
        ))
    print(f"\n相对估值区间=[{fmtv(lo, '.2f')}, {fmtv(hi, '.2f')}]{per_share_unit}  "
          f"可比中位PE={medpe if not isinstance(medpe, float) else round(medpe, 1)}")

    gm_row = addr.get('is_gm_row')
    if gm_row is None:
        for row in range(1, WB['IS'].max_row + 1):
            label = WB['IS'].cell(row=row, column=1).value
            if isinstance(label, str) and '毛利率' in label:
                gm_row = row
                break
    if gm_row is None:
        record(_control_result('毛利率硬边界', 'FAIL', 'IS页未找到毛利率行'))
    else:
        record(_gross_margin_control(series('IS', int(gm_row), YRS)))

    if 'historical_plugs' in meta:
        record(_historical_plug_control(meta.get('historical_plugs')))
    else:
        record(_control_result(
            '历史配平plug重大性', 'NOT_APPLICABLE', 'legacy addr未提供historical_plugs元数据',
        ))

    sens = V('Sensitivity', addr['sens_center'].split('!')[1])
    sg = numguard('敏感矩阵中心=DCF', sens_center=sens, dcf_ps=ps_raw)
    if sg:
        check('敏感矩阵中心=DCF', abs(sg['sens_center'] - sg['dcf_ps']) < 0.01,
              '{0:.2f} vs {1:.2f}'.format(sg['sens_center'], sg['dcf_ps']))
    dpo_c = V('Sensitivity', addr['dpo_center'].split('!')[1])
    dg = numguard('DPO矩阵δ=0=DCF基准', dpo_center=dpo_c, dcf_ps=ps_raw)
    if dg:
        check('DPO矩阵δ=0=DCF基准', abs(dg['dpo_center'] - dg['dcf_ps']) < 0.01,
              '{0:.2f}'.format(dg['dpo_center']))
    env = V('Summary', addr['summary_env'].split('!')[1]); env_hi = V('Summary', addr['summary_env_hi'].split('!')[1])
    print(f"Summary综合参考区间=[{fmtv(env, '.2f')}, {fmtv(env_hi, '.2f')}]{per_share_unit}")

    print('\n[主模型输出]')
    for y in FCST[:3]:
        rv = V('IS', addr['is_rev'][str(y)].split('!')[1])
        npv = V('IS', addr['is_np'][str(y)].split('!')[1])
        print(f'  {y}E: 营收={fmt(rv)}  归母={fmt(npv)}')

    payload = _verification_payload(recalced, meta, controls)
    print('\n' + '=' * 70)
    print('总体:', payload['label'], f"(verdict={payload['verdict']}, exit={payload['exit_code']})")
    print('=' * 70)
    if args.json_summary:
        rendered = json.dumps(payload, ensure_ascii=False, indent=2)
        if args.json_summary == '-':
            print(rendered)
        else:
            summary_dir = os.path.dirname(os.path.abspath(args.json_summary))
            os.makedirs(summary_dir, exist_ok=True)
            with open(args.json_summary, 'w', encoding='utf-8') as fh:
                fh.write(rendered + '\n')
            print(f'[INFO] JSON摘要: {args.json_summary}')
    sys.exit(payload['exit_code'])


if __name__ == '__main__':
    main()
