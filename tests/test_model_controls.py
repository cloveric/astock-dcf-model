import copy
import hashlib
import math
import re
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

import build_model as bm


ROOT = Path(__file__).resolve().parents[1]


def _raw(code="300476"):
    with (ROOT / "configs" / f"{code}.yaml").open(encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def _build(tmp_path, raw, name, research=None):
    out = tmp_path / f"{name}.xlsx"
    addr = tmp_path / f"{name}.addr.json"
    result = bm.build(bm.Cfg(raw), str(out), str(addr), research=research)
    return load_workbook(out, data_only=False), result


def _find_row(ws, label):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            return row
    raise AssertionError(f"missing row label: {label}")


def _cell_from_ref(wb, ref):
    sheet, coord = ref.replace("$", "").split("!", 1)
    return wb[sheet][coord]


def _mark_reconciled(comp, valuation_date):
    comp.pop("ref_only", None)
    comp.update({
        "earnings_verified": True,
        "earnings_basis": "FY1",
        "source_url": "https://example.test/forecast",
        "source_as_of": valuation_date,
        "source_forward_pe": comp["mcap"] / comp["np_f0"],
        "src": "外部一致预期，已与来源 forward PE 对账",
    })


def _raw_with_one_reconciled_comp():
    raw = _raw("688825")
    for comp in raw["relative_val"]["comps"]:
        comp["earnings_verified"] = False
        comp["ref_only"] = True
    comp = raw["relative_val"]["comps"][0]
    _mark_reconciled(comp, raw["model"]["valuation_date"])
    return raw, comp


def test_price_comparables_without_beta_still_drive_true_median(tmp_path):
    raw = _raw()
    raw["wacc"]["beta_unlevered_input"] = 0.9
    raw["wacc"]["beta_unlevered_basis"] = "无beta可比时采用分析师输入"
    for comp in raw["relative_val"]["comps"]:
        comp.pop("beta_l", None)
        comp.pop("d_e", None)
        comp.pop("tax", None)
        comp["earnings_verified"] = False
        comp["ref_only"] = True
    for comp in raw["relative_val"]["comps"][:2]:
        _mark_reconciled(comp, raw["model"]["valuation_date"])

    wb, addr = _build(tmp_path, raw, "no_beta")

    assert addr["rel_med_pe"] == addr["rel_median_cell"]
    assert addr["relative_price_rows"] == ["Relative_Val!F5", "Relative_Val!F6"]
    median_formula = _cell_from_ref(wb, addr["rel_median_cell"]).value
    assert median_formula == "=MEDIAN(F5:F6)"
    scenario_formula = _cell_from_ref(wb, addr["scen_base_rel"]).value
    assert addr["rel_median_cell"].replace("!", "!$").replace("F", "$F$") not in scenario_formula
    assert "Relative_Val!$F$10" in scenario_formula


def test_all_three_scenarios_use_same_engine_and_main_model_has_separate_bridge(tmp_path):
    wb, addr = _build(tmp_path, _raw(), "scenario_engine")

    scenario_formulas = [_cell_from_ref(wb, addr[key]).value for key in (
        "scen_bear_ps", "scen_base_ps", "scen_bull_ps"
    )]
    assert all(formula.startswith("=(") for formula in scenario_formulas)
    assert all("DCF!$D$" in formula for formula in scenario_formulas)
    assert _cell_from_ref(wb, addr["scen_main_ps"]).value.startswith("=DCF!")
    bridge = _cell_from_ref(wb, addr["scen_full_simplified_ratio"]).value
    assert addr["scen_main_ps"].split("!", 1)[1] in bridge
    assert addr["scen_base_ps"].split("!", 1)[1] in bridge
    assert addr["meta"]["scenario_engine"] == "simplified_same_engine_v1"

    checks = wb["Checks"]
    monotonic_row = _find_row(checks, "13. 三情景同引擎估值单调性")
    assert checks.cell(monotonic_row, 2).value.startswith("=AND(")


def test_usd_model_uses_usd_per_share_labels_everywhere(tmp_path):
    raw = _raw()
    raw["company"]["unit"] = "美元百万元"
    raw["company"]["currency_note"] = "美元 / 百万元 (每股数据为美元)"
    wb, addr = _build(tmp_path, raw, "usd_labels")

    assert addr["meta"]["currency_code"] == "USD"
    assert addr["meta"]["per_share_unit"] == "美元/股"
    text_values = [
        cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and not cell.value.startswith("=")
    ]
    assert any("美元/股" in value for value in text_values)
    assert not any(re.search(r"(?<!美)(?<!港)元/股", value) for value in text_values)
    assert not any(re.search(r"\(元\)|低\(元\)|高\(元\)", value) for value in text_values)


def test_auto_consensus_is_labeled_and_excluded_from_summary_envelope(tmp_path):
    # 仓库配置中已完成结构化对账的FY1可比会合法进入正式计价;
    # 本测试守卫的是"未验证可比不得进正式计价", 故显式将其降级为未验证再断言
    raw = _raw("688825")
    for cp in raw.get("relative_val", {}).get("comps", []):
        cp["earnings_verified"] = False
    wb, addr = _build(tmp_path, raw, "auto_consensus")

    assumptions = wb["Assumptions"]
    labels = [assumptions.cell(row, 1).value for row in range(1, assumptions.max_row + 1)]
    assert any(isinstance(label, str) and label.startswith("模型自动外推2026E营收") for label in labels)

    summary = wb["Summary"]
    auto_row = next(
        row for row in range(1, summary.max_row + 1)
        if isinstance(summary.cell(row, 1).value, str)
        and summary.cell(row, 1).value.startswith("相对估值 — 模型自动外推")
    )
    envelope_row = _find_row(summary, "综合参考区间(经验证方法包络)")
    assert summary.cell(envelope_row, 7).alignment.wrap_text is True
    low_formula = summary.cell(envelope_row, 3).value.replace("$", "")
    high_formula = summary.cell(envelope_row, 4).value.replace("$", "")
    assert f"C{auto_row}" not in low_formula
    assert f"D{auto_row}" not in high_formula
    assert addr["relative_price_rows"] == []
    assert addr["rel_med_pe"] != addr["rel_median_cell"]
    model_row = _find_row(summary, "相对估值 — 模型2026E归母")
    assert summary.cell(model_row, 2).value == "PE 5.7x 单点（无正式可比）"
    assert summary.column_dimensions["B"].width >= 30



def test_reconciled_fy1_comp_enters_formal_pricing(tmp_path):
    """Only a source-backed FY1 comp that reconciles to source PE may price."""
    raw, _ = _raw_with_one_reconciled_comp()
    wb, addr = _build(tmp_path, raw, "verified_fy1")
    assert addr["relative_price_rows"] == ["Relative_Val!F5"]
    assert addr["rel_med_pe"] == addr["rel_median_cell"]
    model_row = _find_row(wb["Summary"], "相对估值 — 模型2026E归母")
    assert wb["Summary"].cell(model_row, 2).value == "PE 5.7x ~ 可比中位"


def test_verified_flag_without_structured_provenance_is_not_a_pricing_gate():
    """A self-attested boolean must not turn arbitrary numbers into verified data."""
    comp = {
        "mcap": 100.0,
        "np_f0": 10.0,
        "earnings_verified": True,
        "earnings_basis": "FY1",
        "src": "reviewed",
    }
    assert bm.forward_earnings_verified(comp) is False


def test_legacy_source_text_cannot_bypass_structured_provenance():
    comp = {
        "mcap": 100.0,
        "np_f0": 0.01,
        "src": "券商一致预期",
    }
    assert bm.forward_earnings_verified(comp) is False


@pytest.mark.parametrize(
    ("mutate", "message"),
    [
        (lambda c: c.pop("source_url"), "source_url"),
        (lambda c: c.update(source_url="http://127.0.0.1/forecast"), "source_url"),
        (lambda c: c.update(source_url="https://user:secret@example.com/forecast"), "source_url"),
        (lambda c: c.update(source_as_of="2026-08-19"), "source_as_of"),
        (lambda c: c.update(source_as_of="20260818"), "source_as_of"),
        (lambda c: c.update(source_as_of="2025-01-01"), "source_as_of"),
        (lambda c: c.update(source_forward_pe=c["source_forward_pe"] * 2), "source_forward_pe"),
        (lambda c: c.update(np_f0=0.0), "np_f0"),
        (lambda c: c.update(src="  "), "src"),
    ],
)
def test_verified_comparable_requires_reconciled_structured_provenance(mutate, message):
    """Malformed or unreconciled peer data must fail before workbook generation."""
    raw, comp = _raw_with_one_reconciled_comp()
    mutate(comp)
    with pytest.raises(ValueError, match=rf"relative_val\.comps\[0\].*{message}"):
        bm.validate_config(raw)


def test_comparable_collection_rejects_non_list_even_when_empty():
    raw = _raw()
    raw["relative_val"]["comps"] = {}
    with pytest.raises(ValueError, match="relative_val.comps must be a list"):
        bm.validate_config(raw)


@pytest.mark.parametrize(("field", "value"), [("np_f0", 0.0), ("np_f1", 0.0), ("np_f1", -1.0)])
def test_comparable_earnings_used_in_pe_formulas_must_be_positive(field, value):
    raw = _raw()
    raw["relative_val"]["comps"][0][field] = value
    with pytest.raises(ValueError, match=rf"relative_val\.comps\[0\]\.{field} must be positive"):
        bm.validate_config(raw)


def test_missing_fy2_comparable_is_shown_as_unavailable(tmp_path):
    """A paywalled FY2 estimate must not be duplicated and shown as zero growth."""
    raw = _raw("688825")
    for comp in raw["relative_val"]["comps"]:
        comp["earnings_verified"] = False
        comp["ref_only"] = True
        comp["np_f1"] = None

    wb, _ = _build(tmp_path, raw, "missing_fy2")

    for row in range(5, 5 + len(raw["relative_val"]["comps"])):
        assert [wb["Relative_Val"].cell(row, col).value for col in (7, 8, 9, 10)] == [
            "—", "—", "—", "—",
        ]


def test_flat_or_negative_fy2_growth_does_not_create_divide_by_zero_peg(tmp_path):
    raw, comp = _raw_with_one_reconciled_comp()
    comp["np_f1"] = comp["np_f0"]

    wb, _ = _build(tmp_path, raw, "flat_fy2_growth")

    assert wb["Relative_Val"]["I5"].value == "=G5/E5-1"
    assert wb["Relative_Val"]["J5"].value == "—"
    median_row = _find_row(wb["Relative_Val"], "可比中位数(计价1家/β0家)")
    assert wb["Relative_Val"].cell(median_row, 10).value == "—"


def test_688825_only_reconciled_comparables_enter_formal_pricing(tmp_path):
    """The disclosed outlier remains visible but cannot move the pricing median."""
    raw = _raw("688825")
    wb, addr = _build(tmp_path, raw, "688825_reconciled_comps")

    assert addr["relative_price_rows"] == [
        "Relative_Val!F5",
        "Relative_Val!F6",
        "Relative_Val!F7",
        "Relative_Val!F8",
        "Relative_Val!F10",
    ]
    assert _cell_from_ref(wb, addr["rel_median_cell"]).value == "=MEDIAN(F5,F6,F7,F8,F10)"
    assert "排除" in wb["Relative_Val"]["K9"].value


def test_missing_historical_da_is_shown_as_unavailable(tmp_path):
    """Missing source D&A must remain visibly unavailable while forecasts still calculate."""
    raw = _raw("601138")
    raw["hist"]["cf"]["da"][-1] = None

    wb, _ = _build(tmp_path, raw, "missing_historical_da")

    row = _find_row(wb["CF"], "加: 折旧与摊销(D&A)")
    assert wb["CF"].cell(row, 4).value == "—"
    assert isinstance(wb["CF"].cell(row, 5).value, str)
    assert wb["CF"].cell(row, 5).value.startswith("=")


@pytest.mark.parametrize("invalid", [None, "not-a-vector"])
def test_historical_da_must_be_a_year_aligned_vector(invalid):
    raw = _raw("601138")
    raw["hist"]["cf"]["da"] = invalid
    with pytest.raises(ValueError, match="hist.cf.da must be a vector"):
        bm.validate_config(raw)


def test_addr_metadata_never_publishes_external_absolute_config_path(tmp_path):
    raw = _raw()
    cfg = bm.Cfg(raw, source_path=str(tmp_path / "private-config.yaml"))
    out = tmp_path / "model.xlsx"
    addr_path = tmp_path / "model.addr.json"

    addr = bm.build(cfg, str(out), str(addr_path))

    assert addr["meta"]["config_path"] == "private-config.yaml"
    assert addr["meta"]["config_path_scope"] == "external"
    assert not Path(addr["meta"]["config_path"]).is_absolute()


def test_partial_external_consensus_only_verifies_explicit_series_and_year(tmp_path):
    raw = _raw("688825")
    research = {
        "notes": [],
        "consensus": {
            "rev": {2026: 123_456.0},
            "np": {},
            "target_price": None,
            "source": "pytest external consensus",
            "date": "2026-08-18",
            "path": "pytest.json",
        },
    }

    wb, addr = _build(tmp_path, raw, "partial_external_consensus", research=research)

    assumptions = wb["Assumptions"]
    labels = [assumptions.cell(row, 1).value for row in range(1, assumptions.max_row + 1)]
    assert "一致预期2026E营收" in labels
    assert "模型自动外推2026E归母净利" in labels
    assert addr["meta"]["consensus_np_verified"] is False
    assert addr["meta"]["consensus_verified_by_year"]["np"]["2026"] is False


def test_pending_consensus_basis_is_not_treated_as_verified(tmp_path):
    raw = _raw("688825")
    raw["consensus"]["np"]["basis"] = "一致预期待核，占位数据"

    wb, addr = _build(tmp_path, raw, "pending_consensus")

    assert addr["meta"]["consensus_np_verified"] is False
    labels = [
        wb["Assumptions"].cell(row, 1).value
        for row in range(1, wb["Assumptions"].max_row + 1)
    ]
    assert "模型自动外推2026E归母净利" in labels


def test_fcfe_divergence_has_visible_control_status_in_detail_checks_and_summary(tmp_path):
    wb, addr = _build(tmp_path, _raw(), "fcfe_control")

    for sheet_name in ("FCFE", "Checks", "Summary"):
        ws = wb[sheet_name]
        row = next(
            row for row in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row, 1).value, str)
            and "FCFE/FCFF差异控制" in ws.cell(row, 1).value
        )
        formulas = [cell.value for cell in ws[row] if isinstance(cell.value, str) and cell.value.startswith("=")]
        assert any("IF(ABS(" in formula and "0.3" in formula for formula in formulas)
    assert "fcfe_control" in addr


def test_fcfe_divergence_waiver_requires_reason_and_is_visible(tmp_path):
    raw = _raw()
    raw.setdefault("checks", {})["fcfe_divergence_waiver"] = "董事会扩产期动态杠杆造成口径差异, 已复核"
    wb, addr = _build(tmp_path, raw, "fcfe_waiver")

    assert addr["meta"]["fcfe_divergence_waiver"] == "董事会扩产期动态杠杆造成口径差异, 已复核"
    for sheet_name in ("FCFE", "Checks", "Summary"):
        formulas = [
            cell.value for row in wb[sheet_name].iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        assert any("WAIVED" in formula and "董事会扩产期动态杠杆" in formula for formula in formulas)


def test_interim_lineage_error_is_visible_without_being_conflated_with_unavailable(tmp_path):
    raw = _raw()
    raw["data_lineage"] = {
        "interim": {"status": "error", "detail": "HKF10 timeout after retries"}
    }
    wb, addr = _build(tmp_path, raw, "interim_error")

    cover_text = [
        cell.value for row in wb["Cover"].iter_rows() for cell in row
        if isinstance(cell.value, str)
    ]
    checks_text = [
        cell.value for row in wb["Checks"].iter_rows() for cell in row
        if isinstance(cell.value, str)
    ]
    assert any("中报抓取失败" in value and "timeout" in value for value in cover_text)
    assert any("中报抓取失败" in value and "timeout" in value for value in checks_text)
    assert "ERROR" in checks_text
    assert addr["meta"]["interim_status"] == "error"


@pytest.mark.parametrize(
    ("mutate", "message"),
    [
        (lambda r: r["opex"]["tax_rate"].update(values=[0.15, 0.15]), "opex.tax_rate"),
        (lambda r: r["model"].update(fcst_years=[2026, 2028, 2029]), "model.fcst_years"),
        (lambda r: r["market"]["price"].update(value=0), "market.price"),
        (lambda r: r["market"]["shares"].update(value=-1), "market.shares"),
        (lambda r: r["segments"][0]["gm"].update(values=[1.01] * 5), "segments[0].gm"),
        (lambda r: r["opex"]["tax_rate"].update(values=[0.8] * 5), "opex.tax_rate"),
        (lambda r: r["financing"]["rate_st"].update(value=-0.01), "financing.rate_st"),
        (lambda r: r["scenarios"]["bear"].update(rev_adj=0.2), "scenarios.*.rev_adj"),
        (lambda r: r["wacc"].update(override=0.032), "wacc.override"),
        (lambda r: r["wacc"]["erp"].update(basis="  "), "wacc.erp.basis"),
        (lambda r: r["market"].update(pe_ttm=math.nan), "finite"),
        (lambda r: r.setdefault("checks", {}).update(fcfe_divergence_waiver="  "), "checks.fcfe_divergence_waiver"),
    ],
)
def test_strong_config_validation_rejects_invalid_controls(mutate, message):
    raw = _raw()
    mutate(raw)
    with pytest.raises(ValueError, match=re.escape(message)):
        bm.validate_config(raw)


@pytest.mark.parametrize("code", ["300476", "002463", "688825", "00981", "601138"])
def test_strong_config_validation_keeps_current_valid_configs(code):
    bm.validate_config(_raw(code))


def test_yaml_loader_rejects_duplicate_keys(tmp_path):
    path = tmp_path / "duplicate.yaml"
    path.write_text("market:\n  price: 10\n  price: 20\n", encoding="utf-8")
    with pytest.raises(ValueError, match=r"duplicate YAML key.*market\.price"):
        bm.load_yaml_strict(path)


def test_file_backed_config_uses_raw_file_fingerprint():
    path = ROOT / "configs" / "300476.yaml"
    cfg = bm.load_config(code="300476", allow_fallback=False)
    assert cfg.config_sha256 == hashlib.sha256(path.read_bytes()).hexdigest()
    assert cfg.source_path == str(path)
    assert cfg.is_default_config is True


def test_historical_balance_plugs_allow_rounding_only_and_report_correction():
    raw = _raw()
    normalized, plugs = bm.prepare_historical_bs(raw)

    assert [plug["year"] for plug in plugs] == [2024, 2025]
    assert all(abs(plug["diff"]) <= plug["threshold"] for plug in plugs)
    assets = bm.historical_balance_totals(normalized)
    assert all(abs(item["diff"]) < 1e-9 for item in assets)


def test_historical_balance_plugs_reject_material_difference():
    raw = _raw()
    raw["hist"]["bs"]["cr"][0] -= 100.0
    with pytest.raises(ValueError, match="material balance-sheet difference"):
        bm.prepare_historical_bs(raw)


def test_historical_balance_plugs_reject_material_negative_residual_bucket():
    raw = _raw()
    raw["hist"]["bs"]["oncl"][0] = -10.0
    with pytest.raises(ValueError, match=r"hist.bs.oncl\[2023\]"):
        bm.prepare_historical_bs(raw)


def test_historical_balance_plugs_reject_material_negative_other_equity():
    raw = _raw()
    raw["hist"]["bs"]["oeq"][0] = -10.0
    with pytest.raises(ValueError, match=r"hist.bs.oeq\[2023\]"):
        bm.prepare_historical_bs(raw)
