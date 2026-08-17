# -*- coding: utf-8 -*-
"""离线冒烟校验 (CI用, 不依赖LibreOffice/外网):
   python tests/smoke_check.py <model.xlsx> [model.addr.json]
校验: 17张基础工作表齐全 / named ranges≥10 / Checks页含12项校验 /
     FCFE页关键行存在 / addr.json关键地址可解析。
"""
import json
import os
import sys

from openpyxl import load_workbook

BASE_SHEETS = ['Cover', 'Summary', 'Assumptions', 'Revenue_Segments', 'IS', 'BS', 'CF',
               'Schedules', 'PPE', 'FIN', 'Equity_Roll', 'DCF', 'FCFE', 'Relative_Val',
               'Sensitivity', 'Scenarios', 'Checks']


def main():
    xlsx = sys.argv[1]
    addr_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(xlsx)[0] + '.addr.json'
    wb = load_workbook(xlsx)
    fails = []

    missing = [s for s in BASE_SHEETS if s not in wb.sheetnames]
    if missing:
        fails.append(f'缺少工作表: {missing}')

    n_names = len(wb.defined_names)
    if n_names < 10:
        fails.append(f'named ranges不足10个: {n_names}')

    chk_labels = [wb['Checks'].cell(row=r, column=1).value or '' for r in range(1, 40)]
    for i in range(1, 13):
        if not any(str(l).startswith(f'{i}.') for l in chk_labels):
            fails.append(f'Checks缺第{i}项校验行')

    fcfe_labels = [wb['FCFE'].cell(row=r, column=1).value or '' for r in range(1, 30)]
    for need in ('FCFE', '加: 净新增借款', '两法差异原因'):
        if not any(need in str(l) for l in fcfe_labels):
            fails.append(f'FCFE页缺少行: {need}')

    if os.path.exists(addr_path):
        addr = json.load(open(addr_path, encoding='utf-8'))
        for k in ('dcf_ps', 'fcfe_ps', 'checks_bool', 'named_ranges'):
            if k not in addr:
                fails.append(f'addr.json缺键: {k}')
        # E6: 已有失败(含缺键)时不再继续索引 addr['dcf_ps'], 直接汇总退出, 防KeyError掩盖真实失败清单
        if fails:
            print('SMOKE FAIL:')
            for f in fails:
                print(' -', f)
            sys.exit(1)
        dcf_sheet, dcf_cell = addr['dcf_ps'].split('!')
        v = wb[dcf_sheet][dcf_cell].value
        if not (isinstance(v, str) and v.startswith('=')):
            fails.append(f'DCF每股单元格非公式: {v!r}')
    else:
        fails.append(f'addr.json不存在: {addr_path}')

    if fails:
        print('SMOKE FAIL:')
        for f in fails:
            print(' -', f)
        sys.exit(1)
    print(f'SMOKE PASS: {len(BASE_SHEETS)}表齐全, {n_names}个named ranges, Checks 12项, addr完整')


if __name__ == '__main__':
    main()
