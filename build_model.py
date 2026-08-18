# -*- coding: utf-8 -*-
"""
A股机构级估值模型生成器 (通用版)

用法:
    python build_model.py --code 300476                 # 读 configs/300476.yaml, 生成 out/*.xlsx
    python build_model.py --config configs/300476.yaml  # 指定配置文件
    python build_model.py --code 002463 --out examples/002463.xlsx

研究层 (全部默认关闭, 不传任何研究参数时输出与基础模型完全一致):
    --dr <档案.md>        导入深度研究档案: 量化结论回填Assumptions依据列(标注dr档案§章节),
                          新增 DR研究(全文存档) 与 研究摘要(假设×采用值×来源×置信度) 两页
    --consensus <json/csv> 聚源/gildata一致预期文件(营收/归母/目标价), 写入研究摘要并
                          用于Relative_Val与Checks的一致性对照行
    --announcements       东财业绩预告/业绩快报(系统curl)抓最新一期要点进研究摘要
    --llm auto|claude|codex|off  本机存在对应CLI时生成研究备忘录写入研究摘要页眉 (默认off)

配置驱动: 个股差异全部收敛到 YAML 配置 (公司信息/分部业务量价假设/费用率/税率/营运资本/
资本开支/股利/债务融资/可比公司/WACC/三情景), 每个假设均带"依据"文字字段 (硬性要求)。
配置缺失或缺 segments 时自动走兜底模式: 经 fetch_data 用东财F10(系统curl)拉取三表+
主营构成, 自动推导单段/少段收入驱动与费率默认值, 模型照样跑通 (精度局限见README)。

模型结构 (17 sheets, 研究层激活时追加 DR研究/研究摘要):
    Cover / Summary / Assumptions / Revenue_Segments / IS / BS / CF / Schedules /
    PPE / FIN / Equity_Roll / DCF / FCFE / Relative_Val / Sensitivity / Scenarios / Checks
特性: FIN页revolver+现金sweep+理财承接, 利息=平均余额×利率的循环链在FIN页表内
迭代展开4轮(全簿无循环引用, LibreOffice headless可直接重算); WACC可比unlever/relever;
DCF年中折现; FCFE股权现金流双视图; 关键驱动named range审计轨迹; Checks页自动校验。

单位: 人民币百万元(除标注外); 输入蓝/公式黑/外链绿/警告红。
"""
import argparse
import datetime
import hashlib
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from model_labels import (
    consensus_label,
    consensus_year_is_verified,
    derive_currency_labels,
    forward_earnings_verified,
)
from model_validation import (
    historical_balance_totals,
    load_yaml_strict,
    prepare_historical_bs,
    validate_config,
)

# ============================================================
# 配置加载与规范化
# ============================================================


def _public_source_reference(path):
    """Return a portable, non-sensitive config reference and its scope."""
    if not path:
        return None, None
    source = Path(path).resolve()
    project_root = Path(__file__).resolve().parent
    try:
        return source.relative_to(project_root).as_posix(), 'project'
    except ValueError:
        return source.name, 'external'

def _norm_entry(e):
    """假设条目规范化 → {'values': [...] 或标量, 'basis': str}
    支持写法: 标量 / 列表 / {value: x, basis: s} / {values: [...], basis: s}
    列表元素可为 {value, basis} 以覆盖单点依据 (取第一个非空basis为行依据)。"""
    if isinstance(e, dict) and ('value' in e or 'values' in e):
        basis = e.get('basis', '')
        vals = e['values'] if 'values' in e else e['value']
    else:
        basis, vals = '', e
    if isinstance(vals, (list, tuple)):
        out = []
        for v in vals:
            if isinstance(v, dict):
                out.append(v['value'])
                if not basis and v.get('basis'):
                    basis = v['basis']
            else:
                out.append(v)
        return {'values': out, 'basis': basis}
    return {'values': vals, 'basis': basis}


def _broadcast(vals, n):
    """标量显式广播为n元列表; 向量必须精确匹配, 禁止静默截断/补齐。"""
    if isinstance(vals, (list, tuple)):
        vals = list(vals)
        if len(vals) != n:
            raise ValueError(f'配置错误: assumption vector length must be exactly {n} (got {len(vals)}); use a scalar for explicit broadcasting')
        return vals
    return [vals] * n


class Cfg:
    """配置访问封装: e(path) 取假设条目, v(path) 取值, s(path) 取字符串"""

    def __init__(self, raw, source_path=None, is_default_config=False, config_sha256=None):
        self.raw = raw
        self.source_path = os.path.abspath(source_path) if source_path else None
        self.is_default_config = bool(is_default_config)
        canonical = json.dumps(raw, ensure_ascii=False, sort_keys=True, separators=(',', ':'), allow_nan=False)
        self.config_sha256 = config_sha256 or hashlib.sha256(canonical.encode('utf-8')).hexdigest()

    def get(self, path, default=None):
        node = self.raw
        for p in path.split('.'):
            if not isinstance(node, dict) or p not in node:
                return default
            node = node[p]
        return node

    def e(self, path):
        node = self.get(path)
        if node is None:
            raise KeyError(f'配置缺失假设项: {path}')
        return _norm_entry(node)

    def v(self, path, default=None):
        node = self.get(path)
        if node is None:
            return default
        if isinstance(node, dict) and ('value' in node or 'values' in node):
            return _norm_entry(node)['values']
        return node


def _as_date(value, label):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    rendered = str(value)
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', rendered):
        raise ValueError(f'配置错误: {label} 必须是 YYYY-MM-DD 日期')
    try:
        return datetime.date.fromisoformat(rendered)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'配置错误: {label} 必须是 YYYY-MM-DD 日期') from exc


def _enforce_freshness(raw, *, as_of=None, stale_after_days=30,
                       fail_on_stale=False, require_interim=False, source='配置'):
    """Apply look-ahead, staleness and latest-report gates to a loaded config."""
    reference = _as_date(as_of, '--as-of') if as_of else datetime.date.today()
    model = raw.get('model') if isinstance(raw, dict) else None
    valuation_date = (model or {}).get('valuation_date')
    if valuation_date:
        valuation = _as_date(valuation_date, 'model.valuation_date')
        if valuation > reference:
            raise ValueError(
                f'配置错误: model.valuation_date={valuation} 晚于 --as-of {reference}，'
                '拒绝使用未来数据')
        if stale_after_days is not None:
            if isinstance(stale_after_days, bool) or stale_after_days < 0:
                raise ValueError('配置错误: stale_after_days 必须是非负整数或None')
            age = (reference - valuation).days
            if age > stale_after_days:
                message = (f'配置已过期: {source}估值日{valuation}距as-of {reference}已有'
                           f'{age}天，阈值{stale_after_days}天')
                if fail_on_stale:
                    raise ValueError(message)
                print(f'WARNING: {message}', file=sys.stderr)
    elif fail_on_stale:
        raise ValueError('配置已过期: 缺少 model.valuation_date，无法证明数据新鲜度')
    else:
        print('WARNING: 配置缺少 model.valuation_date，无法判断数据新鲜度', file=sys.stderr)

    if require_interim:
        interim = raw.get('interim') if isinstance(raw, dict) else None
        lineage = raw.get('data_lineage') if isinstance(raw, dict) else None
        fetch_state = ((lineage or {}).get('interim') or {}) if isinstance(lineage, dict) else {}
        if (not isinstance(interim, dict) or interim.get('anchor') is not True
                or fetch_state.get('status') == 'error'
                or interim.get('fetch_status') == 'error'):
            detail = fetch_state.get('error') or (interim or {}).get('basis') or '未提供有效中期锚点'
            raise ValueError(f'require-interim 门控失败: {detail}')


def load_config(code=None, config_path=None, allow_fallback=True, *, refresh=False,
                as_of=None, stale_after_days=30, fail_on_stale=False,
                require_interim=False):
    """加载并校验YAML；可显式刷新公开兜底数据并执行新鲜度/中期门控。"""
    is_default_config = config_path is None
    if config_path is None:
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   'configs', f'{code}.yaml')
    if refresh:
        if code is None and os.path.exists(config_path):
            code = str(load_yaml_strict(config_path).get('company', {}).get('code') or '')
        if not code:
            raise ValueError('--refresh 需要 --code，或配置内 company.code')
        print(f'显式刷新: 忽略现有配置数值，在内存中重建{code}公开数据兜底模型')
        from fetch_data import build_fallback_config
        raw = build_fallback_config(code)
        source_hash = None
        is_default_config = False
    elif os.path.exists(config_path):
        with open(config_path, 'rb') as source_file:
            source_hash = hashlib.sha256(source_file.read()).hexdigest()
        raw = load_yaml_strict(config_path)
        config_ref, _ = _public_source_reference(config_path)
        print(f'配置: {config_ref}')
    else:
        if not allow_fallback:
            raise FileNotFoundError(config_path)
        if code is None:
            raise ValueError('配置文件不存在且未提供--code, 无法兜底')
        config_ref, _ = _public_source_reference(config_path)
        print(f'配置文件不存在: {config_ref} → 兜底模式(自动拉取{code}公开数据生成)')
        from fetch_data import build_fallback_config
        raw = build_fallback_config(code)
        source_hash = None
    cfg = Cfg(raw, source_path=config_path if not refresh and os.path.exists(config_path) else None,
              is_default_config=is_default_config, config_sha256=source_hash)
    need_fallback = not cfg.get('segments') or not cfg.get('hist')
    if need_fallback and allow_fallback:
        from fetch_data import apply_fallback
        code = code or cfg.get('company.code')
        apply_fallback(cfg.raw, code)
        print('兜底模式: 已用东财F10主营构成/三表自动补全缺失配置 (精度局限见README)')
        cfg = Cfg(cfg.raw, source_path=cfg.source_path, is_default_config=is_default_config)
    _enforce_freshness(
        cfg.raw, as_of=as_of, stale_after_days=stale_after_days,
        fail_on_stale=fail_on_stale, require_interim=require_interim,
        source=(_public_source_reference(cfg.source_path)[0]
                if cfg.source_path else f'{code or cfg.get("company.code")}公开数据'),
    )
    validate_config(cfg.raw)
    return cfg


# ---------- 样式 ----------
C_BLUE = 'FF0033CC'    # 输入/假设
C_GREEN = 'FF1E7A1E'   # 外部数据/链接
C_BLACK = 'FF000000'   # 公式
C_RED = 'FFCC0000'     # 警告
C_GREY = 'FF7F7F7F'
FONT = 'Microsoft YaHei'
HDR_FILL = PatternFill('solid', fgColor='FF1F3864')
SEC_FILL = PatternFill('solid', fgColor='FFDCE6F1')
TOT_FILL = PatternFill('solid', fgColor='FFF2F2F2')
CHK_FILL = PatternFill('solid', fgColor='FFFFF2CC')
THIN = Side(style='thin', color='FFBFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0.0'
PCT = '0.0%'
MULT = '0.0"x"'
PS = '0.00'
DAYS = '#,##0'


def put(ws, r, c, v, kind='f', fmt=None, bold=False, fill=None, align=None,
        wrap=False, size=10, italic=False):
    # 非公式kind的文本若以'='开头会被当作公式, 转义为全角'＝'
    if kind != 'f' and isinstance(v, str) and v.startswith('='):
        v = '＝' + v[1:]
    cell = ws.cell(row=r, column=c, value=v)
    color = {'in': C_BLUE, 'x': C_GREEN, 'f': C_BLACK, 'w': C_RED,
             't': C_BLACK, 'g': C_GREY}[kind]
    cell.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return cell


def title_bar(ws, text, note, ncols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=13, bold=True, color='FFFFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=note)
    c2.font = Font(name=FONT, size=9, color=C_GREY)
    ws.row_dimensions[2].height = 14


def sec(ws, r, text, ncols=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color='FF1F3864')
    c.fill = SEC_FILL
    ws.row_dimensions[r].height = 16


# ============================================================
# 主构建流程
# ============================================================

def build(cfg, out_path, addr_path, research=None):
    validate_config(cfg.raw)
    co = cfg.raw['company']
    NAME, CODE_FULL = co['name'], co['code_full']
    HIST = [int(y) for y in cfg.get('model.hist_years')]
    FCST = [int(y) for y in cfg.get('model.fcst_years')]
    # C2④: 版面结构性前提显式校验 (YoY引用上年列 / 一致预期取前3个预测年 / 假设页C起NF列等)
    if len(HIST) < 2:
        raise ValueError(f'配置错误: model.hist_years 需≥2年(现{len(HIST)}年) — 收入YoY与期初余额锚点依赖上一年列')
    if not (3 <= len(FCST) <= 5):
        raise ValueError(f'配置错误: model.fcst_years 需3-5年(现{len(FCST)}年) — 一致预期区取前3年、版面按3-5列设计')
    YRS = HIST + FCST
    NH, NF = len(HIST), len(FCST)
    H0, F0, F1 = HIST[-1], FCST[0], FCST[-1]
    VAL_DATE = cfg.get('model.valuation_date', '')
    BUILD_DATE = cfg.get('model.build_date', '')
    UNIT = cfg.get('company.unit', '人民币百万元')        # 港股配置可覆盖为财报币种口径
    _currency = derive_currency_labels(co)
    CURRENCY_CODE = _currency['currency_code']
    PER_SHARE_UNIT = _currency['per_share_unit']
    PS_DENOM = _currency['denomination']
    _normalized_bs, HISTORICAL_PLUGS = prepare_historical_bs(cfg.raw)
    FCFE_DIVERGENCE_WAIVER = str(cfg.get('checks.fcfe_divergence_waiver') or '').strip()
    _interim_lineage = cfg.get('data_lineage.interim') or {}
    INTERIM_STATUS = str(_interim_lineage.get('status') or '') if isinstance(_interim_lineage, dict) else ''
    INTERIM_ERROR = ''
    if INTERIM_STATUS == 'error':
        INTERIM_ERROR = str(_interim_lineage.get('error') or _interim_lineage.get('detail') or '未提供错误详情')
    _fcfe_waiver_excel = FCFE_DIVERGENCE_WAIVER.replace('"', '""')

    def fcfe_control_formula(diff_ref):
        return (f'=IF("{_fcfe_waiver_excel}"<>"","WAIVED: {_fcfe_waiver_excel}",'
                f'IF(ABS({diff_ref})>=0.3,"警告: 差异≥30%, 复核债务调度/资本结构/终值","通过: 差异<30%"))')
    IS_HK = str(CODE_FULL).upper().endswith('.HK')
    YC = {y: get_column_letter(2 + i) for i, y in enumerate(YRS)}   # 报表页 B..
    AC = {y: get_column_letter(3 + i) for i, y in enumerate(FCST)}  # 假设页 C..
    NOTE_COL = 2 + len(YRS)   # C2③: 报表页备注列=年份列后一列 (3历史+5预测→列10, 与既有版面等值)
    AB_COL = 3 + NF           # C2②: Assumptions依据列=C起NF列预测值后一列 (3+5=8即'H', 与既有版面等值)
    _cy = [str(y)[2:] for y in FCST[:3]]   # C1: 一致预期行key年份后缀, 随fcst_years联动 (2026起步时=['26','27','28'])

    SEGS = cfg.get('segments')
    for sg in SEGS:
        sg.setdefault('driver', 'growth' if sg.get('asp') is None else 'vol_asp')
    # Relative_Val 布局常量 (Assumptions的beta引用需先行确定)
    comps = cfg.get('relative_val.comps') or []
    beta_comps = [c for c in comps if c.get('beta_l') is not None]
    other_comps = [c for c in comps if c.get('beta_l') is None]
    comps = beta_comps + other_comps       # 带beta的A股可比在前, 中位数区连续
    RV_COMP_START = 5
    RV_MED_ROW = RV_COMP_START + len(comps)
    N_BETA = len(beta_comps)

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # Assumptions
    # ============================================================
    ws = wb.create_sheet('Assumptions')
    title_bar(ws, f'{NAME}({CODE_FULL}) — Assumptions 驱动总表',
              f'单位: {UNIT}(除标注外) | 蓝色=输入/假设, 绿色=外部数据, 黑色=公式 | 数据日期: {VAL_DATE}')
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 8
    for _i in range(NF):
        ws.column_dimensions[get_column_letter(3 + _i)].width = 11
    ws.column_dimensions[get_column_letter(AB_COL)].width = 95
    ws.freeze_panes = 'C4'

    A = {}   # key -> row

    # dr档案量化结论 → 依据列回填 (研究层, 默认关闭; 仅追加标注, 不改动任何假设值)
    _dr_findings = ((research or {}).get('dr') or {}).get('findings') or []
    if _dr_findings:
        from research.dr_report import annotate as _dr_annotate

        def _ann(label, basis):
            hits = _dr_annotate(_dr_findings, label)
            if not hits:
                return basis
            return (basis + ' | ' if basis else '') + ' | '.join(hits)
    else:
        def _ann(label, basis):
            return basis

    def arow(r, label, unit, vals, basis, kind='in', fmt=PCT):
        """写入一行假设: vals 为标量或NF元列表"""
        put(ws, r, 1, label, 't')
        put(ws, r, 2, unit, 'g', align='center')
        if isinstance(vals, (list, tuple)):
            for i, v in enumerate(vals):
                put(ws, r, 3 + i, v, kind, fmt)
        else:
            put(ws, r, 3, vals, kind, fmt)
        put(ws, r, AB_COL, _ann(label, basis), 'g', size=9, wrap=True)
        return r

    def a5(path):
        return _broadcast(cfg.e(path)['values'], NF)

    def ab(path):
        return cfg.e(path)['basis']

    def a1(path):
        e = cfg.e(path)
        return e['values'][0] if isinstance(e['values'], list) else e['values']

    r = 4
    sec(ws, r, '一、全局与市场数据'); r += 1
    A['sw'] = arow(r, '情景开关 (1=熊/2=基准/3=牛)', '-', 2,
                   '三情景总开关, 驱动Revenue_Segments分部增速与毛利率、进而贯穿IS/BS/CF/DCF; 默认2=基准', 'in', '0'); r += 1
    put(ws, r, 1, '估值基准日 / 行情数据日期', 't'); put(ws, r, 3, VAL_DATE, 'x', align='center')
    put(ws, r, AB_COL, f'现价与可比公司市值均为{VAL_DATE}收盘(腾讯行情); DCF采用年中折现(见DCF页)', 'g', size=9); r += 1
    # F1: 港股双字段口径 — 同时存在market.price_hkd与market.fx_hkd(均>0)时按 price_hkd/fx_hkd
    #     重折算模型价(fetch_data接口约定: price字段仅为生成时的折算快照); 否则沿用market.price
    _px_hkd, _fx_hkd = cfg.v('market.price_hkd'), cfg.v('market.fx_hkd')
    if (isinstance(_px_hkd, (int, float)) and isinstance(_fx_hkd, (int, float))
            and _px_hkd > 0 and _fx_hkd > 0):
        _px_val = _px_hkd / _fx_hkd
        _px_basis = _norm_entry(cfg.get('market.price_hkd'))['basis'] or '港元现价/汇率折算为财报币种'
        _px_basis = f'{_px_basis} | =港元{_px_hkd:g}/fx{_fx_hkd:g}={_px_val:.4f}'
    else:
        _px_val, _px_basis = cfg.v('market.price'), ab('market.price')
    A['px'] = arow(r, '现价', PER_SHARE_UNIT, _px_val, _px_basis, 'x', PS); r += 1
    A['sh'] = arow(r, '总股本', '百万股', cfg.v('market.shares'), ab('market.shares'), 'x', NUM); r += 1
    put(ws, r, 1, '总市值', 't'); put(ws, r, 2, '百万', 'g', align='center')
    put(ws, r, 3, '=C{}*C{}'.format(A['px'], A['sh']), 'f', NUM)
    _mcap_yi = _px_val * cfg.v('market.shares') / 100
    put(ws, r, AB_COL, f'=现价×总股本, 约{_mcap_yi:.0f}亿元, 与行情软件一致', 'g', size=9); A['mcap'] = r; r += 1
    r += 1

    cons_rev = cfg.e('consensus.rev'); cons_np = cfg.e('consensus.np')
    _external_consensus = (research or {}).get('consensus')

    def _consensus_series(key):
        if _external_consensus is None:
            return None
        values = _external_consensus.get(key) if isinstance(_external_consensus, dict) else None
        return values if isinstance(values, dict) else {}

    CONSENSUS_VERIFIED = {
        'rev': {
            y: consensus_year_is_verified(cons_rev['basis'], y, _consensus_series('rev'))
            for y in FCST[:3]
        },
        'np': {
            y: consensus_year_is_verified(cons_np['basis'], y, _consensus_series('np'))
            for y in FCST[:3]
        },
    }
    CONSENSUS_LABELS = {
        key: {y: consensus_label(verified) for y, verified in by_year.items()}
        for key, by_year in CONSENSUS_VERIFIED.items()
    }
    CONS_REV_LABEL = CONSENSUS_LABELS['rev'][F0]
    CONS_NP_LABEL = CONSENSUS_LABELS['np'][F0]
    CONS_NP_VERIFIED = CONSENSUS_VERIFIED['np'][F0]
    _all_consensus_verified = all(
        verified for by_year in CONSENSUS_VERIFIED.values() for verified in by_year.values()
    )
    _cons_section_label = ('市场一致预期' if _all_consensus_verified
                           else '模型自动外推 / 经验证一致预期')
    sec(ws, r, f'二、{_cons_section_label}与最新财报 (外部/衍生数据, 绿色)'); r += 1
    cons_rev_v = _broadcast(cons_rev['values'], 3)
    cons_np_v = _broadcast(cons_np['values'], 3)
    for i, y in enumerate(FCST[:3]):
        A[f'c_rev{str(y)[2:]}'] = arow(r, f'{CONSENSUS_LABELS["rev"][y]}{y}E营收', '百万', cons_rev_v[i],
                                       f"{cons_rev['basis']}", 'x', NUM); r += 1
        A[f'c_np{str(y)[2:]}'] = arow(r, f'{CONSENSUS_LABELS["np"][y]}{y}E归母净利', '百万', cons_np_v[i],
                                      f"{cons_np['basis']}", 'x', NUM); r += 1
    lq = cfg.raw.get('latest_quarter') or {}
    if lq:
        A['q1_rev'] = arow(r, f"{lq.get('label', '最新季度')}实际营收", '百万',
                           _norm_entry(lq['rev'])['values'], _norm_entry(lq['rev'])['basis'], 'x', NUM); r += 1
        A['q1_np'] = arow(r, f"{lq.get('label', '最新季度')}实际归母净利", '百万',
                          _norm_entry(lq['np'])['values'], _norm_entry(lq['np'])['basis'], 'x', NUM); r += 1
    r += 1

    sec(ws, r, f'三、分部收入驱动 — 量/价增速 (蓝色输入, {F0}E-{F1}E)'); r += 1
    put(ws, r, 1, '年份', 't', bold=True)
    for i, y in enumerate(FCST):
        put(ws, r, 3 + i, f'{y}E', 't', bold=True, align='center')
    put(ws, r, AB_COL, '假设依据 / 来源', 't', bold=True); r += 1
    for sg in SEGS:
        k = sg['key']
        vol_lab = f"{sg['short']} 销量增速" if sg['driver'] == 'vol_asp' else f"{sg['short']} 收入增速"
        A[f'{k}_vol'] = arow(r, vol_lab, '%',
                             _broadcast(_norm_entry(sg['vol'])['values'], NF),
                             _norm_entry(sg['vol'])['basis']); r += 1
        if sg['driver'] == 'vol_asp':
            A[f'{k}_asp'] = arow(r, f"{sg['short']} ASP增速", '%',
                                 _broadcast(_norm_entry(sg['asp'])['values'], NF),
                                 _norm_entry(sg['asp'])['basis']); r += 1
    r += 1

    sec(ws, r, '四、分部毛利率 (蓝色输入)'); r += 1
    for sg in SEGS:
        k = sg['key']
        A[f'{k}_gm'] = arow(r, f"{sg['short']} 毛利率", '%',
                            _broadcast(_norm_entry(sg['gm'])['values'], NF),
                            _norm_entry(sg['gm'])['basis']); r += 1
    r += 1

    sec(ws, r, '五、费用率 / 其他损益 / 税率'); r += 1
    A['rt_taxadd'] = arow(r, '税金及附加率', '%', a5('opex.tax_add_rate'), ab('opex.tax_add_rate')); r += 1
    A['rt_sale'] = arow(r, '销售费用率', '%', a5('opex.sale_rate'), ab('opex.sale_rate')); r += 1
    A['rt_adm'] = arow(r, '管理费用率', '%', a5('opex.adm_rate'), ab('opex.adm_rate')); r += 1
    A['rt_rd'] = arow(r, '研发费用率', '%', a5('opex.rd_rate'), ab('opex.rd_rate')); r += 1
    A['oth_op'] = arow(r, '其他经营损益', '百万', a5('opex.oth_op'), ab('opex.oth_op'), 'in', NUM); r += 1
    A['nonop'] = arow(r, '营业外收支净额', '百万', a5('opex.nonop'), ab('opex.nonop'), 'in', NUM); r += 1
    A['tax'] = arow(r, '有效税率', '%', a5('opex.tax_rate'), ab('opex.tax_rate')); r += 1
    # A1: 少数股东损益占净利润比 — 预测期 归母净利润=合并净利润×(1-本行);
    #     默认按历史(合并净利-归母)/合并净利均值自动推导并clamp到[0,50%], config可经model.mi_share覆盖;
    #     历史合并净利按IS同口径倒推(收入-成本-税金-三费-财务+其他经营+营业外-所得税, othop含轧差)
    _is_hist = cfg.raw['hist']['is']
    _mi_ratios = []
    for _i in range(NH):
        _np_merge = (_is_hist['rev'][_i] - _is_hist['cost'][_i] - _is_hist['taxadd'][_i]
                     - _is_hist['sale'][_i] - _is_hist['adm'][_i] - _is_hist['rd'][_i]
                     - _is_hist['fin'][_i] + _is_hist['othop'][_i] + _is_hist['nonop'][_i]
                     - _is_hist['taxexp'][_i])
        if abs(_np_merge) > 1e-9:
            _mi_ratios.append((_np_merge - _is_hist['np_p'][_i]) / _np_merge)
    _mi_raw = sum(_mi_ratios) / len(_mi_ratios) if _mi_ratios else 0.0
    _mi_auto = min(0.5, max(0.0, round(_mi_raw, 6)))   # round防浮点噪声, 负值(少数股东亏损)clamp为0
    _mi_cfg = cfg.get('model.mi_share')
    if _mi_cfg is not None:
        _mi_ent = _norm_entry(_mi_cfg)
        _mi_vals = _broadcast(_mi_ent['values'], NF)
        _mi_basis = _mi_ent['basis'] or 'config覆盖(model.mi_share)'
    else:
        _mi_vals = [_mi_auto] * NF
        _mi_basis = (f'按历史均值自动推导: (合并净利-归母)/合并净利 {NH}年均值{_mi_raw:+.2%}, '
                     f'clamp[0,50%]→采用{_mi_auto:.2%}; config可经model.mi_share覆盖')
    A['mi_share'] = arow(r, '少数股东损益占净利润比', '%', _mi_vals, _mi_basis); r += 1
    A['oth_rt'] = arow(r, '其他及财务费用率合计(情景页用)', '%', a1('opex.oth_rate'),
                       ab('opex.oth_rate')); r += 1
    r += 1

    sec(ws, r, '六、营运资本天数'); r += 1
    A['dso'] = arow(r, 'DSO 应收(票据+账款)天数', '天', a5('working_capital.dso'), ab('working_capital.dso'), 'in', DAYS); r += 1
    A['dio'] = arow(r, 'DIO 存货天数', '天', a5('working_capital.dio'), ab('working_capital.dio'), 'in', DAYS); r += 1
    A['dpo'] = arow(r, 'DPO 应付(票据+账款)天数', '天', a5('working_capital.dpo'), ab('working_capital.dpo'), 'in', DAYS); r += 1
    A['rt_pre'] = arow(r, '预付款项/收入', '%', a5('working_capital.pre_rate'), ab('working_capital.pre_rate')); r += 1
    A['rt_staff'] = arow(r, '应付职工薪酬/收入', '%', a5('working_capital.staff_rate'), ab('working_capital.staff_rate')); r += 1
    A['rt_taxp'] = arow(r, '应交税费/收入', '%', a5('working_capital.taxp_rate'), ab('working_capital.taxp_rate')); r += 1
    r += 1

    sec(ws, r, '七、资本开支与折旧摊销 (PP&E页驱动)'); r += 1
    A['rt_capex'] = arow(r, 'Capex占收入比', '%', a5('capex.capex_rate'), ab('capex.capex_rate')); r += 1
    A['rt_trans'] = arow(r, '在建工程转固率', '%', a1('capex.trans_rate'), ab('capex.trans_rate')); r += 1
    A['rt_dep'] = arow(r, '固定资产折旧率(期初净值)', '%', a1('capex.dep_rate'), ab('capex.dep_rate')); r += 1
    A['rt_dep_new'] = arow(r, '当年转固部分折旧率', '%', a1('capex.dep_new_rate'), ab('capex.dep_new_rate')); r += 1
    A['rt_disp'] = arow(r, '固定资产处置率(期初净值)', '%', a1('capex.disp_rate'), ab('capex.disp_rate')); r += 1
    A['rt_amort'] = arow(r, '无形资产摊销率(期初)', '%', a1('capex.amort_rate'), ab('capex.amort_rate')); r += 1
    r += 1

    sec(ws, r, '八、股利与盈余公积'); r += 1
    A['payout'] = arow(r, f'股利支付率(按上年归母)', '%', a5('dividend.payout'), ab('dividend.payout')); r += 1
    A['sr_rate'] = arow(r, '法定盈余公积计提率(按当年归母)', '%', a1('dividend.surplus_rate'), ab('dividend.surplus_rate')); r += 1
    r += 1

    sec(ws, r, '九、债务与融资 (FIN页调度输入)'); r += 1
    A['min_cash_pct'] = arow(r, '最低现金占收入比', '%', a1('financing.min_cash_pct'), ab('financing.min_cash_pct')); r += 1
    A['rep_st'] = arow(r, '短期借款(revolver)计划还款', '百万', a5('financing.rep_st'), ab('financing.rep_st'), 'in', NUM); r += 1
    A['rep_cur'] = arow(r, '一年内到期非流动负债计划还款', '百万', a5('financing.rep_cur'), ab('financing.rep_cur'), 'in', NUM); r += 1
    A['rep_lt'] = arow(r, '长期借款计划还款', '百万', a5('financing.rep_lt'), ab('financing.rep_lt'), 'in', NUM); r += 1
    A['rep_lease'] = arow(r, '租赁负债计划还款', '百万', a5('financing.rep_lease'), ab('financing.rep_lease'), 'in', NUM); r += 1
    A['rate_st'] = arow(r, '短期借款利率', '%', a1('financing.rate_st'), ab('financing.rate_st')); r += 1
    A['rate_cur'] = arow(r, '一年内到期非流动负债利率', '%', a1('financing.rate_cur'), ab('financing.rate_cur')); r += 1
    A['rate_lt'] = arow(r, '长期借款利率', '%', a1('financing.rate_lt'), ab('financing.rate_lt')); r += 1
    A['rate_lease'] = arow(r, '租赁负债利率', '%', a1('financing.rate_lease'), ab('financing.rate_lease')); r += 1
    A['rt_cash'] = arow(r, '货币资金收益率(按平均余额)', '%', a1('financing.cash_yield'), ab('financing.cash_yield')); r += 1
    A['fin_oth'] = arow(r, '其他财务费用(手续费/汇兑)', '百万', a1('financing.fin_oth'), ab('financing.fin_oth'), 'in', NUM); r += 1
    _bs_hist = cfg.raw['hist']['bs']
    _d0 = round(_bs_hist['stl'][-1] + _bs_hist['cur1y'][-1] + _bs_hist['ltl'][-1] + _bs_hist['lease'][-1], 4)
    put(ws, r, 1, f'{H0}A末有息负债合计', 't'); put(ws, r, 2, '百万', 'g', align='center')
    put(ws, r, 3, _d0, 'x', NUM)
    put(ws, r, AB_COL, f"=短借{_bs_hist['stl'][-1]:.1f}+一年内{_bs_hist['cur1y'][-1]:.1f}+长借{_bs_hist['ltl'][-1]:.1f}"
                  f"+租赁{_bs_hist['lease'][-1]:.1f} (东财F10 {H0}年报); WACC的Wd市值口径引用本行", 'g', size=9)
    A['debt25'] = r; r += 1
    r += 1

    sec(ws, r, '十、WACC 与终值 (全公式化, 联动Relative_Val可比beta)'); r += 1
    A['rf'] = arow(r, '无风险利率 rf', '%', a1('wacc.rf'), ab('wacc.rf')); r += 1
    A['erp'] = arow(r, '股权风险溢价 ERP', '%', a1('wacc.erp'), ab('wacc.erp')); r += 1
    put(ws, r, 1, '可比公司无杠杆βu中位数', 't'); put(ws, r, 2, 'x', 'g', align='center')
    if N_BETA:
        put(ws, r, 3, f"=Relative_Val!$O${RV_MED_ROW}", 'f', '0.000')
        put(ws, r, AB_COL, '=MEDIAN(可比公司βl/(1+(1-t)×D/E)), 见Relative_Val页L-O列; 改可比beta输入即联动', 'g', size=9)
        # F3: 存在可比beta时βu以可比中位数为准; 若同时配置了beta_unlevered_input, 显式提示其未生效
        if cfg.v('wacc.beta_unlevered_input') is not None:
            print(f"警告: 配置了 beta_unlevered_input={cfg.v('wacc.beta_unlevered_input')}, "
                  f"但存在{N_BETA}家可比beta, 已采用可比中位数, 该输入未生效")
    else:
        _bu = cfg.v('wacc.beta_unlevered_input', 1.0)
        put(ws, r, 3, _bu, 'in', '0.000')
        put(ws, r, AB_COL, cfg.v('wacc.beta_unlevered_basis', '未配置可比公司, βu为蓝色输入(兜底), 建议补充relative_val.comps'), 'g', size=9)
    A['beta_u'] = r; r += 1
    put(ws, r, 1, '目标债务权重 Wd (市值口径)', 't'); put(ws, r, 2, '%', 'g', align='center')
    put(ws, r, 3, f"=C{A['debt25']}/(C{A['debt25']}+C{A['mcap']})", 'f', PCT)
    put(ws, r, AB_COL, cfg.v('wacc.wd_basis',
        f'={H0}A末有息负债/(有息负债+总市值): 以市值口径现状结构为目标结构'), 'g', size=9)
    A['wd'] = r; r += 1
    put(ws, r, 1, '目标 D/E = Wd/(1-Wd)', 't'); put(ws, r, 2, 'x', 'g', align='center')
    put(ws, r, 3, f"=C{A['wd']}/(1-C{A['wd']})", 'f', '0.000')
    put(ws, r, AB_COL, 'relever用目标产权比率', 'g', size=9); A['de_t'] = r; r += 1
    put(ws, r, 1, '再杠杆βl = βu×(1+(1-t)×D/E)', 't'); put(ws, r, 2, 'x', 'g', align='center')
    put(ws, r, 3, f"=C{A['beta_u']}*(1+(1-C{A['tax']})*C{A['de_t']})", 'f', '0.000')
    put(ws, r, AB_COL, f'Hamada公式, t={F0}E有效税率', 'g', size=9); A['beta'] = r; r += 1
    A['srp'] = arow(r, '个股/执行风险溢价 (已删除=0)', '%', a1('wacc.srp'), ab('wacc.srp')); r += 1
    put(ws, r, 1, '股权成本 Ke = rf+βl×ERP+溢价', 't'); put(ws, r, 2, '%', 'g', align='center')
    put(ws, r, 3, f"=C{A['rf']}+C{A['beta']}*C{A['erp']}+C{A['srp']}", 'f', PCT)
    put(ws, r, AB_COL, 'CAPM公式(溢价默认为0)', 'g', size=9); A['ke'] = r; r += 1
    A['kd'] = arow(r, '债务成本 Kd', '%', a1('wacc.kd'), ab('wacc.kd')); r += 1
    put(ws, r, 1, 'WACC (计算值)', 't'); put(ws, r, 2, '%', 'g', align='center')
    put(ws, r, 3, f"=(1-C{A['wd']})*C{A['ke']}+C{A['wd']}*C{A['kd']}*(1-C{A['tax']})", 'f', PCT)
    put(ws, r, AB_COL, '=(1-Wd)×Ke + Wd×Kd×(1-t)', 'g', size=9); A['wacc_calc'] = r; r += 1
    put(ws, r, 1, 'WACC override (留空=用计算值)', 't'); put(ws, r, 2, '%', 'g', align='center')
    put(ws, r, 3, cfg.v('wacc.override', None), 'in', PCT)
    put(ws, r, AB_COL, '蓝色输入, 留空则采用上方计算值; 填入数值则覆盖(Checks页校验采用值=计算值或override非空)', 'g', size=9)
    A['wacc_ovr'] = r; r += 1
    put(ws, r, 1, 'WACC (采用值)', 't', ); put(ws, r, 2, '%', 'g', align='center')
    c = put(ws, r, 3, f"=IF(C{A['wacc_ovr']}=\"\",C{A['wacc_calc']},C{A['wacc_ovr']})", 'f', PCT, bold=True, fill=CHK_FILL)
    put(ws, r, AB_COL, '=IF(override="", 计算值, override); DCF/Sensitivity/Scenarios全部引用本行', 'g', size=9)
    A['wacc'] = r; r += 1
    A['tg'] = arow(r, '永续增长率 g', '%', a1('wacc.tg'), ab('wacc.tg')); r += 1
    r += 1

    sec(ws, r, '十一、情景参数 (驱动情景开关 & Scenarios页)'); r += 1
    put(ws, r, 1, '情景', 't', bold=True)
    put(ws, r, 3, '收入增速调整', 't', bold=True, align='center')
    put(ws, r, 4, '净利率调整', 't', bold=True, align='center')
    put(ws, r, AB_COL, '情景逻辑', 't', bold=True); r += 1
    for skey, sname, skind in [('bear', '熊市 Bear', 'w'), ('base', '基准 Base', 't'), ('bull', '牛市 Bull', 't')]:
        scfg = cfg.raw['scenarios'][skey]
        put(ws, r, 1, sname, skind, bold=True)
        put(ws, r, 3, scfg['rev_adj'], 'in', PCT); put(ws, r, 4, scfg['npm_adj'], 'in', PCT)
        put(ws, r, AB_COL, scfg.get('logic', ''), 'g', size=9)
        A[f'sc_{skey}'] = r; r += 1

    ASheet = 'Assumptions'
    print('Assumptions rows:', A)

    # ============================================================
    # Revenue_Segments
    # ============================================================
    ws = wb.create_sheet('Revenue_Segments')
    title_bar(ws, f'{NAME} — Revenue_Segments 分业务收入拆解 (量×价驱动)',
              f'单位: {UNIT} | 历史收入=产品占比×总收入; 预测收入=上年×(1+量增速)×(1+价增速), 增速引用Assumptions并受情景开关调整')
    ws.column_dimensions['A'].width = 30
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 90
    ws.freeze_panes = 'B4'

    SW = f"{ASheet}!$C${A['sw']}"
    SC_BEAR = f"{ASheet}!$C${A['sc_bear']}"
    SC_BASE = f"{ASheet}!$C${A['sc_base']}"
    SC_BULL = f"{ASheet}!$C${A['sc_bull']}"
    GM_BEAR = f"{ASheet}!$D${A['sc_bear']}"
    GM_BASE = f"{ASheet}!$D${A['sc_base']}"
    GM_BULL = f"{ASheet}!$D${A['sc_bull']}"

    def adj_g(rowkey):
        """情景调整后的增速引用: Assumptions基础值 + CHOOSE(开关,熊,基,牛)"""
        def f(y):
            return f"({ASheet}!{AC[y]}${A[rowkey]}+CHOOSE({SW},{SC_BEAR},{SC_BASE},{SC_BULL}))"
        return f

    def adj_gm(rowkey):
        def f(y):
            return f"({ASheet}!{AC[y]}${A[rowkey]}+CHOOSE({SW},{GM_BEAR},{GM_BASE},{GM_BULL}))"
        return f

    r = 3
    put(ws, r, 1, '项目', 't', bold=True)
    for y in YRS:
        lab = f'{y}A' if y in HIST else f'{y}E'
        c = put(ws, r, 2 + YRS.index(y), lab, 't', bold=True, align='center')
        c.fill = TOT_FILL
    put(ws, r, NOTE_COL, '备注 / 来源', 't', bold=True)
    r += 1
    # 历史总收入参考行
    put(ws, r, 1, '营业收入合计(历史实际)', 't')
    for y, v in zip(HIST, cfg.raw['hist']['is']['rev']):
        put(ws, r, 2 + YRS.index(y), v, 'x', NUM)
    put(ws, r, NOTE_COL, f'东财F10利润表: {HIST[0]}-{H0}实际营收', 'g', size=9)
    TOT_HIST = r
    r += 1

    RSEG = {}   # segment -> dict of rows
    for sg in SEGS:
        key, name = sg['key'], sg['name']
        mix, gm_hist = sg['hist_share'], sg['hist_gm']
        volk, gmk = f"{key}_vol", f"{key}_gm"
        aspk = f"{key}_asp" if sg['driver'] == 'vol_asp' else None
        logic = sg.get('logic', '')
        sec(ws, r, name); r += 1
        rows = {}
        # 收入
        put(ws, r, 1, '收入', 't')
        for i, y in enumerate(HIST):
            cl = YC[y]
            put(ws, r, 2 + i, f"={cl}{r + 1}*{cl}${TOT_HIST}", 'f', NUM)
        rows['rev'] = r
        r += 1
        # 占比
        put(ws, r, 1, '  收入占比', 't')
        for i, y in enumerate(HIST):
            put(ws, r, 2 + i, mix[i], 'x', PCT)
        put(ws, r, NOTE_COL, sg.get('share_basis', '占比: 东财F10主营构成(按产品)'), 'g', size=9)
        rows['share'] = r
        r += 1
        # 量增速 (growth驱动业务为合并增速)
        lab = '  销量增速' if aspk else '  收入增速'
        put(ws, r, 1, lab, 't')
        for i in range(NH):
            put(ws, r, 2 + i, '—', 'g', align='center')
        for y in FCST:
            put(ws, r, 2 + YRS.index(y), f"={adj_g(volk)(y)}", 'f', PCT)
        put(ws, r, NOTE_COL, '=Assumptions基础增速+CHOOSE(情景开关,熊,基,牛)', 'g', size=9)
        rows['vol'] = r
        r += 1
        # ASP增速
        if aspk:
            put(ws, r, 1, '  ASP增速', 't')
            for i in range(NH):
                put(ws, r, 2 + i, '—', 'g', align='center')
            for y in FCST:
                put(ws, r, 2 + YRS.index(y), f"={ASheet}!{AC[y]}${A[aspk]}", 'f', PCT)
            rows['asp'] = r
            r += 1
        # 毛利率
        put(ws, r, 1, '  毛利率', 't')
        for i in range(NH):
            put(ws, r, 2 + i, gm_hist[i], 'x', PCT)
        for y in FCST:
            put(ws, r, 2 + YRS.index(y), f"={adj_gm(gmk)(y)}", 'f', PCT)
        put(ws, r, NOTE_COL, '历史: 主营构成分产品毛利率; 预测: Assumptions+情景调整', 'g', size=9)
        rows['gm'] = r
        r += 1
        # 毛利
        put(ws, r, 1, '  毛利', 't')
        for y in YRS:
            cl = YC[y]
            put(ws, r, 2 + YRS.index(y), f"={cl}{rows['rev']}*{cl}{rows['gm']}", 'f', NUM)
        rows['gp'] = r
        r += 1
        # 预测收入公式 (放在占比等行确定后回填)
        rr = rows['rev']
        for y in FCST:
            cl, pl = YC[y], YC[y - 1]
            if aspk:
                f = f"={pl}{rr}*(1+{cl}{rows['vol']})*(1+{cl}{rows['asp']})"
            else:
                f = f"={pl}{rr}*(1+{cl}{rows['vol']})"
            put(ws, rr, 2 + YRS.index(y), f, 'f', NUM)
        # 占比预测列在REV_TOT行号确定后统一写入(见"回填占比预测列"), 此处不再先写占位公式
        # 逻辑
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NOTE_COL)
        put(ws, r, 1, '  逻辑: ' + logic, 'g', size=9, wrap=True)
        rows['logic'] = r
        r += 1
        RSEG[key] = rows

    # 合计
    sec(ws, r, '合计 (自动汇总进IS)'); r += 1
    seg_keys = [sg['key'] for sg in SEGS]
    put(ws, r, 1, '营业收入合计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        expr = '+'.join(f"{cl}{RSEG[k]['rev']}" for k in seg_keys)
        put(ws, r, 2 + YRS.index(y), f"={expr}", 'f', NUM, bold=True, fill=TOT_FILL)
    REV_TOT = r; r += 1
    put(ws, r, 1, '毛利合计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        expr = '+'.join(f"{cl}{RSEG[k]['gp']}" for k in seg_keys)
        put(ws, r, 2 + YRS.index(y), f"={expr}", 'f', NUM, bold=True)
    GP_TOT = r; r += 1
    put(ws, r, 1, '综合毛利率', 't')
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{GP_TOT}/{cl}{REV_TOT}", 'f', PCT)
    GM_TOT = r; r += 1
    put(ws, r, 1, '收入YoY', 't')
    for y in YRS[1:]:
        cl, pl = YC[y], YC[y - 1]
        put(ws, r, 2 + YRS.index(y), f"={cl}{REV_TOT}/{pl}{REV_TOT}-1", 'f', PCT)
    r += 1
    # H4: 回填占比预测列 — REV_TOT行号确定后一次性写入完整公式与格式(不再先写坏公式再补丁)
    for key in RSEG:
        for y in FCST:
            cl = YC[y]
            put(ws, RSEG[key]['share'], 2 + YRS.index(y), f"={cl}{RSEG[key]['rev']}/{cl}{REV_TOT}", 'f', PCT)
    r += 1

    # 基准情形演算 (不受情景开关影响, 供Scenarios页引用)
    sec(ws, r, '基准情形演算 (仅引用Assumptions基础值, 不受情景开关影响, 供Scenarios页引用)'); r += 1
    BASE_SEG = {}
    base_defs = [(sg['key'], f"{sg['key']}_vol",
                  f"{sg['key']}_asp" if sg['driver'] == 'vol_asp' else None,
                  f"{sg['key']}_gm") for sg in SEGS]
    for key, volk, aspk, gmk in base_defs:
        put(ws, r, 1, f'基准收入 — {key}', 't')
        # C2①: 种子写入H0列(动态), 3历史年时=第4列与旧版面等值
        put(ws, r, 2 + YRS.index(H0), f"={YC[H0]}{RSEG[key]['rev']}", 'f', NUM)
        for y in FCST:
            cl, pl = YC[y], YC[y - 1]
            if aspk:
                f = f"={pl}{r}*(1+{ASheet}!{AC[y]}${A[volk]})*(1+{ASheet}!{AC[y]}${A[aspk]})"
            else:
                f = f"={pl}{r}*(1+{ASheet}!{AC[y]}${A[volk]})"
            put(ws, r, 2 + YRS.index(y), f, 'f', NUM)
        BASE_SEG[key] = r
        r += 1
    put(ws, r, 1, '基准合计收入', 't', bold=True)
    for y in FCST:
        cl = YC[y]
        expr = '+'.join(f"{cl}{BASE_SEG[k]}" for k in seg_keys)
        put(ws, r, 2 + YRS.index(y), f"={expr}", 'f', NUM)
    # C2①: 合计种子同样写入H0列(动态)
    put(ws, r, 2 + YRS.index(H0), '=' + '+'.join(f"{YC[H0]}{RSEG[k]['rev']}" for k in seg_keys), 'f', NUM)
    BASE_REV = r; r += 1
    put(ws, r, 1, '基准收入增速', 't')
    for y in FCST:
        cl, pl = YC[y], YC[y - 1]
        put(ws, r, 2 + YRS.index(y), f"={cl}{BASE_REV}/{pl}{BASE_REV}-1", 'f', PCT)
    BASE_G = r; r += 1
    put(ws, r, 1, '基准综合毛利率', 't')
    for y in FCST:
        cl = YC[y]
        num = '+'.join(f"{cl}{BASE_SEG[k]}*{ASheet}!{AC[y]}${A[gmk]}"
                       for k, _, _, gmk in base_defs)
        put(ws, r, 2 + YRS.index(y), f"=({num})/{cl}{BASE_REV}", 'f', PCT)
    BASE_GM = r; r += 1
    r += 1
    # 产能/客户/行业参考 (绿色外部信息, 可选)
    refs = cfg.get('references') or []
    if refs:
        sec(ws, r, '产能 / 客户 / 行业参考 (外部信息)'); r += 1
        for txt in refs:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NOTE_COL)
            put(ws, r, 1, txt, 'x', size=9, wrap=True)
            ws.row_dimensions[r].height = 24
            r += 1

    RS = 'Revenue_Segments'
    print('RevSeg: REV_TOT', REV_TOT, 'GP_TOT', GP_TOT, 'GM_TOT', GM_TOT, 'BASE_REV', BASE_REV, 'BASE_G', BASE_G, 'BASE_GM', BASE_GM)

    # ============================================================
    # Schedules (NWC / 无形资产 / 股利) — 骨架, 公式后填
    # ============================================================
    ws = wb.create_sheet('Schedules')
    title_bar(ws, f'{NAME} — Schedules 营运资本 / 无形资产 / 股利',
              f'单位: {UNIT} | PP&E见PP&E页, 债务与利息见FIN页; 预测行全公式')
    ws.column_dimensions['A'].width = 30
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 60
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    S = {}
    r = 5
    sec(ws, r, '一、营运资本 (NWC)'); r += 1
    for key, lab, note in [
        ('ar', '应收票据及应收账款', '=BS, 预测=DSO/365×收入'),
        ('pre', '预付款项', '=BS, 预测=收入×预付率'),
        ('orec', '其他应收款', '=BS, 预测持平'),
        ('inv', '存货', '=BS, 预测=DIO/365×营业成本'),
        ('nwc_a', '经营性流动资产小计', ''),
        ('ap', '应付票据及应付账款', '=BS, 预测=DPO/365×营业成本'),
        ('contract', '合同负债', '=BS, 持平'),
        ('staff', '应付职工薪酬', '=BS, 预测=收入×比率'),
        ('taxp', '应交税费', '=BS, 预测=收入×比率'),
        ('opay', '其他应付款', '=BS, 持平'),
        ('nwc_l', '经营性流动负债小计', ''),
        ('nwc', '净营运资本 NWC', '=经营性流动资产-经营性流动负债'),
        ('dnwc', 'ΔNWC (增加为正)', 'CF中扣减项'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key in ('nwc_a', 'nwc_l', 'nwc')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        S[key] = r
        r += 1
    r += 1
    sec(ws, r, '二、无形资产摊销'); r += 1
    _amort_rt = a1('capex.amort_rate')
    for key, lab, note in [
        ('ia0', '期初无形资产', f'{F0}期初={H0}A期末'),
        ('amort', '无形资产摊销', f'=期初×{_amort_rt:.0%}'),
        ('ia1', '期末无形资产', '=期初-摊销'),
    ]:
        put(ws, r, 1, lab, 't')
        put(ws, r, NOTE_COL, note, 'g', size=9)
        S[key] = r
        r += 1
    r += 1
    sec(ws, r, '三、股利'); r += 1
    _payout0 = a5('dividend.payout')[0]
    put(ws, r, 1, '现金分红(当年支付)', 't')
    put(ws, r, NOTE_COL, f'=上年归母净利×支付率{_payout0:.0%}', 'g', size=9)
    S['div'] = r
    r += 1
    SCH = 'Schedules'
    print('Schedules rows:', S)

    # ============================================================
    # PPE 固定资产页 — 骨架
    # ============================================================
    _fa_net0 = cfg.v('hist.ppe_split.fa_net')
    _cip0 = cfg.v('hist.ppe_split.cip')
    _dep_rt = a1('capex.dep_rate')
    _dep_new = a1('capex.dep_new_rate')
    _disp_rt = a1('capex.disp_rate')
    _trans_rt = a1('capex.trans_rate')
    ws = wb.create_sheet('PPE')
    title_bar(ws, f'{NAME} — PP&E 固定资产/在建工程滚动',
              f'单位: {UNIT} | 固定资产与在建工程分行滚动: 期初+转固-折旧-处置 / 期初+Capex-转固; 隐含折旧率/年限校验行在底部')
    ws.column_dimensions['A'].width = 30
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 66
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    P = {}
    r = 5
    sec(ws, r, '一、固定资产 (净值口径)'); r += 1
    for key, lab, note in [
        ('fa0', '期初固定资产净值', f'{F0}期初={H0}A固定资产{_fa_net0/100:.1f}亿(绿色)'),
        ('trans', '加: 在建工程转入', f'=(期初在建+当年Capex)×转固率{_trans_rt:.0%}'),
        ('dep', '减: 固定资产折旧', f'=期初净值×{_dep_rt:.0%}+当年转固×{_dep_new:.0%}'),
        ('disp', '减: 处置', f'=期初净值×{_disp_rt:.0%}'),
        ('fa1', '期末固定资产净值', '=期初+转固-折旧-处置'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key == 'fa1'))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        P[key] = r
        r += 1
    r += 1
    sec(ws, r, '二、在建工程'); r += 1
    for key, lab, note in [
        ('cip0', '期初在建工程', f'{F0}期初={H0}A在建工程{_cip0/100:.1f}亿(绿色)'),
        ('capex', '加: 资本开支 Capex', '=收入×Capex占收入比(Assumptions)'),
        ('cip1', '期末在建工程', '=期初+Capex-转固'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key == 'cip1'))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        P[key] = r
        r += 1
    r += 1
    sec(ws, r, '三、合计与校验'); r += 1
    for key, lab, note in [
        ('ppe1', '固定资产+在建工程 期末合计', f'=BS固定资产及在建工程; {H0}A合计{(_fa_net0+_cip0)/100:.1f}亿'),
        ('da', 'D&A 合计 (折旧+无形摊销)', '=折旧+Schedules无形摊销, 进CF与DCF'),
        ('dep_rt', '校验: 隐含折旧率(折旧/期初净值)', '信息行; 合理带宽8%-18%'),
        ('dep_yrs', '校验: 隐含折旧年限(期初净值/折旧)', '信息行; 机器设备为主, 5-10年为合理带'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key in ('ppe1', 'da')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        P[key] = r
        r += 1
    PPEs = 'PPE'
    print('PPE rows:', P)

    # ============================================================
    # FIN 债务与融资调度页 — 骨架
    # ============================================================
    ws = wb.create_sheet('FIN')
    title_bar(ws, f'{NAME} — FIN 债务与融资调度 (迭代展开·无循环引用)',
              f'单位: {UNIT} | 瀑布: 期初现金+经营CF+投资CF(不含理财)+分红+计划还款+利息→缺口=revolver新增借款→超额现金sweep还短借/长借→溢出购理财; 利息=平均余额×利率的循环链在下方"四、迭代展开"区表内展开4轮, 末轮输出供全簿引用, 无需开启迭代计算')
    ws.column_dimensions['A'].width = 34
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 70
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        lab = f'{y}A' if y in HIST else f'{y}E'
        put(ws, 3, 2 + YRS.index(y), lab, 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    F = {}
    r = 5
    sec(ws, r, '一、资金瀑布 (融资需求与可动用现金)'); r += 1
    for key, lab, note in [
        ('cash0', '期初货币资金', f'{F0}={H0}A末; 以后=上年期末'),
        ('ocf', '加: 经营活动现金流(不含利息)', '=CF经营净额(利息已重分类至筹资)'),
        ('icf', '加: 投资活动现金流(不含理财)', '=-Capex+处置回收(PP&E页)'),
        ('div', '减: 计划分红(负值)', f'=-上年归母×{_payout0:.0%} (Schedules)'),
        ('intpay', '减: 偿付利息(负值)', '=-本页利息支出合计(=第4轮迭代展开输出)'),
        ('rep', '减: 实际还款合计(负值)', '=四档实际还款加总(实际还款=MIN(计划还款,期初余额))'),
        ('cpre', '融资前现金结余', '=期初+经营+投资+分红+利息+还款'),
        ('mincash', '最低现金', f'=收入×{a1("financing.min_cash_pct"):.0%} (Assumptions)'),
        ('new', '资金缺口→新增借款(revolver)', '=MAX(0, 最低现金-融资前结余)'),
        ('avail', '可动用现金(还债+购理财)', '=MAX(0, 融资前结余+新增借款-最低现金)'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key in ('cpre', 'new', 'avail')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        F[key] = r
        r += 1
    r += 1
    for tranche, tname, repk, ratek, sweepable in [
        ('st', '短期借款 (revolver: 新增借款入口, sweep第一顺位)', 'rep_st', 'rate_st', True),
        ('cur', '一年内到期的非流动负债 (仅计划还款)', 'rep_cur', 'rate_cur', False),
        ('lt', '长期借款 (计划还款+sweep第二顺位)', 'rep_lt', 'rate_lt', True),
        ('le', '租赁负债 (仅计划还款)', 'rep_lease', 'rate_lease', False),
    ]:
        sec(ws, r, f'· {tname}')
        r += 1
        rows_def = [('t0', '期初余额', '上年期末')]
        rows_def.append(('rep', '计划还款(输入)', f'=Assumptions {repk}'))
        rows_def.append(('rep_a', '实际还款(≤期初)', '=MIN(计划还款, 期初余额); 防止债务清零后仍按计划额扣现金'))
        if sweepable:
            if tranche == 'st':
                rows_def.append(('new', '新增借款', '=资金缺口(全部经revolver融入)'))
            rows_def.append(('pre', 'sweep前未偿余额', '=期初-实际还款+新增'))
            rows_def.append(('sw', '现金sweep偿还', '=MIN(可动用现金余额, sweep前未偿)'))
            rows_def.append(('t1', '期末余额', '=sweep前未偿-sweep'))
        else:
            rows_def.append(('t1', '期末余额', '=期初-实际还款'))
        rows_def += [('avg', '平均余额', '=(期初+期末)/2'),
                     ('int', '利息支出', '=平均余额×利率(按期末余额重算的隐含值; 与采用值(第4轮迭代输出)之差见下方残差行)')]
        for key2, lab, note in rows_def:
            put(ws, r, 1, lab, 't', bold=(key2 == 't1'))
            put(ws, r, NOTE_COL, note, 'g', size=9)
            F[f'{tranche}_{key2}'] = r
            r += 1
        r += 1
    sec(ws, r, '三、汇总 / 理财 / 现金期末'); r += 1
    _fin_oth_v = a1('financing.fin_oth')
    for key, lab, note in [
        ('dtot0', '有息负债期初合计', '四档期初加总'),
        ('dtot1', '有息负债期末合计', '四档期末加总'),
        ('ddebt', '债务净变动(增加为正)', '=期末-期初, 进CF筹资'),
        ('sw_tot', '现金sweep合计', '=短借sweep+长借sweep'),
        ('overflow', '溢出资金→购买理财', '=可动用现金-sweep合计(债务清零后剩余)'),
        ('tfa0', '期初交易性金融资产(理财)', f'{F0}={H0}A末{_bs_hist["tfa"][-1]/100:.2f}亿'),
        ('tfa1', '期末交易性金融资产(理财)', '=期初+溢出购买; 简化: 缺口年不赎回'),
        ('dtfa', '理财净变动', '=期末-期初, 进CF投资活动'),
        ('cash1', '期末货币资金(=最低现金)', '=融资前结余+新增-sweep-溢出, 恒等于最低现金'),
        ('intexp', '利息支出合计', '=第4轮迭代展开输出(四档平均余额×利率); 上方分档利息行为隐含重算值, 差额见残差行'),
        ('intinc', '利息收入', f'=第4轮迭代展开输出(货币资金平均余额×{a1("financing.cash_yield"):.1%})'),
        ('finoth', '其他财务费用', f'=Assumptions {_fin_oth_v:.0f}百万'),
        ('finnet', '财务费用净额(→IS)', '=利息支出-利息收入+其他'),
    ]:
        put(ws, r, 1, lab, 't', bold=(key in ('dtot1', 'cash1', 'finnet')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        F[key] = r
        r += 1
    # --- 四、迭代展开区 (利息↔净利↔现金↔sweep↔利息循环的表内显式展开, 全簿无循环引用) ---
    r += 1
    sec(ws, r, '四、迭代展开 (iteration unrolling — 循环链表内展开, 替代Excel迭代计算)'); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NOTE_COL)
    put(ws, r, 1, '方法: 第1轮按"期初债务×利率"计息→算期末债务₁/现金₁; 第k轮(k=2..4)按"(期初+上轮期末)/2×利率"重算→期末_k; '
                  '第4轮输出回填上方汇总行(利息支出合计/利息收入/财务费用净额)供IS/CF/BS引用; '
                  '每轮残差=|Δ利息支出|+|Δ利息收入|(相对上轮), <0.01即收敛(外部不动点法实测6轮内收敛至1e-8, 4轮充足)', 'g', size=9, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1
    IT = {}
    for k in range(1, 5):
        sec(ws, r, f'· 第{k}轮迭代 (iter{k})'); r += 1
        d = {}
        for key2, lab in [
            ('intexp_in', '采用利息支出' + (' (=期初债务×利率)' if k == 1 else ' (=上轮重算值)')),
            ('intinc_in', '采用利息收入' + (' (=期初现金×收益率)' if k == 1 else ' (=上轮重算值)')),
            ('finnet_in', '财务费用净额(采用口径)'),
            ('np', '归母净利(按采用利息重算)'),
            ('ocf', '经营CF(按采用利息重算)'),
            ('cpre', '融资前现金结余'),
            ('new', '新增借款(revolver)'),
            ('avail', '可动用现金'),
            ('st1', '短期借款期末'),
            ('lt1', '长期借款期末'),
            ('intexp_out', '利息支出(按期末余额重算)'),
            ('cash1', '期末现金'),
            ('intinc_out', '利息收入(按期末现金重算)'),
            ('finnet_out', '财务费用净额(重算)'),
            ('resid', '残差 |Δ利息支出|+|Δ利息收入|'),
        ]:
            put(ws, r, 1, f'iter{k}: {lab}', 't', bold=(key2 == 'resid'))
            d[key2] = r
            r += 1
        IT[k] = d
    r += 1
    sec(ws, r, '五、收敛残差 (第4轮采用值 vs 按期末余额重算隐含值, <0.01即收敛)'); r += 1
    for key2, lab, note in [
        ('resid_int', '残差: 利息支出 |采用值-隐含值|', '隐含值=上方四档利息行(按期末余额重算)加总'),
        ('resid_inc', '残差: 利息收入 |采用值-隐含值|', f'隐含值=(期初+期末货币资金)/2×{a1("financing.cash_yield"):.1%}'),
        ('resid_max', '残差上限 (应<0.01)', '两项取大; Checks页第11项汇总校验'),
        ('resid_ok', '收敛判定 (TRUE=收敛)', ''),
    ]:
        put(ws, r, 1, lab, 't', bold=(key2 in ('resid_max', 'resid_ok')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        F[key2] = r
        r += 1
    FINs = 'FIN'
    print('FIN rows:', F)
    print('FIN iter blocks intexp_in rows:', {k: v['intexp_in'] for k, v in IT.items()})

    # ============================================================
    # Equity_Roll 权益滚动页 — 骨架
    # ============================================================
    ws = wb.create_sheet('Equity_Roll')
    title_bar(ws, f'{NAME} — Equity_Roll 所有者权益滚动',
              f'单位: {UNIT} | 股本/资本公积/库存股/OCI持平滚动; 盈余公积按归母计提; 未分配=期初+归母-分红-盈余公积计提')
    ws.column_dimensions['A'].width = 30
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 60
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    Q = {}
    r = 5
    _sr_rt = a1('dividend.surplus_rate')
    for key, lab, note in [
        ('sc', '股本', '预测持平(不考虑增发)'),
        ('cr', '资本公积', '预测持平'),
        ('ts', '减: 库存股', '预测持平'),
        ('oci', '其他综合收益', '预测持平(汇率/公允变动不预测)'),
        ('sr_acc', '盈余公积计提(流量)', f'=当年归母×{_sr_rt:.0%} (Assumptions)'),
        ('sr', '盈余公积(期末)', '=期初+计提'),
        ('div', '现金分红(流量, 引用)', '=Schedules'),
        ('re', '未分配利润(期末)', '=期初+归母净利-分红-盈余公积计提'),
        ('oeq', '其他权益项目(轧差, 引用BS)', '含专项储备及换算尾差, 持平'),
        ('te_p', '归母所有者权益合计', ''),
        ('mi', '少数股东权益', ''),
        ('te', '所有者权益合计', ''),
    ]:
        put(ws, r, 1, lab, 't', bold=(key in ('te_p', 'te')))
        put(ws, r, NOTE_COL, note, 'g', size=9)
        Q[key] = r
        r += 1
    r += 1
    sec(ws, r, '稀释股数'); r += 1
    put(ws, r, 1, '基本股数(总股本, 百万股)', 't')
    put(ws, r, 3, f"={ASheet}!$C${A['sh']}", 'f', NUM)
    put(ws, r, NOTE_COL, f'=Assumptions总股本{cfg.v("market.shares")/100:.2f}亿股', 'g', size=9)
    Q['sh_basic'] = r; r += 1
    put(ws, r, 1, '稀释工具(可转债/期权等, 百万股)', 't')
    put(ws, r, 3, cfg.v('equity.dilutive', 0.0), 'in', NUM)
    put(ws, r, NOTE_COL, cfg.v('equity.dilutive_basis', '占位输入: 如有流通可转债/期权等稀释工具在此输入'), 'g', size=9)
    Q['sh_dil'] = r; r += 1
    put(ws, r, 1, '稀释后股数(百万股)', 't', bold=True)
    c = put(ws, r, 3, f"=C{Q['sh_basic']}+C{Q['sh_dil']}", 'f', NUM, bold=True, fill=CHK_FILL)
    put(ws, r, NOTE_COL, 'DCF/相对估值/Scenarios每股口径统一引用本行', 'g', size=9)
    Q['shares'] = r; r += 1
    EQs = 'Equity_Roll'
    print('Equity_Roll rows:', Q)

    # ============================================================
    # IS 利润表
    # ============================================================
    _is_h = cfg.raw['hist']['is']
    _rev_h = _is_h['rev']
    _g1 = _rev_h[1] / _rev_h[0] - 1
    _g2 = _rev_h[2] / _rev_h[2 - 1] - 1 if NH >= 3 else None
    ws = wb.create_sheet('IS')
    title_bar(ws, f'{NAME} — IS 利润表 ({HIST[0]}A-{F1}E)',
              f'单位: {UNIT} | 收入/成本引用Revenue_Segments; 费用率/税率引用Assumptions; 财务费用引用FIN(平均余额计息, FIN页迭代展开第4轮输出)')
    ws.column_dimensions['A'].width = 26
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 70
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注 / 来源', 't', bold=True)

    I = {}
    r = 4
    def isrow(key, label, hist_vals, fcst_fn, note='', fmt=NUM, bold=False, hist_kind='x'):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        if hist_vals is not None:
            for i, v in enumerate(hist_vals):
                put(ws, r, 2 + i, v, hist_kind, fmt, bold=bold)
        if fcst_fn is not None:
            for y in FCST:
                put(ws, r, 2 + YRS.index(y), fcst_fn(y), 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        I[key] = r
        r += 1

    def isf(key, label, all_fn, note='', fmt=NUM, bold=False):
        """所有年份(含历史)均为公式的行"""
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        for y in YRS:
            put(ws, r, 2 + YRS.index(y), all_fn(y), 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        I[key] = r
        r += 1

    _rev_note = '=Revenue_Segments分部合计; 历史: 东财F10(' + '/'.join(f'{v/100:.1f}' for v in _rev_h) + '亿)'
    isf('rev', '营业收入', lambda y: f"={RS}!{YC[y]}{REV_TOT}", _rev_note, NUM, True)
    put(ws, r, 1, '  收入YoY', 't')
    for y in YRS[1:]:
        put(ws, r, 2 + YRS.index(y), f"={YC[y]}{I['rev']}/{YC[y-1]}{I['rev']}-1", 'f', PCT)
    _gyoy = f'{HIST[1]}A {_g1:+.1%}' + (f' / {H0}A {_g2:+.1%}' if _g2 is not None else '')
    put(ws, r, NOTE_COL, f'历史: {_gyoy}', 'g', size=9)
    I['rev_g'] = r; r += 1
    isrow('cost', '营业成本', _is_h['cost'], lambda y: f"={YC[y]}{I['rev']}-{RS}!{YC[y]}{GP_TOT}", '历史: 东财F10实际值; 预测=收入-分部毛利合计(占比四舍五入, 历史分部毛利与总额存微小口径差)')
    isf('gp', '毛利', lambda y: f"={YC[y]}{I['rev']}-{YC[y]}{I['cost']}", '', NUM, True)
    put(ws, r, 1, '  毛利率', 't')
    for y in YRS:
        put(ws, r, 2 + YRS.index(y), f"={YC[y]}{I['gp']}/{YC[y]}{I['rev']}", 'f', PCT)
    _gm_h = [(_rev_h[i] - _is_h['cost'][i]) / _rev_h[i] for i in range(NH)]
    put(ws, r, NOTE_COL, f'{HIST[0]}A {_gm_h[0]:.1%}→{H0}A {_gm_h[-1]:.1%}; 预测由分部毛利率加权', 'g', size=9)
    I['gm'] = r; r += 1
    isrow('taxadd', '税金及附加', _is_h['taxadd'], lambda y: f"={YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_taxadd']}", f"预测=收入×{a5('opex.tax_add_rate')[0]:.2%}")
    isrow('sale', '销售费用', _is_h['sale'], lambda y: f"={YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_sale']}", f"预测=收入×{a5('opex.sale_rate')[0]:.2%}")
    isrow('adm', '管理费用', _is_h['adm'], lambda y: f"={YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_adm']}", f"预测=收入×{a5('opex.adm_rate')[0]:.2%}")
    isrow('rd', '研发费用', _is_h['rd'], lambda y: f"={YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_rd']}", f"预测=收入×{a5('opex.rd_rate')[0]:.2%}")
    isrow('fin', '财务费用', _is_h['fin'], lambda y: f"={FINs}!{YC[y]}{F['finnet']}", '预测=FIN净财务费用(平均余额计息, 第4轮迭代展开输出)')
    isrow('othop', '其他经营损益(投资/公允/减值等)', _is_h['othop'], lambda y: f"={ASheet}!{AC[y]}${A['oth_op']}", f"历史为轧差项; 预测取{a5('opex.oth_op')[0]:.0f}百万")
    isf('ebit_op', '营业利润', lambda y: (f"={YC[y]}{I['gp']}-{YC[y]}{I['taxadd']}-{YC[y]}{I['sale']}"
        f"-{YC[y]}{I['adm']}-{YC[y]}{I['rd']}-{YC[y]}{I['fin']}+{YC[y]}{I['othop']}"), '', NUM, True)
    isrow('nonop', '营业外收支净额', _is_h['nonop'], lambda y: f"={ASheet}!{AC[y]}${A['nonop']}", f"{H0}A {_is_h['nonop'][-1]:.1f}百万")
    isf('tprofit', '利润总额', lambda y: f"={YC[y]}{I['ebit_op']}+{YC[y]}{I['nonop']}", '', NUM, True)
    isrow('taxexp', '所得税', _is_h['taxexp'], lambda y: f"={YC[y]}{I['tprofit']}*{ASheet}!{AC[y]}${A['tax']}", f"预测=利润总额×{a5('opex.tax_rate')[0]:.1%}")
    isf('np', '净利润', lambda y: f"={YC[y]}{I['tprofit']}-{YC[y]}{I['taxexp']}", '合并口径(含少数股东损益)', NUM, True)
    # A1: 归母≠合并 — 历史归母取config硬数(hist.is.np_p, 与hist.cf.np同为归母口径, 绿色);
    #     预测 归母净利润=合并净利润×(1-mi_share); mi_share=0的标的与合并口径数值恒等(np×(1-0)=np)
    _np_note = '历史: ' + '/'.join(f'{y}A {v/100:.1f}亿' for y, v in zip(HIST, _is_h['np_p']))
    isrow('np_p', '归母净利润', _is_h['np_p'],
          lambda y: f"={YC[y]}{I['np']}*(1-{ASheet}!{AC[y]}${A['mi_share']})",
          _np_note + ' (config硬数, 归母口径=hist.cf.np); 预测=净利润×(1-少数股东损益占比)', NUM, bold=True)
    isf('mi', '少数股东损益', lambda y: f"={YC[y]}{I['np']}-{YC[y]}{I['np_p']}",
        '=净利润-归母净利润; 历史为合并与归母口径差(倒挤), 预测=净利润×mi_share')
    put(ws, r, 1, '  归母净利率', 't')
    for y in YRS:
        put(ws, r, 2 + YRS.index(y), f"={YC[y]}{I['np_p']}/{YC[y]}{I['rev']}", 'f', PCT)
    put(ws, r, NOTE_COL, f"{H0}A {_is_h['np_p'][-1]/_rev_h[-1]:.1%}", 'g', size=9)
    I['npm'] = r; r += 1
    isf('eps', f'EPS({PER_SHARE_UNIT})', lambda y: f"={YC[y]}{I['np_p']}/{EQs}!$C${Q['shares']}", f'=归母/稀释股数(Equity_Roll页, 现={cfg.v("market.shares")/100:.2f}亿股)', PS)
    isf('ebit', 'EBIT (利润总额+财务费用)', lambda y: f"={YC[y]}{I['tprofit']}+{YC[y]}{I['fin']}", '供DCF页FCFF计算', NUM, True)
    ISs = 'IS'

    # ============================================================
    # BS 资产负债表
    # ============================================================
    # --- 历史尾差控制: 构建前已做物质性校验, 仅允许微小舍入差并入其他权益项目 ---
    _bs = _normalized_bs
    _A_K = ['cash', 'tfa', 'ar', 'pre', 'orec', 'inv', 'oca', 'ppe', 'rou', 'ia', 'gw', 'lpe', 'dta', 'oei', 'onca']
    _L_K = ['stl', 'ap', 'contract', 'staff', 'taxp', 'opay', 'cur1y', 'ocl', 'ltl', 'lease', 'oncl']
    _E_K = ['sc', 'cr', 'oci', 'sr', 're', 'oeq', 'mi']
    for plug in HISTORICAL_PLUGS:
        print(f'尾差修正(仅舍入): {plug["year"]}A 尾差 {plug["diff"]:+.4f} '
              f'(资产占比{plug["ratio"]:.2e}) 并入其他权益项目')

    _bs_notes = cfg.get('hist.notes') or {}
    ws = wb.create_sheet('BS')
    title_bar(ws, f'{NAME} — BS 资产负债表 ({HIST[0]}A-{F1}E)',
              f'单位: {UNIT} | 配平机制=FIN页revolver新增借款/现金sweep/理财承接(货币资金=最低现金); 配平差额检查行应恒=0')
    ws.column_dimensions['A'].width = 28
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 66
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注 / 来源', 't', bold=True)

    B = {}
    r = 4
    def brow(key, label, hist_vals, fcst_fn, note='', bold=False, fmt=NUM):
        nonlocal r
        B[key] = r  # 先登记行号, 供hold()自引用
        put(ws, r, 1, label, 't', bold=bold)
        for i, v in enumerate(hist_vals):
            put(ws, r, 2 + i, v, 'x', fmt, bold=bold)
        if fcst_fn is not None:
            for y in FCST:
                put(ws, r, 2 + YRS.index(y), fcst_fn(y), 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        r += 1

    def bsum(key, label, keys, note=''):
        nonlocal r
        put(ws, r, 1, label, 't', bold=True)
        for y in YRS:
            cl = YC[y]
            expr = '+'.join(f"{cl}{B[k]}" for k in keys)
            put(ws, r, 2 + YRS.index(y), f"={expr}", 'f', NUM, bold=True, fill=TOT_FILL)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        B[key] = r
        r += 1

    hold = lambda key: (lambda y: f"={YC[y-1]}{B[key]}")

    brow('cash', '货币资金', _bs['cash'], lambda y: f"={FINs}!{YC[y]}{F['cash1']}", '预测=FIN期末现金(=最低现金); 历史: 东财F10')
    brow('tfa', '交易性金融资产', _bs['tfa'], lambda y: f"={FINs}!{YC[y]}{F['tfa1']}", '预测=FIN理财期末(sweep溢出承接); 历史: 东财F10')
    brow('ar', '应收票据及应收账款', _bs['ar'],
         lambda y: f"={ASheet}!{AC[y]}${A['dso']}/365*{ISs}!{YC[y]}{I['rev']}", '预测=DSO/365×收入')
    brow('pre', '预付款项', _bs['pre'],
         lambda y: f"={ISs}!{YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_pre']}", f"预测=收入×{a5('working_capital.pre_rate')[0]:.1%}")
    brow('orec', '其他应收款', _bs['orec'], hold('orec'), '预测持平')
    brow('inv', '存货', _bs['inv'],
         lambda y: f"={ASheet}!{AC[y]}${A['dio']}/365*{ISs}!{YC[y]}{I['cost']}", '预测=DIO/365×营业成本')
    brow('oca', '应收款项融资及其他流动资产', _bs['oca'], hold('oca'), _bs_notes.get('oca', '含应收款项融资/其他流动资产, 历史为轧差, 预测持平'))
    bsum('tca', '流动资产合计', ['cash', 'tfa', 'ar', 'pre', 'orec', 'inv', 'oca'])
    brow('ppe', '固定资产及在建工程', _bs['ppe'],
         lambda y: f"={PPEs}!{YC[y]}{P['ppe1']}", f"预测=PP&E页期末合计; {H0}A固资{_fa_net0/100:.1f}亿+在建{_cip0/100:.1f}亿")
    brow('rou', '使用权资产', _bs['rou'], hold('rou'), '预测持平')
    brow('ia', '无形资产', _bs['ia'],
         lambda y: f"={SCH}!{YC[y]}{S['ia1']}", '预测=期初-摊销(Schedules)')
    brow('gw', '商誉', _bs['gw'], hold('gw'), '预测持平')
    brow('lpe', '长期待摊费用', _bs['lpe'], hold('lpe'), '预测持平')
    brow('dta', '递延所得税资产', _bs['dta'], hold('dta'), '预测持平')
    brow('oei', '其他权益工具投资', _bs['oei'], hold('oei'), f"预测持平; {H0}A末{_bs['oei'][-1]/100:.1f}亿已计入DCF的EV→Equity桥")
    brow('onca', '其他非流动资产', _bs['onca'], hold('onca'), _bs_notes.get('onca', '历史为轧差, 预测持平'))
    bsum('tnca', '非流动资产合计', ['ppe', 'rou', 'ia', 'gw', 'lpe', 'dta', 'oei', 'onca'])
    put(ws, r, 1, '资产总计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{B['tca']}+{cl}{B['tnca']}", 'f', NUM, bold=True, fill=TOT_FILL)
    _ta_h0 = sum(_bs[k][-1] for k in _A_K)
    put(ws, r, NOTE_COL, f'{H0}A {_ta_h0/100:.1f}亿', 'g', size=9)
    B['ta'] = r; r += 1
    brow('stl', '短期借款', _bs['stl'], lambda y: f"={FINs}!{YC[y]}{F['st_t1']}", '预测=FIN revolver期末(缺口新增-sweep偿还)')
    brow('ap', '应付票据及应付账款', _bs['ap'],
         lambda y: f"={ASheet}!{AC[y]}${A['dpo']}/365*{ISs}!{YC[y]}{I['cost']}", '预测=DPO/365×营业成本')
    brow('contract', '合同负债', _bs['contract'], hold('contract'), '预测持平')
    brow('staff', '应付职工薪酬', _bs['staff'],
         lambda y: f"={ISs}!{YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_staff']}", f"预测=收入×{a5('working_capital.staff_rate')[0]:.1%}")
    brow('taxp', '应交税费', _bs['taxp'],
         lambda y: f"={ISs}!{YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_taxp']}", f"预测=收入×{a5('working_capital.taxp_rate')[0]:.1%}")
    brow('opay', '其他应付款', _bs['opay'], hold('opay'), '预测持平')
    brow('cur1y', '一年内到期的非流动负债', _bs['cur1y'], lambda y: f"={FINs}!{YC[y]}{F['cur_t1']}", '预测=FIN期末(计划还款摊还)')
    brow('ocl', '其他流动负债', _bs['ocl'], hold('ocl'), '历史为轧差, 预测持平')
    bsum('tcl', '流动负债合计', ['stl', 'ap', 'contract', 'staff', 'taxp', 'opay', 'cur1y', 'ocl'])
    brow('ltl', '长期借款', _bs['ltl'], lambda y: f"={FINs}!{YC[y]}{F['lt_t1']}", '预测=FIN期末(计划还款+sweep)')
    brow('lease', '租赁负债', _bs['lease'], lambda y: f"={FINs}!{YC[y]}{F['le_t1']}", '预测=FIN期末(计划还款摊还)')
    brow('oncl', '其他非流动负债', _bs['oncl'], hold('oncl'), '含递延收益/递延所得税负债, 历史轧差, 预测持平')
    bsum('tncl', '非流动负债合计', ['ltl', 'lease', 'oncl'])
    put(ws, r, 1, '负债合计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{B['tcl']}+{cl}{B['tncl']}", 'f', NUM, bold=True, fill=TOT_FILL)
    B['tl'] = r; r += 1
    brow('sc', '股本', _bs['sc'], lambda y: f"={EQs}!{YC[y]}{Q['sc']}", '预测=Equity_Roll(持平)')
    brow('cr', '资本公积', _bs['cr'], lambda y: f"={EQs}!{YC[y]}{Q['cr']}", '预测=Equity_Roll(持平)')
    brow('ts', '减:库存股', _bs['ts'], lambda y: f"={EQs}!{YC[y]}{Q['ts']}", '预测=Equity_Roll(持平)')
    brow('oci', '其他综合收益', _bs['oci'], lambda y: f"={EQs}!{YC[y]}{Q['oci']}", '预测=Equity_Roll(持平)')
    brow('sr', '盈余公积', _bs['sr'], lambda y: f"={EQs}!{YC[y]}{Q['sr']}", f'预测=Equity_Roll(期初+归母×{_sr_rt:.0%})')
    brow('re', '未分配利润', _bs['re'],
         lambda y: f"={EQs}!{YC[y]}{Q['re']}", '预测=Equity_Roll(期初+归母-分红-盈余公积计提)')
    brow('oeq', '其他权益项目(轧差)', _bs['oeq'], hold('oeq'), '含专项储备及换算尾差, 预测持平; 历史尾差已并入本行清零')
    put(ws, r, 1, '归母所有者权益合计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y),
            f"={cl}{B['sc']}+{cl}{B['cr']}-{cl}{B['ts']}+{cl}{B['oci']}+{cl}{B['sr']}+{cl}{B['re']}+{cl}{B['oeq']}",
            'f', NUM, bold=True, fill=TOT_FILL)
    B['te_p'] = r; r += 1
    brow('mi', '少数股东权益', _bs['mi'], hold('mi'), _bs_notes.get('mi', '预测持平'))
    put(ws, r, 1, '所有者权益合计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{B['te_p']}+{cl}{B['mi']}", 'f', NUM, bold=True)
    B['te'] = r; r += 1
    put(ws, r, 1, '负债和所有者权益总计', 't', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{B['tl']}+{cl}{B['te']}", 'f', NUM, bold=True, fill=TOT_FILL)
    B['tle'] = r; r += 1
    put(ws, r, 1, '配平差额检查(资产-负债权益)', 'w', bold=True)
    for y in YRS:
        cl = YC[y]
        put(ws, r, 2 + YRS.index(y), f"={cl}{B['ta']}-{cl}{B['tle']}", 'f', NUM, bold=True, fill=CHK_FILL)
    put(ws, r, NOTE_COL, '应恒=0(容差0.01); Checks页汇总校验', 'g', size=9)
    B['chk'] = r; r += 1
    BSs = 'BS'

    # ============================================================
    # CF 现金流量表 (间接法, 利息重分类至筹资, 投资含理财净变动)
    # ============================================================
    _cf_h = cfg.raw['hist']['cf']
    ws = wb.create_sheet('CF')
    title_bar(ws, f'{NAME} — CF 现金流量表 (间接法, {HIST[0]}A-{F1}E)',
              f'单位: {UNIT} | 利息支出经营加回、筹资列"偿付利息"; 投资含"理财净变动"; 期末货币资金=FIN期末现金=BS货币资金')
    ws.column_dimensions['A'].width = 30
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 66
    ws.freeze_panes = 'B4'
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注 / 来源', 't', bold=True)

    C = {}
    r = 4
    def crow(key, label, hist_vals, fcst_fn, note='', bold=False):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        if hist_vals is not None:
            for i, v in enumerate(hist_vals):
                if v is None:
                    put(ws, r, 2 + i, '—', 'g', align='center')
                else:
                    put(ws, r, 2 + i, v, 'x', NUM, bold=bold)
            if fcst_fn is not None:
                for y in FCST:
                    put(ws, r, 2 + YRS.index(y), fcst_fn(y), 'f', NUM, bold=bold)
        else:
            for i in range(NH):
                put(ws, r, 2 + i, '—', 'g', align='center')
            if fcst_fn is not None:
                for y in FCST:
                    put(ws, r, 2 + YRS.index(y), fcst_fn(y), 'f', NUM, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        C[key] = r
        r += 1

    crow('np', '归母净利润', _cf_h['np'], lambda y: f"={ISs}!{YC[y]}{I['np_p']}", '=IS归母净利')
    crow('da', '加: 折旧与摊销(D&A)', _cf_h['da'], lambda y: f"={PPEs}!{YC[y]}{P['da']}", '历史=固定资产折旧+无形/长待摊摊销; 预测=PP&E页合计')
    crow('intadd', '加: 利息支出(重分类至筹资)', None, lambda y: f"={FINs}!{YC[y]}{F['intexp']}", '利润表中利息属筹资活动, 经营现金流加回; 历史未拆(列—)')
    crow('dnwc', '减: 营运资本变动(增加为负)', None, lambda y: f"=-{SCH}!{YC[y]}{S['dnwc']}", '预测=-ΔNWC(Schedules)')
    crow('ocf', '经营活动现金流量净额', _cf_h['ocf'],
         lambda y: f"={YC[y]}{C['np']}+{YC[y]}{C['da']}+{YC[y]}{C['intadd']}+{YC[y]}{C['dnwc']}", '历史为年报实际值(含利息, 口径与预测略有差异)', True)
    crow('capex', '资本开支(购建长期资产)', _cf_h['capex'], lambda y: f"=-{PPEs}!{YC[y]}{P['capex']}", '预测=-收入×Capex占比(PP&E页)')
    crow('disp', '加: 处置固定资产回收(账面值)', None, lambda y: f"={PPEs}!{YC[y]}{P['disp']}", '预测=PP&E页处置额(按账面值回收, 无损益); 历史含在投资净额内(列—)')
    crow('wmm', '理财净变动(交易性金融资产)', None, lambda y: f"=-{FINs}!{YC[y]}{F['dtfa']}", '=-FIN理财净变动(sweep溢出购买为流出); 历史含在投资净额内(列—)')
    crow('icf', '投资活动现金流量净额', _cf_h['icf'], lambda y: f"={YC[y]}{C['capex']}+{YC[y]}{C['disp']}+{YC[y]}{C['wmm']}", '历史为年报实际值; 预测=资本开支+处置回收+理财净变动', True)
    crow('ddebt', '债务净变动(新增为正)', None, lambda y: f"={FINs}!{YC[y]}{F['ddebt']}", '=FIN四档债务期末-期初(含revolver新增/sweep偿还)')
    crow('intpay', '偿付利息', None, lambda y: f"=-{FINs}!{YC[y]}{F['intexp']}", '利息支出列筹资活动')
    crow('div', '股利支付', None, lambda y: f"=-{SCH}!{YC[y]}{S['div']}", f'=-上年归母×{_payout0:.0%}')
    crow('fcf', '筹资活动现金流量净额', _cf_h['fcf'],
         lambda y: f"={YC[y]}{C['ddebt']}+{YC[y]}{C['intpay']}+{YC[y]}{C['div']}", '历史为年报实际值; 预测=债务净变动-利息-股利', True)
    crow('dcash', '货币资金净增加额', None,
         lambda y: f"={YC[y]}{C['ocf']}+{YC[y]}{C['icf']}+{YC[y]}{C['fcf']}", '', True)
    put(ws, r, 1, '期初货币资金', 't')
    for i in range(NH):
        put(ws, r, 2 + i, '—', 'g', align='center')
    for y in FCST:
        if y == F0:
            put(ws, r, 2 + YRS.index(y), f"={BSs}!{YC[H0]}{B['cash']}", 'f', NUM)
        else:
            put(ws, r, 2 + YRS.index(y), f"={YC[y-1]}{r + 1}", 'f', NUM)
    put(ws, r, NOTE_COL, f'{F0}期初={H0}A末货币资金', 'g', size=9)
    C['cash0'] = r; r += 1
    crow('cash1', '期末货币资金(=BS货币资金)', _cf_h['cash1'],
         lambda y: f"={YC[y]}{C['cash0']}+{YC[y]}{C['dcash']}", '历史=BS货币资金; 预测=期初+净增加=FIN期末现金(最低现金)', True)
    CFs = 'CF'

    # ============================================================
    # 回填 Schedules / PPE / FIN / Equity_Roll 全部公式
    # ============================================================
    wS = wb[SCH]
    def srow_fill(w, rows, key, hist_fn, fcst_fn, fmt=NUM):
        """H3: 显式传worksheet(w)与行号dict(rows), 不再按dict身份反查所属sheet"""
        rr = rows[key]
        if hist_fn:
            for y in HIST:
                put(w, rr, 2 + YRS.index(y), hist_fn(y), 'f', fmt)
        for y in FCST:
            put(w, rr, 2 + YRS.index(y), fcst_fn(y), 'f', fmt)

    # --- Schedules: NWC 引用BS ---
    for key, bk in [('ar', 'ar'), ('pre', 'pre'), ('orec', 'orec'), ('inv', 'inv'),
                    ('ap', 'ap'), ('contract', 'contract'), ('staff', 'staff'), ('taxp', 'taxp'), ('opay', 'opay')]:
        srow_fill(wS, S, key, lambda y, bk=bk: f"={BSs}!{YC[y]}{B[bk]}",
                  lambda y, bk=bk: f"={BSs}!{YC[y]}{B[bk]}")
    srow_fill(wS, S, 'nwc_a', lambda y: f"=SUM({YC[y]}{S['ar']}:{YC[y]}{S['inv']})",
              lambda y: f"=SUM({YC[y]}{S['ar']}:{YC[y]}{S['inv']})")
    srow_fill(wS, S, 'nwc_l', lambda y: f"=SUM({YC[y]}{S['ap']}:{YC[y]}{S['opay']})",
              lambda y: f"=SUM({YC[y]}{S['ap']}:{YC[y]}{S['opay']})")
    srow_fill(wS, S, 'nwc', lambda y: f"={YC[y]}{S['nwc_a']}-{YC[y]}{S['nwc_l']}",
              lambda y: f"={YC[y]}{S['nwc_a']}-{YC[y]}{S['nwc_l']}")
    srow_fill(wS, S, 'dnwc', None, lambda y: f"={YC[y]}{S['nwc']}-{YC[y-1]}{S['nwc']}")
    # 无形资产
    srow_fill(wS, S, 'ia0', None, lambda y: (f"={BSs}!{YC[H0]}{B['ia']}" if y == F0 else f"={YC[y-1]}{S['ia1']}"))
    srow_fill(wS, S, 'amort', None, lambda y: f"={YC[y]}{S['ia0']}*{ASheet}!$C${A['rt_amort']}")
    srow_fill(wS, S, 'ia1', None, lambda y: f"={YC[y]}{S['ia0']}-{YC[y]}{S['amort']}")
    # 股利
    srow_fill(wS, S, 'div', None, lambda y: f"={ISs}!{YC[y-1]}{I['np_p']}*{ASheet}!{AC[y]}${A['payout']}")

    # --- PPE ---
    # fa0/cip0 的F0期初为绿色输入(H0实际拆分), 以后年度=上年期末
    wP = wb[PPEs]
    put(wP, P['fa0'], 2 + YRS.index(F0), _fa_net0, 'x', NUM)
    for y in FCST[1:]:
        put(wP, P['fa0'], 2 + YRS.index(y), f"={YC[y-1]}{P['fa1']}", 'f', NUM)
    put(wP, P['fa0'], NOTE_COL, f'{F0}期初={H0}A固定资产{_fa_net0/100:.1f}亿(绿色输入, {H0}年报)', 'g', size=9)
    put(wP, P['cip0'], 2 + YRS.index(F0), _cip0, 'x', NUM)
    for y in FCST[1:]:
        put(wP, P['cip0'], 2 + YRS.index(y), f"={YC[y-1]}{P['cip1']}", 'f', NUM)
    put(wP, P['cip0'], NOTE_COL, f'{F0}期初={H0}A在建工程{_cip0/100:.1f}亿(绿色输入); {_fa_net0:.1f}+{_cip0:.1f}={_fa_net0+_cip0:.1f}与BS一致', 'g', size=9)
    srow_fill(wP, P, 'capex', None, lambda y: f"={ISs}!{YC[y]}{I['rev']}*{ASheet}!{AC[y]}${A['rt_capex']}")
    srow_fill(wP, P, 'trans', None, lambda y: f"=({YC[y]}{P['cip0']}+{YC[y]}{P['capex']})*{ASheet}!$C${A['rt_trans']}")
    srow_fill(wP, P, 'dep', None, lambda y: f"={YC[y]}{P['fa0']}*{ASheet}!$C${A['rt_dep']}+{YC[y]}{P['trans']}*{ASheet}!$C${A['rt_dep_new']}")
    srow_fill(wP, P, 'disp', None, lambda y: f"={YC[y]}{P['fa0']}*{ASheet}!$C${A['rt_disp']}")
    srow_fill(wP, P, 'fa1', None, lambda y: f"={YC[y]}{P['fa0']}+{YC[y]}{P['trans']}-{YC[y]}{P['dep']}-{YC[y]}{P['disp']}")
    srow_fill(wP, P, 'cip1', None, lambda y: f"={YC[y]}{P['cip0']}+{YC[y]}{P['capex']}-{YC[y]}{P['trans']}")
    srow_fill(wP, P, 'ppe1', lambda y: f"={BSs}!{YC[y]}{B['ppe']}",
              lambda y: f"={YC[y]}{P['fa1']}+{YC[y]}{P['cip1']}")
    srow_fill(wP, P, 'da', lambda y: f"={CFs}!{YC[y]}{C['da']}",
              lambda y: f"={YC[y]}{P['dep']}+{SCH}!{YC[y]}{S['amort']}")
    srow_fill(wP, P, 'dep_rt', None, lambda y: f"={YC[y]}{P['dep']}/{YC[y]}{P['fa0']}", PCT)
    srow_fill(wP, P, 'dep_yrs', None, lambda y: f"={YC[y]}{P['fa0']}/{YC[y]}{P['dep']}", '0.0"年"')

    # --- FIN ---
    wF = wb[FINs]
    # H0列锚点: 各期末行引用BS H0实际值
    put(wF, F['st_t1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['stl']}", 'f', NUM)
    put(wF, F['cur_t1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['cur1y']}", 'f', NUM)
    put(wF, F['lt_t1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['ltl']}", 'f', NUM)
    put(wF, F['le_t1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['lease']}", 'f', NUM)
    put(wF, F['tfa1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['tfa']}", 'f', NUM)
    put(wF, F['cash1'], 2 + YRS.index(H0), f"={BSs}!{YC[H0]}{B['cash']}", 'f', NUM)
    put(wF, F['dtot1'], 2 + YRS.index(H0), f"={YC[H0]}{F['st_t1']}+{YC[H0]}{F['cur_t1']}+{YC[H0]}{F['lt_t1']}+{YC[H0]}{F['le_t1']}", 'f', NUM)
    put(wF, F['dtot0'], 2 + YRS.index(H0), f"={YC[H0]}{F['dtot1']}", 'f', NUM)
    def fin_fill(key, fn, fmt=NUM):
        for y in FCST:
            put(wF, F[key], 2 + YRS.index(y), fn(y), 'f', fmt)
    def prevcl(y):
        return YC[y - 1]
    fin_fill('cash0', lambda y: f"={prevcl(y)}{F['cash1']}")
    fin_fill('ocf', lambda y: f"={CFs}!{YC[y]}{C['ocf']}")
    fin_fill('icf', lambda y: f"=-{PPEs}!{YC[y]}{P['capex']}+{PPEs}!{YC[y]}{P['disp']}")
    fin_fill('div', lambda y: f"=-{SCH}!{YC[y]}{S['div']}")
    fin_fill('intpay', lambda y: f"=-{YC[y]}{F['intexp']}")
    fin_fill('rep', lambda y: f"=-({YC[y]}{F['st_rep_a']}+{YC[y]}{F['cur_rep_a']}+{YC[y]}{F['lt_rep_a']}+{YC[y]}{F['le_rep_a']})")
    fin_fill('cpre', lambda y: (f"={YC[y]}{F['cash0']}+{YC[y]}{F['ocf']}+{YC[y]}{F['icf']}"
                                f"+{YC[y]}{F['div']}+{YC[y]}{F['intpay']}+{YC[y]}{F['rep']}"))
    fin_fill('mincash', lambda y: f"={ISs}!{YC[y]}{I['rev']}*{ASheet}!$C${A['min_cash_pct']}")
    fin_fill('new', lambda y: f"=MAX(0,{YC[y]}{F['mincash']}-{YC[y]}{F['cpre']})")
    fin_fill('avail', lambda y: f"=MAX(0,{YC[y]}{F['cpre']}+{YC[y]}{F['new']}-{YC[y]}{F['mincash']})")
    # 四档债务
    fin_fill('st_t0', lambda y: f"={prevcl(y)}{F['st_t1']}")
    fin_fill('st_rep', lambda y: f"={ASheet}!{AC[y]}${A['rep_st']}")
    fin_fill('st_rep_a', lambda y: f"=MIN({YC[y]}{F['st_rep']},{YC[y]}{F['st_t0']})")
    fin_fill('st_new', lambda y: f"={YC[y]}{F['new']}")
    fin_fill('st_pre', lambda y: f"={YC[y]}{F['st_t0']}-{YC[y]}{F['st_rep_a']}+{YC[y]}{F['st_new']}")
    fin_fill('st_sw', lambda y: f"=MIN({YC[y]}{F['avail']},{YC[y]}{F['st_pre']})")
    fin_fill('st_t1', lambda y: f"={YC[y]}{F['st_pre']}-{YC[y]}{F['st_sw']}")
    fin_fill('st_avg', lambda y: f"=({YC[y]}{F['st_t0']}+{YC[y]}{F['st_t1']})/2")
    fin_fill('st_int', lambda y: f"={YC[y]}{F['st_avg']}*{ASheet}!$C${A['rate_st']}")
    fin_fill('cur_t0', lambda y: f"={prevcl(y)}{F['cur_t1']}")
    fin_fill('cur_rep', lambda y: f"={ASheet}!{AC[y]}${A['rep_cur']}")
    fin_fill('cur_rep_a', lambda y: f"=MIN({YC[y]}{F['cur_rep']},{YC[y]}{F['cur_t0']})")
    fin_fill('cur_t1', lambda y: f"={YC[y]}{F['cur_t0']}-{YC[y]}{F['cur_rep_a']}")
    fin_fill('cur_avg', lambda y: f"=({YC[y]}{F['cur_t0']}+{YC[y]}{F['cur_t1']})/2")
    fin_fill('cur_int', lambda y: f"={YC[y]}{F['cur_avg']}*{ASheet}!$C${A['rate_cur']}")
    fin_fill('lt_t0', lambda y: f"={prevcl(y)}{F['lt_t1']}")
    fin_fill('lt_rep', lambda y: f"={ASheet}!{AC[y]}${A['rep_lt']}")
    fin_fill('lt_rep_a', lambda y: f"=MIN({YC[y]}{F['lt_rep']},{YC[y]}{F['lt_t0']})")
    fin_fill('lt_pre', lambda y: f"={YC[y]}{F['lt_t0']}-{YC[y]}{F['lt_rep_a']}")
    fin_fill('lt_sw', lambda y: f"=MIN({YC[y]}{F['avail']}-{YC[y]}{F['st_sw']},{YC[y]}{F['lt_pre']})")
    fin_fill('lt_t1', lambda y: f"={YC[y]}{F['lt_pre']}-{YC[y]}{F['lt_sw']}")
    fin_fill('lt_avg', lambda y: f"=({YC[y]}{F['lt_t0']}+{YC[y]}{F['lt_t1']})/2")
    fin_fill('lt_int', lambda y: f"={YC[y]}{F['lt_avg']}*{ASheet}!$C${A['rate_lt']}")
    fin_fill('le_t0', lambda y: f"={prevcl(y)}{F['le_t1']}")
    fin_fill('le_rep', lambda y: f"={ASheet}!{AC[y]}${A['rep_lease']}")
    fin_fill('le_rep_a', lambda y: f"=MIN({YC[y]}{F['le_rep']},{YC[y]}{F['le_t0']})")
    fin_fill('le_t1', lambda y: f"={YC[y]}{F['le_t0']}-{YC[y]}{F['le_rep_a']}")
    fin_fill('le_avg', lambda y: f"=({YC[y]}{F['le_t0']}+{YC[y]}{F['le_t1']})/2")
    fin_fill('le_int', lambda y: f"={YC[y]}{F['le_avg']}*{ASheet}!$C${A['rate_lease']}")
    # 汇总
    fin_fill('dtot0', lambda y: f"={YC[y]}{F['st_t0']}+{YC[y]}{F['cur_t0']}+{YC[y]}{F['lt_t0']}+{YC[y]}{F['le_t0']}")
    fin_fill('dtot1', lambda y: f"={YC[y]}{F['st_t1']}+{YC[y]}{F['cur_t1']}+{YC[y]}{F['lt_t1']}+{YC[y]}{F['le_t1']}")
    fin_fill('ddebt', lambda y: f"={YC[y]}{F['dtot1']}-{YC[y]}{F['dtot0']}")
    fin_fill('sw_tot', lambda y: f"={YC[y]}{F['st_sw']}+{YC[y]}{F['lt_sw']}")
    fin_fill('overflow', lambda y: f"={YC[y]}{F['avail']}-{YC[y]}{F['sw_tot']}")
    fin_fill('tfa0', lambda y: f"={prevcl(y)}{F['tfa1']}")
    fin_fill('tfa1', lambda y: f"={YC[y]}{F['tfa0']}+{YC[y]}{F['overflow']}")
    fin_fill('dtfa', lambda y: f"={YC[y]}{F['tfa1']}-{YC[y]}{F['tfa0']}")
    fin_fill('cash1', lambda y: f"={YC[y]}{F['cpre']}+{YC[y]}{F['new']}-{YC[y]}{F['sw_tot']}-{YC[y]}{F['overflow']}")
    fin_fill('finoth', lambda y: f"={ASheet}!$C${A['fin_oth']}")
    # --- 迭代展开块公式 (循环链显式展开; 引用均为无环上游: IS中fin之前的行/Assumptions/PPE/Schedules) ---
    def itf(k, key, fn, fmt=NUM):
        for y in FCST:
            put(wF, IT[k][key], 2 + YRS.index(y), fn(y), 'f', fmt)

    for k in range(1, 5):
        itf(k, 'intexp_in', lambda y, k=k: (
            f"={YC[y]}{F['st_t0']}*{ASheet}!$C${A['rate_st']}+{YC[y]}{F['cur_t0']}*{ASheet}!$C${A['rate_cur']}"
            f"+{YC[y]}{F['lt_t0']}*{ASheet}!$C${A['rate_lt']}+{YC[y]}{F['le_t0']}*{ASheet}!$C${A['rate_lease']}"
            if k == 1 else f"={YC[y]}{IT[k-1]['intexp_out']}"))
        itf(k, 'intinc_in', lambda y, k=k: (
            f"={YC[y]}{F['cash0']}*{ASheet}!$C${A['rt_cash']}" if k == 1 else f"={YC[y]}{IT[k-1]['intinc_out']}"))
        itf(k, 'finnet_in', lambda y, k=k: f"={YC[y]}{IT[k]['intexp_in']}-{YC[y]}{IT[k]['intinc_in']}+{YC[y]}{F['finoth']}")
        # 归母净利重算: (毛利-税金附加-三费+其他经营损益+营业外-财务费用净额)×(1-t)×(1-mi_share);
        # 只引用IS中fin之前的无环行; A1: ×(1-mi_share)与IS归母口径一致, mi_share=0时数值恒等
        itf(k, 'np', lambda y, k=k: (
            f"=({ISs}!{YC[y]}{I['gp']}-{ISs}!{YC[y]}{I['taxadd']}-{ISs}!{YC[y]}{I['sale']}-{ISs}!{YC[y]}{I['adm']}"
            f"-{ISs}!{YC[y]}{I['rd']}+{ISs}!{YC[y]}{I['othop']}+{ISs}!{YC[y]}{I['nonop']}-{YC[y]}{IT[k]['finnet_in']})"
            f"*(1-{ASheet}!{AC[y]}${A['tax']})*(1-{ASheet}!{AC[y]}${A['mi_share']})"))
        itf(k, 'ocf', lambda y, k=k: f"={YC[y]}{IT[k]['np']}+{PPEs}!{YC[y]}{P['da']}+{YC[y]}{IT[k]['intexp_in']}-{SCH}!{YC[y]}{S['dnwc']}")
        itf(k, 'cpre', lambda y, k=k: (f"={YC[y]}{F['cash0']}+{YC[y]}{IT[k]['ocf']}+{YC[y]}{F['icf']}"
                                       f"+{YC[y]}{F['div']}-{YC[y]}{IT[k]['intexp_in']}+{YC[y]}{F['rep']}"))
        itf(k, 'new', lambda y, k=k: f"=MAX(0,{YC[y]}{F['mincash']}-{YC[y]}{IT[k]['cpre']})")
        itf(k, 'avail', lambda y, k=k: f"=MAX(0,{YC[y]}{IT[k]['cpre']}+{YC[y]}{IT[k]['new']}-{YC[y]}{F['mincash']})")
        itf(k, 'st1', lambda y, k=k: (f"=({YC[y]}{F['st_t0']}-{YC[y]}{F['st_rep_a']}+{YC[y]}{IT[k]['new']})"
                                      f"-MIN({YC[y]}{IT[k]['avail']},{YC[y]}{F['st_t0']}-{YC[y]}{F['st_rep_a']}+{YC[y]}{IT[k]['new']})"))
        itf(k, 'lt1', lambda y, k=k: (f"=({YC[y]}{F['lt_t0']}-{YC[y]}{F['lt_rep_a']})"
                                      f"-MIN({YC[y]}{IT[k]['avail']}-({YC[y]}{F['st_t0']}-{YC[y]}{F['st_rep_a']}+{YC[y]}{IT[k]['new']}-{YC[y]}{IT[k]['st1']}),"
                                      f"{YC[y]}{F['lt_t0']}-{YC[y]}{F['lt_rep_a']})"))
        itf(k, 'intexp_out', lambda y, k=k: (
            f"=({YC[y]}{F['st_t0']}+{YC[y]}{IT[k]['st1']})/2*{ASheet}!$C${A['rate_st']}"
            f"+({YC[y]}{F['cur_t0']}+{YC[y]}{F['cur_t1']})/2*{ASheet}!$C${A['rate_cur']}"
            f"+({YC[y]}{F['lt_t0']}+{YC[y]}{IT[k]['lt1']})/2*{ASheet}!$C${A['rate_lt']}"
            f"+({YC[y]}{F['le_t0']}+{YC[y]}{F['le_t1']})/2*{ASheet}!$C${A['rate_lease']}"))
        itf(k, 'cash1', lambda y, k=k: f"={YC[y]}{IT[k]['cpre']}+{YC[y]}{IT[k]['new']}-{YC[y]}{IT[k]['avail']}")
        itf(k, 'intinc_out', lambda y, k=k: f"=({YC[y]}{F['cash0']}+{YC[y]}{IT[k]['cash1']})/2*{ASheet}!$C${A['rt_cash']}")
        itf(k, 'finnet_out', lambda y, k=k: f"={YC[y]}{IT[k]['intexp_out']}-{YC[y]}{IT[k]['intinc_out']}+{YC[y]}{F['finoth']}")
        itf(k, 'resid', lambda y, k=k: f"=ABS({YC[y]}{IT[k]['intexp_out']}-{YC[y]}{IT[k]['intexp_in']})+ABS({YC[y]}{IT[k]['intinc_out']}-{YC[y]}{IT[k]['intinc_in']})", '0.0000')
    # 末轮(iter4)输出回填汇总行: IS/CF/BS全簿引用即第4轮迭代值
    for y in FCST:
        cl = YC[y]; col = 2 + YRS.index(y)
        put(wF, F['intexp'], col, f"={cl}{IT[4]['intexp_out']}", 'f', NUM)
        put(wF, F['intinc'], col, f"={cl}{IT[4]['intinc_out']}", 'f', NUM)
        put(wF, F['finnet'], col, f"={cl}{IT[4]['finnet_out']}", 'f', NUM)
    # 收敛残差行: 采用值(第4轮) vs 按期末余额重算隐含值(=再迭代一轮), <0.01即收敛
    for y in FCST:
        cl = YC[y]; col = 2 + YRS.index(y)
        put(wF, F['resid_int'], col,
            f"=ABS({cl}{F['intexp']}-({cl}{F['st_int']}+{cl}{F['cur_int']}+{cl}{F['lt_int']}+{cl}{F['le_int']}))", 'f', '0.0000')
        put(wF, F['resid_inc'], col,
            f"=ABS({cl}{F['intinc']}-({cl}{F['cash0']}+{cl}{F['cash1']})/2*{ASheet}!$C${A['rt_cash']})", 'f', '0.0000')
        put(wF, F['resid_max'], col, f"=MAX({cl}{F['resid_int']},{cl}{F['resid_inc']})", 'f', '0.0000', bold=True, fill=CHK_FILL)
        put(wF, F['resid_ok'], col, f"={cl}{F['resid_max']}<0.01", 'f', 'General', bold=True, align='center')

    # --- Equity_Roll ---
    wQ = wb[EQs]
    eq_hist = {'sc': _bs['sc'], 'cr': _bs['cr'], 'ts': _bs['ts'], 'oci': _bs['oci'],
               'sr': _bs['sr'], 're': _bs['re'], 'oeq': _bs['oeq'], 'mi': _bs['mi']}
    for k, vals in eq_hist.items():
        for i, v in enumerate(vals):
            put(wQ, Q[k], 2 + i, v, 'x', NUM)
    def eq_fill(key, fn, years=FCST):
        for y in years:
            put(wQ, Q[key], 2 + YRS.index(y), fn(y), 'f', NUM)
    eq_fill('sc', lambda y: f"={YC[y-1]}{Q['sc']}")
    eq_fill('cr', lambda y: f"={YC[y-1]}{Q['cr']}")
    eq_fill('ts', lambda y: f"={YC[y-1]}{Q['ts']}")
    eq_fill('oci', lambda y: f"={YC[y-1]}{Q['oci']}")
    eq_fill('sr_acc', lambda y: f"={ISs}!{YC[y]}{I['np_p']}*{ASheet}!$C${A['sr_rate']}")
    eq_fill('sr', lambda y: f"={YC[y-1]}{Q['sr']}+{YC[y]}{Q['sr_acc']}")
    eq_fill('div', lambda y: f"={SCH}!{YC[y]}{S['div']}")
    eq_fill('re', lambda y: f"={YC[y-1]}{Q['re']}+{ISs}!{YC[y]}{I['np_p']}-{YC[y]}{Q['div']}-{YC[y]}{Q['sr_acc']}")
    eq_fill('oeq', lambda y: f"={YC[y-1]}{Q['oeq']}")
    eq_fill('mi', lambda y: f"={YC[y-1]}{Q['mi']}")
    for y in YRS:
        cl = YC[y]
        put(wQ, Q['te_p'], 2 + YRS.index(y),
            f"={cl}{Q['sc']}+{cl}{Q['cr']}-{cl}{Q['ts']}+{cl}{Q['oci']}+{cl}{Q['sr']}+{cl}{Q['re']}+{cl}{Q['oeq']}",
            'f', NUM, bold=True, fill=TOT_FILL)
        put(wQ, Q['te'], 2 + YRS.index(y), f"={cl}{Q['te_p']}+{cl}{Q['mi']}", 'f', NUM, bold=True)
    print('IS/BS/CF/Schedules/PPE/FIN/Equity_Roll done')

    # ============================================================
    # DCF (FCFF, 年中折现, Gordon终值)
    # ============================================================
    ws = wb.create_sheet('DCF')
    title_bar(ws, f'{NAME} — DCF 绝对估值 (FCFF, 显性期{F0}-{F1} + Gordon终值)',
              f'单位: {UNIT} | 估值基准日{VAL_DATE}, 采用年中折现(mid-year convention: t=0.5/1.5/.../{NF-0.5}); WACC/g引用Assumptions(采用值=IF(override为空,计算值,override))')
    ws.column_dimensions['A'].width = 32
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 60
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in FCST:
        put(ws, 3, 2 + YRS.index(y), f'{y}E', 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    D = {}
    r = 4
    def drow(key, label, fn, fmt=NUM, bold=False, note=''):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        for y in FCST:
            put(ws, r, 2 + YRS.index(y), fn(y), 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        D[key] = r
        r += 1

    WACC_REF = f"{ASheet}!$C${A['wacc']}"
    TG_REF = f"{ASheet}!$C${A['tg']}"
    SH_REF = f"{EQs}!$C${Q['shares']}"

    drow('ebit', 'EBIT', lambda y: f"={ISs}!{YC[y]}{I['ebit']}", NUM, False, '=IS: 利润总额+财务费用')
    drow('t', '税率', lambda y: f"={ASheet}!{AC[y]}${A['tax']}", PCT)
    drow('nopat', 'NOPAT = EBIT×(1-t)', lambda y: f"={YC[y]}{D['ebit']}*(1-{YC[y]}{D['t']})", NUM, True)
    drow('da', '加: D&A', lambda y: f"={PPEs}!{YC[y]}{P['da']}")
    drow('dnwc', '减: ΔNWC', lambda y: f"={SCH}!{YC[y]}{S['dnwc']}", NUM, False, 'NWC增加为现金流出')
    drow('capex', '减: Capex', lambda y: f"={PPEs}!{YC[y]}{P['capex']}")
    drow('fcff', 'FCFF', lambda y: f"={YC[y]}{D['nopat']}+{YC[y]}{D['da']}-{YC[y]}{D['dnwc']}-{YC[y]}{D['capex']}", NUM, True)
    drow('t_idx', '折现年序(年中)', lambda y: 0.5 + FCST.index(y), '0.0', False, f'年中折现(mid-year); 基准日{VAL_DATE}, 近似年中收到现金流')
    drow('df', '折现因子', lambda y: f"=1/(1+{WACC_REF})^{YC[y]}{D['t_idx']}", '0.000')
    drow('pv', 'PV(FCFF)', lambda y: f"={YC[y]}{D['fcff']}*{YC[y]}{D['df']}")
    r += 1
    def dcell(key, label, formula, fmt=NUM, bold=False, note=''):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        put(ws, r, 4, formula, 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        D[key] = r
        r += 1

    dcell('pv_sum', f'显性期PV合计 ({F0}-{F1})', f"=SUM({YC[F0]}{D['pv']}:{YC[F1]}{D['pv']})", NUM, True)
    dcell('tv', '终值 TV (Gordon)', f"={YC[F1]}{D['fcff']}*(1+{TG_REF})/({WACC_REF}-{TG_REF})", NUM, False, f'=FCFF{F1}×(1+g)/(WACC-g)')
    dcell('pv_tv', 'PV(终值)', f"=D{D['tv']}*{YC[F1]}{D['df']}", NUM, False, f'终值按t={NF-0.5}折现(与显性期最后一年一致)')
    dcell('ev', '企业价值 EV', f"=D{D['pv_sum']}+D{D['pv_tv']}", NUM, True)
    dcell('cash', f'加: 货币资金({H0}A末)', f"={BSs}!{YC[H0]}{B['cash']}")
    dcell('tfa', f'加: 交易性金融资产({H0}A末)', f"={BSs}!{YC[H0]}{B['tfa']}")
    dcell('oei', f'加: 其他权益工具投资({H0}A末)', f"={BSs}!{YC[H0]}{B['oei']}", NUM, False, f"{H0}A末{_bs['oei'][-1]/100:.1f}亿, 参股投资按账面值计入股权价值")
    dcell('debt', f'减: 有息负债({H0}A末)', f"=-({BSs}!{YC[H0]}{B['stl']}+{BSs}!{YC[H0]}{B['cur1y']}+{BSs}!{YC[H0]}{B['ltl']}+{BSs}!{YC[H0]}{B['lease']})")
    dcell('mi', f'减: 少数股东权益({H0}A末)', f"=-{BSs}!{YC[H0]}{B['mi']}", NUM, False,
          'EV桥得归母口径股权价值须扣少数股东权益(账面值); 无少数股东权益的标的为0, 不影响结果')
    dcell('netadj', '净现金调整合计', f"=D{D['cash']}+D{D['tfa']}+D{D['oei']}+D{D['debt']}+D{D['mi']}", NUM, False, '供Sensitivity/Scenarios页引用')
    dcell('eq', '股权价值', f"=D{D['ev']}+D{D['netadj']}", NUM, True)
    dcell('ps', f'每股价值 ({PER_SHARE_UNIT})', f"=D{D['eq']}/{SH_REF}", PS, True, '=股权价值/稀释股数(Equity_Roll页)')
    dcell('px', f'现价 ({VAL_DATE})', f"={ASheet}!$C${A['px']}", PS)
    dcell('upside', '隐含涨跌幅', f"=D{D['ps']}/D{D['px']}-1", PCT, True)
    dcell('ipe', f'隐含{F0}E P/E', f"=D{D['eq']}/{ISs}!{YC[F0]}{I['np_p']}", MULT, False, f'=股权价值/{F0}E归母')
    dcell('iev', f'隐含{F0}E EV/EBITDA', f"=D{D['ev']}/({ISs}!{YC[F0]}{I['ebit']}+{PPEs}!{YC[F0]}{P['da']})", MULT)
    put(ws, r, 1, f'注: 估值基准日{VAL_DATE}, 与{F0}财年现金流起点存在时间错位; 年中折现(t=0.5起)为该错位的标准近似处理, 不再单独做stub调整', 'g', size=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NOTE_COL)
    r += 1
    DCFs = 'DCF'

    # ============================================================
    # FCFE (股权自由现金流, Ke折现, 与FCFF双视图对照)
    # ============================================================
    ws = wb.create_sheet('FCFE')
    title_bar(ws, f'{NAME} — FCFE 股权自由现金流估值 (Ke折现, 与FCFF双视图)',
              f'单位: {UNIT} | FCFE=归母净利+D&A-ΔNWC-Capex+净新增借款(FIN页调度); 按股权成本Ke折现(年中口径与DCF页一致), '
              '直接得归母股权价值(少数股东权益见DCF页EV→Equity桥); 终值取正常化FCFE(净新增借款按D×g稳态正常化)')
    ws.column_dimensions['A'].width = 32
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 60
    put(ws, 3, 1, '项目', 't', bold=True)
    for y in FCST:
        put(ws, 3, 2 + YRS.index(y), f'{y}E', 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '备注', 't', bold=True)

    E = {}
    r = 4
    def erow(key, label, fn, fmt=NUM, bold=False, note=''):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        for y in FCST:
            put(ws, r, 2 + YRS.index(y), fn(y), 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        E[key] = r
        r += 1

    erow('np', '归母净利润', lambda y: f"={ISs}!{YC[y]}{I['np_p']}", NUM, False, '=IS归母净利(含FIN调度后的利息)')
    erow('da', '加: D&A', lambda y: f"={PPEs}!{YC[y]}{P['da']}", NUM, False, '=PP&E页折旧+无形摊销')
    erow('dnwc', '减: ΔNWC', lambda y: f"={SCH}!{YC[y]}{S['dnwc']}", NUM, False, '=Schedules; NWC增加为现金流出')
    erow('capex', '减: Capex', lambda y: f"={PPEs}!{YC[y]}{P['capex']}", NUM, False, '=PP&E页资本开支')
    erow('ddebt', '加: 净新增借款', lambda y: f"={FINs}!{YC[y]}{F['ddebt']}", NUM, False,
         '=FIN四档债务期末-期初(revolver新增-sweep偿还-计划还款); 股权口径下债务净变动归股东')
    erow('wmm', '理财净变动(勾稽参考, 不计入FCFE)', lambda y: f"=-{FINs}!{YC[y]}{F['dtfa']}", NUM, False,
         '=-FIN理财净变动: sweep溢出购买理财为现金↔理财的形态转换(资产仍在表内归属股东), 故不计入FCFE; '
         '显性列示供FCFE↔FCFF差异勾稽(FCFF也未扣理财购买, 两法在此口径一致)')
    erow('fcfe', 'FCFE', lambda y: f"={YC[y]}{E['np']}+{YC[y]}{E['da']}-{YC[y]}{E['dnwc']}-{YC[y]}{E['capex']}+{YC[y]}{E['ddebt']}", NUM, True,
         '=归母+D&A-ΔNWC-Capex+净新增借款')
    erow('t_idx', '折现年序(年中)', lambda y: 0.5 + FCST.index(y), '0.0', False, '年中折现, 与DCF页同口径')
    erow('df', '折现因子 (Ke)', lambda y: f"=1/(1+KE)^{YC[y]}{E['t_idx']}", '0.000', False, '=1/(1+Ke)^t; Ke为named range(Assumptions股权成本)')
    erow('pv', 'PV(FCFE)', lambda y: f"={YC[y]}{E['fcfe']}*{YC[y]}{E['df']}")
    r += 1

    def ecell(key, label, formula, fmt=NUM, bold=False, note=''):
        nonlocal r
        put(ws, r, 1, label, 't', bold=bold)
        put(ws, r, 4, formula, 'f', fmt, bold=bold)
        put(ws, r, NOTE_COL, note, 'g', size=9)
        E[key] = r
        r += 1

    ecell('pv_sum', f'显性期PV合计 ({F0}-{F1})', f"=SUM({YC[F0]}{E['pv']}:{YC[F1]}{E['pv']})", NUM, True)
    # A3: 终值净新增借款按 D×g 稳态正常化(对称处理, 不再MIN单边封顶) —
    #     大额净融资与大额净偿还均不应被Gordon公式永续外推
    ecell('fcfe_n', f'正常化FCFE ({F1}E, 终值口径)',
          f"={YC[F1]}{E['fcfe']}-{YC[F1]}{E['ddebt']}+{FINs}!{YC[F1]}{F['dtot1']}*TERM_G",
          NUM, False,
          f'终值净借款按D×g稳态正常化: 终值年净新增借款={F1}E期末有息负债×g(替代实际调度值, '
          '对称防止显性期末大额净融资/净偿还被Gordon公式永续外推)')
    ecell('tv', '终值 TV (Gordon, 正常化FCFE)', f"=D{E['fcfe_n']}*(1+TERM_G)/(KE-TERM_G)", NUM, False,
          f'=正常化FCFE{F1}×(1+g)/(Ke-g); Ke/g均为named range')
    ecell('pv_tv', 'PV(终值)', f"=D{E['tv']}*{YC[F1]}{E['df']}", NUM, False, f'终值按t={NF-0.5}折现(与DCF页一致)')
    ecell('eq', '股权价值 (FCFE口径)', f"=D{E['pv_sum']}+D{E['pv_tv']}", NUM, True, 'FCFE直接折现为股权价值, 无需净现金桥')
    ecell('ps', f'每股价值 ({PER_SHARE_UNIT})', f"=D{E['eq']}/SHARES_DIL", PS, True, '=股权价值/稀释股数(named range)')
    ecell('ps_fcff', f'对照: FCFF口径每股价值 ({PER_SHARE_UNIT})', f"={DCFs}!D{D['ps']}", PS, False, '=DCF页主输出')
    ecell('diff', '两法差异 (FCFE/FCFF-1)', f"=D{E['ps']}/D{E['ps_fcff']}-1", PCT, True, ' Checks页含一致性信息行')
    ecell('control', 'FCFE/FCFF差异控制 (阈值30%)',
          fcfe_control_formula(f'D{E["diff"]}'),
          None, True, '该状态为估值口径复核警报, 不替代两种方法各自的模型校验')
    put(ws, r, 1, '两法差异原因', 't', bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NOTE_COL)
    put(ws, r, 2, '①折现率不同: FCFF按WACC(全资本成本, 税盾在分母)折现, FCFE按Ke(纯股权成本)折现; '
                  '②口径不同: FCFF为企业自由现金流, 须经EV→Equity桥加回净现金/扣有息负债/扣少数股东权益, FCFE基于归母净利折现、'
                  '直接对应归母股权价值, 少数股东权益已在DCF页EV→Equity桥扣除, 本页无需再扣; '
                  '③杠杆路径不同: FCFE逐年反映FIN页revolver/sweep实际债务调度(资本结构动态变化), FCFF隐含目标结构(Wd)恒定; '
                  '④理财净变动(上行勾稽参考行)为现金形态转换, 两法均未计入, 不构成差异来源; '
                  '⑤两法在"恒定资本结构+一致税盾处理"下理论等价, 实务差异主要来自债务调度时点、净现金桥与折现率差; 差异幅度大时应复查融资假设',
        'g', size=9, wrap=True)
    ws.row_dimensions[r].height = 52
    r += 1
    FCFEs = 'FCFE'

    # ============================================================
    # Relative_Val (可比公司 + Beta unlever/relever)
    # ============================================================
    ws = wb.create_sheet('Relative_Val')
    title_bar(ws, f'{NAME} — Relative_Val 可比公司相对估值 + Beta unlever/relever',
              f'市值/PE-TTM为{VAL_DATE}收盘(腾讯行情); 预测归母为{CONS_NP_LABEL}/分析师输入; L-O列: βl/(1+(1-t)D/E) unlever取中位→Assumptions按目标结构relever', 15)
    ws.column_dimensions['A'].width = 14
    for col, w in zip('BCDEFGHIJ', [10, 12, 10, 14, 11, 14, 11, 12, 10]):
        ws.column_dimensions[col].width = w
    for col, w in zip('LMNO', [9, 9, 8, 9]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions['K'].width = 56
    hdrs = ['公司', '代码', '市值(亿元)', 'PE-TTM', f'{F0}E归母(亿)', f'{F0}E PE', f'{F0+1}E归母(亿)', f'{F0+1}E PE', f'{str(F0)[2:]}→{str(F0+1)[2:]}增速', 'PEG',
            '数据来源 / 备注', 'βl(输入)', 'D/E(输入)', '税率(输入)', 'βu(公式)']
    for i, h in enumerate(hdrs):
        c = put(ws, 4, 1 + i, h, 't', bold=True, align='center')
        c.fill = TOT_FILL
    r = RV_COMP_START
    COMP_ROWS = []
    BETA_ROWS = []
    PRICE_ROWS = []
    F1_ROWS = []
    PEG_ROWS = []
    _prior_comp_source = ''
    for cp in comps:
        put(ws, r, 1, cp['name'], 't')
        put(ws, r, 2, cp['code'], 'g', align='center')
        put(ws, r, 3, cp['mcap'], 'x', NUM)
        put(ws, r, 4, cp['pe_ttm'], 'x', MULT)
        put(ws, r, 5, cp['np_f0'], 'x', NUM)
        put(ws, r, 6, f"=C{r}/E{r}", 'f', MULT)
        has_f1 = cp.get('np_f1') is not None
        has_peg = False
        if not has_f1:
            for cc in (7, 8, 9, 10):
                put(ws, r, cc, '—', 'g', align='center')
        else:
            put(ws, r, 7, cp['np_f1'], 'x', NUM)
            put(ws, r, 8, f"=C{r}/G{r}", 'f', MULT)
            put(ws, r, 9, f"=G{r}/E{r}-1", 'f', PCT)
            has_peg = cp['np_f1'] > cp['np_f0']
            if has_peg:
                put(ws, r, 10, f"=F{r}/(I{r}*100)", 'f', '0.00')
            else:
                put(ws, r, 10, '—', 'g', align='center')
        _cp_source = str(cp.get('src') or '')
        _pricing_ok = forward_earnings_verified(cp, _prior_comp_source, VAL_DATE)
        _pricing_status = ('定价资格: 已验证FY1/NTM' if _pricing_ok
                           else '定价资格: 排除(未验证远期盈利/仅参照)')
        put(ws, r, 11, (_cp_source + ' | ' if _cp_source else '') + _pricing_status, 'g', size=9, wrap=True)
        if _cp_source not in ('同上', '同上。'):
            _prior_comp_source = _cp_source
        if cp.get('beta_l') is not None:
            put(ws, r, 12, cp['beta_l'], 'in', '0.00')
            put(ws, r, 13, cp['d_e'], 'in', PCT)
            put(ws, r, 14, cp['tax'], 'in', PCT)
            put(ws, r, 15, f"=L{r}/(1+(1-N{r})*M{r})", 'f', '0.000')
            BETA_ROWS.append(r)
        else:
            for cc in (12, 13, 14, 15):
                put(ws, r, cc, '—', 'g', align='center')
        if _pricing_ok:
            PRICE_ROWS.append(r)
            if has_f1:
                F1_ROWS.append(r)
            if has_peg:
                PEG_ROWS.append(r)
        COMP_ROWS.append(r)
        r += 1
    assert r == RV_MED_ROW, f'RV median row mismatch: {r} != {RV_MED_ROW}'
    # 0.5.1: PE中位与β中位解耦 —— PE中位统计全部计价可比(非ref_only),
    # β中位仍只统计带beta_l的可比; 两集合相同时保持旧版label/公式逐字节不变
    _price_same = (PRICE_ROWS == BETA_ROWS)
    def _med(colL, rows):
        if rows == list(range(rows[0], rows[-1] + 1)):
            return f"=MEDIAN({colL}{rows[0]}:{colL}{rows[-1]})"
        return "=MEDIAN(" + ",".join(f"{colL}{x}" for x in rows) + ")"
    if _price_same or not PRICE_ROWS:
        put(ws, r, 1, f'可比中位数(A股{N_BETA}家)', 't', bold=True)
    else:
        put(ws, r, 1, f'可比中位数(计价{len(PRICE_ROWS)}家/β{N_BETA}家)', 't', bold=True)
    if PRICE_ROWS:
        put(ws, r, 6, _med('F', PRICE_ROWS), 'f', MULT, bold=True)
    if F1_ROWS:
        put(ws, r, 8, _med('H', F1_ROWS), 'f', MULT, bold=True)
    elif COMP_ROWS:
        put(ws, r, 8, '—', 'g', bold=True, align='center')
    if PEG_ROWS:
        put(ws, r, 10, _med('J', PEG_ROWS), 'f', '0.00', bold=True)
    elif COMP_ROWS:
        put(ws, r, 10, '—', 'g', bold=True, align='center')
    if N_BETA:
        _b0, _b1 = BETA_ROWS[0], BETA_ROWS[-1]
        c = put(ws, r, 15, f"=MEDIAN(O{_b0}:O{_b1})", 'f', '0.000', bold=True, fill=CHK_FILL)
        put(ws, r, 11, 'βu中位数→Assumptions WACC区relever; βl/D/E为分析师输入(参考行情终端β与各公司最新年报杠杆), 可按终端数据更新', 'g', size=9, wrap=True)
    elif PRICE_ROWS:
        put(ws, r, 11, '计价可比PE中位→目标PE上限; 无beta_l可比, βu走Assumptions蓝色输入(不影响折现率)', 'g', size=9, wrap=True)
    else:
        put(ws, r, 11, '无可比公司配置(兜底模式): βu走Assumptions蓝色输入, 相对估值上下限=目标PE下限', 'g', size=9, wrap=True)
    MED_ROW = r
    r += 1
    put(ws, r, 1, f'{NAME}(现价)', 't', bold=True)
    put(ws, r, 2, CODE_FULL, 'g', align='center')
    put(ws, r, 3, f"={ASheet}!$C${A['mcap']}/100", 'f', NUM)
    put(ws, r, 4, cfg.v('market.pe_ttm'), 'x', MULT)
    put(ws, r, 5, f"={ASheet}!$C${A[f'c_np{_cy[0]}']}/100", 'f', NUM)
    put(ws, r, 6, f"=C{r}/E{r}", 'f', MULT, bold=True)
    put(ws, r, 7, f"={ASheet}!$C${A[f'c_np{_cy[1]}']}/100", 'f', NUM)
    put(ws, r, 8, f"=C{r}/G{r}", 'f', MULT)
    put(ws, r, 11, f'{F0}E/{F0+1}E归母={CONS_NP_LABEL}; {cfg.v("market.pe_ttm_basis", "")}', 'g', size=9, wrap=True)
    SH_ROW = r
    r += 2
    sec(ws, r, f'{NAME}定价 (基于模型基准{F0}E归母)', 15); r += 1
    _pe_lo_e = cfg.e('relative_val.target_pe_lo')
    # F5: 经_norm_entry规整后 list/{values:[...]} 取首元素为float标量,
    #     避免openpyxl写list报ValueError与后续:g格式化TypeError
    _pe_lo_v = _pe_lo_e['values']
    if isinstance(_pe_lo_v, (list, tuple)):
        _pe_lo_v = _pe_lo_v[0]
    _pe_lo_v = float(_pe_lo_v)
    put(ws, r, 1, '目标PE下限', 't'); put(ws, r, 3, _pe_lo_v, 'in', MULT)
    put(ws, r, 11, _pe_lo_e['basis'], 'g', size=9)
    PE_LO = r; r += 1
    put(ws, r, 1, '目标PE上限', 't')
    if PRICE_ROWS and _price_same:
        put(ws, r, 3, f"=F{MED_ROW}", 'f', MULT)
        put(ws, r, 11, f'=A股可比{F0}E PE中位数(公式)', 'g', size=9)
    elif PRICE_ROWS:
        put(ws, r, 3, f"=F{MED_ROW}", 'f', MULT)
        put(ws, r, 11, f'=计价可比{F0}E PE中位数(公式, 含海外; ref_only标记的仅参照行不参与)', 'g', size=9)
    else:
        put(ws, r, 3, f"=C{PE_LO}", 'f', MULT)
        put(ws, r, 11, '=目标PE下限(无可比公司配置, 兜底)', 'g', size=9)
    PE_HI = r; r += 1
    put(ws, r, 1, f'模型基准{F0}E归母(百万)', 't'); put(ws, r, 3, f"={ISs}!{YC[F0]}{I['np_p']}", 'f', NUM)
    NP26 = r; r += 1
    put(ws, r, 1, f'对应股价下限({PS_DENOM})', 't', bold=True)
    put(ws, r, 3, f"=C{PE_LO}*C{NP26}/{SH_REF}", 'f', PS, bold=True)
    P_LO = r; r += 1
    put(ws, r, 1, f'对应股价上限({PS_DENOM})', 't', bold=True)
    put(ws, r, 3, f"=C{PE_HI}*C{NP26}/{SH_REF}", 'f', PS, bold=True)
    P_HI = r; r += 1
    put(ws, r, 1, f'以{CONS_NP_LABEL}{F0}E归母{cons_np_v[0]/100:.1f}亿计股价区间({PS_DENOM})', 't')
    put(ws, r, 3, f"=C{PE_LO}*{ASheet}!$C${A[f'c_np{_cy[0]}']}/{SH_REF}", 'f', PS)
    put(ws, r, 4, f"=C{PE_HI}*{ASheet}!$C${A[f'c_np{_cy[0]}']}/{SH_REF}", 'f', PS)
    put(ws, r, 11, '左=PE下限, 右=PE上限', 'g', size=9)
    P_CONS_LO = r
    r += 1
    put(ws, r, 1, f'相对现价{_px_val}{PS_DENOM}空间', 't')   # F1: 现价与Assumptions采用值同源(含price_hkd/fx折算口径)
    put(ws, r, 3, f"=C{P_LO}/{ASheet}!$C${A['px']}-1", 'f', PCT)
    put(ws, r, 4, f"=C{P_HI}/{ASheet}!$C${A['px']}-1", 'f', PCT)
    r += 1
    # 研究层: 一致预期目标价对照行 (仅 --consensus 激活)
    TP_ROW = None
    _cons_ctx = (research or {}).get('consensus')
    if _cons_ctx and _cons_ctx.get('target_price'):
        _tp = float(_cons_ctx['target_price'])
        put(ws, r, 1, '一致预期目标价 (外部文件)', 't', bold=True)
        put(ws, r, 3, _tp, 'x', PS)
        put(ws, r, 4, f"=C{r}/{ASheet}!$C${A['px']}-1", 'f', PCT)
        put(ws, r, 11, f"来源: {_cons_ctx['source']}{' ' + _cons_ctx['date'] if _cons_ctx['date'] else ''}; "
                       '右=较现价空间; 与模型DCF对照见Checks页信息行', 'g', size=9)
        TP_ROW = r
        r += 1
    RVs = 'Relative_Val'
    # 定价中位数只依赖非ref_only计价可比, 与是否配置beta完全解耦。
    REL_MED_PE = f"{RVs}!$F${MED_ROW}" if PRICE_ROWS else f"{RVs}!$C${PE_LO}"

    # ============================================================
    # Sensitivity (矩阵轴公式化自动对中; DPO单维)
    # ============================================================
    ws = wb.create_sheet('Sensitivity')
    title_bar(ws, f'{NAME} — Sensitivity 敏感性分析',
              f'单位: {PER_SHARE_UNIT} | 矩阵轴为公式(自动对中WACC采用值/g), 两矩阵引用DCF页FCFF与Equity_Roll稀释股数; 另附DPO单维敏感性')
    ws.column_dimensions['A'].width = 26
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['I'].width = 60

    r = 4
    sec(ws, r, '矩阵一: DCF每股价值 — WACC × 永续增长率g (年中折现, 与DCF页同口径)', 9); r += 1
    put(ws, r, 1, 'WACC \\ g', 't', bold=True, align='center')
    g_offsets = [-0.01, -0.005, 0.0, 0.005, 0.01]
    for i, off in enumerate(g_offsets):
        f = f"={TG_REF}{off:+.4f}".replace('+0.0000', '').replace('-0.0000', '')
        put(ws, r, 3 + i, f, 'f', PCT, bold=True)
    G_HDR = r
    r += 1
    w_offsets = [-0.01, -0.005, 0.0, 0.005, 0.01]
    FCFF_R = D['fcff']; NET_R = D['netadj']
    M1_FIRST = r
    for off in w_offsets:
        wf = f"={WACC_REF}{off:+.4f}".replace('+0.0000', '').replace('-0.0000', '')
        put(ws, r, 2, wf, 'f', PCT, bold=True)
        for i in range(len(g_offsets)):
            cl = get_column_letter(3 + i)
            terms = '+'.join(f"DCF!${YC[y]}${FCFF_R}/(1+$B{r})^{0.5 + j}" for j, y in enumerate(FCST))
            f = (f"=({terms}"
                 f"+DCF!${YC[F1]}${FCFF_R}*(1+{cl}${G_HDR})/($B{r}-{cl}${G_HDR})/(1+$B{r})^{NF-0.5}"
                 f"+DCF!$D${NET_R})/{SH_REF}")
            fill = CHK_FILL if off == 0.0 and g_offsets[i] == 0.0 else None
            put(ws, r, 3 + i, f, 'f', PS, fill=fill)
        r += 1
    M1_CENTER = f"E{M1_FIRST + 2}"
    put(ws, r, 1, '高亮格=当前采用值(WACC/g轴自动对中); 与DCF页同为年中折现口径', 'g', size=9); r += 2

    _pes = cfg.get('sensitivity.pe_list') or [25, 30, 35, 40, 45, 50]
    _npgs = cfg.get('sensitivity.np_growth_list') or [0.40, 0.60, 0.80, 1.00, 1.115, 1.30]
    _hl = cfg.get('sensitivity.highlight') or {}
    sec(ws, r, f'矩阵二: 股价 — {F0}E归母增速 × 目标PE ({H0}A归母{_is_h["np_p"][-1]:.0f}百万为基数)', 9); r += 1
    put(ws, r, 1, f'{F0}E归母增速 \\ PE', 't', bold=True, align='center')
    for i, pe in enumerate(_pes):
        put(ws, r, 3 + i, pe, 'in', '0"x"', bold=True)
    PE_HDR = r
    r += 1
    for g in _npgs:
        put(ws, r, 2, g, 'in', PCT, bold=True)
        for i, pe in enumerate(_pes):
            cl = get_column_letter(3 + i)
            f = f"={ISs}!${YC[H0]}${I['np_p']}*(1+$B{r})*{cl}${PE_HDR}/{SH_REF}"
            fill = CHK_FILL if _hl and abs(g - _hl.get('growth', -9)) < 1e-9 and pe == _hl.get('pe', -9) else None
            put(ws, r, 3 + i, f, 'f', PS, fill=fill)
        r += 1
    put(ws, r, 1, cfg.get('sensitivity.matrix2_note', f'高亮行≈{CONS_NP_LABEL}增速; 矩阵与DCF独立, 供交叉验证'), 'g', size=9)
    r += 2

    _dpo_deltas = cfg.get('sensitivity.dpo_deltas') or [-40, -20, 0, 20, 40]
    # E4: δ=0为DCF基准锚点(DPO_CENTER引用), 配置不含0时自动补0; 排序去重保证行序与锚点恒存在
    _dpo_deltas = sorted(set(_dpo_deltas) | {0})
    _dpo_path = a5('working_capital.dpo')
    sec(ws, r, '矩阵三: DPO单维敏感性 — 应付天数变动 × DCF每股价值', 9); r += 1
    put(ws, r, 1, 'DPO变动(天) \\ 每股价值', 't', bold=True, align='center')
    put(ws, r, 3, f'每股价值({PER_SHARE_UNIT})', 't', bold=True, align='center')
    put(ws, r, 4, 'vs基准', 't', bold=True, align='center')
    put(ws, r, 9, '机制: DPO变动δ→应付变动→ΔNWC变动→FCFF一阶调整; 首年按全额COGS重定价, 以后年按ΔCOGS; '
                  '未反映利息二阶效应(金额小); δ=0格=DCF基准', 'g', size=9, wrap=True)
    r += 1
    # A2: 第一预测年应付重定价基数=全额COGS_F0 (AP_{H0}为既成事实, δ作用于F0全部应付重建),
    #     后续年保持年际差分(ΔCOGS); δ=0行不受影响, 仍恒等于DCF基准
    dcogs = {}
    for y in FCST:
        if y == F0:
            dcogs[y] = f"({ISs}!{YC[y]}${I['cost']})"
        else:
            dcogs[y] = f"({ISs}!{YC[y]}${I['cost']}-{ISs}!{YC[y-1]}${I['cost']})"
    DPO_CENTER = None
    for dpo_d in _dpo_deltas:
        put(ws, r, 2, dpo_d, 'in', DAYS, bold=True)
        terms = []
        for j, y in enumerate(FCST):
            t = 0.5 + j
            terms.append(f"(DCF!{YC[y]}${FCFF_R}+$B{r}/365*{dcogs[y]})/(1+{WACC_REF})^{t}")
        tv_term = (f"(DCF!${YC[F1]}${FCFF_R}+$B{r}/365*{dcogs[F1]})*(1+{TG_REF})/({WACC_REF}-{TG_REF})/(1+{WACC_REF})^{NF-0.5}")
        f = f"=({'+'.join(terms)}+{tv_term}+DCF!$D${NET_R})/{SH_REF}"
        fill = CHK_FILL if dpo_d == 0 else None
        put(ws, r, 3, f, 'f', PS, fill=fill)
        put(ws, r, 4, f"=C{r}/{DCFs}!$D${D['ps']}-1", 'f', PCT)
        if dpo_d == 0:
            DPO_CENTER = f"C{r}"
        r += 1
    put(ws, r, 1, f'DPO每±20天: 首年≈FCFF±(营业成本×20/365), 以后年≈±(Δ营业成本×20/365); '
                  f'基准DPO路径见Assumptions({_dpo_path[0]:.0f}→{_dpo_path[-1]:.0f}天)', 'g', size=9)
    SENs = 'Sensitivity'

    # ============================================================
    # Scenarios (基准引用主模型输出; 年中折现同步)
    # ============================================================
    ws = wb.create_sheet('Scenarios')
    title_bar(ws, f'{NAME} — Scenarios 熊/基/牛三情景估值',
              f'单位: {UNIT} | 基准估值=主模型输出(DCF页/IS页直接引用); 熊/牛保留简化净利率法并附桥接说明; 折现为年中口径与DCF一致')
    ws.column_dimensions['A'].width = 26
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 56
    put(ws, 3, 1, '情景开关当前值', 't', bold=True)
    put(ws, 3, 2, f"={ASheet}!$C${A['sw']}", 'f', '0', bold=True)
    put(ws, 3, 3, '=CHOOSE(B3,"熊市","基准","牛市")', 'f', align='center')
    put(ws, 3, NOTE_COL, '主模型(IS/BS/CF/FIN/DCF)随开关切换; 本页熊/基/牛横向比较全部采用同一简化引擎, 主模型另列桥接', 'g', size=9)

    # NWC密度 rho (基准NWC_F0/收入_F0)
    put(ws, 4, 1, '基准NWC/收入密度ρ', 't')
    put(ws, 4, 2, f"={SCH}!{YC[F0]}{S['nwc']}/{ISs}!{YC[F0]}{I['rev']}", 'f', PCT)
    put(ws, 4, NOTE_COL, '情景页ΔNWC=收入增量×ρ(负密度=现金流来源); 熊/基/牛同一简化引擎均使用', 'g', size=9)
    RHO = '$B$4'

    scen_rows = {}
    r = 6
    for skey, sname, arow_ in [('bear', '熊市 Bear', A['sc_bear']), ('base', '基准 Base', A['sc_base']), ('bull', '牛市 Bull', A['sc_bull'])]:
        sec(ws, r, f'{sname} — 收入增速调整/净利率调整引用Assumptions第{arow_}行'); r += 1
        put(ws, r, 1, '项目', 't', bold=True)
        for y in FCST:
            put(ws, r, 2 + YRS.index(y), f'{y}E', 't', bold=True, align='center').fill = TOT_FILL
        r += 1
        ADJG = f"{ASheet}!$C${arow_}"
        ADJM = f"{ASheet}!$D${arow_}"
        rows = {}
        put(ws, r, 1, '营业收入', 't')
        for y in FCST:
            cl = YC[y]
            if y == F0:
                f = f"={ISs}!{YC[H0]}{I['rev']}*(1+{RS}!{cl}{BASE_G}+{ADJG})"
            else:
                f = f"={YC[y-1]}{r}*(1+{RS}!{cl}{BASE_G}+{ADJG})"
            put(ws, r, 2 + YRS.index(y), f, 'f', NUM)
        rows['rev'] = r; r += 1
        put(ws, r, 1, '归母净利率', 't')
        for y in FCST:
            cl = YC[y]
            # A1: 简化法净利率同步×(1-mi_share)保持归母口径, mi_share=0时数值恒等
            f = (f"=({RS}!{cl}{BASE_GM}+{ADJM}-{ASheet}!{AC[y]}${A['rt_taxadd']}-{ASheet}!{AC[y]}${A['rt_sale']}"
                 f"-{ASheet}!{AC[y]}${A['rt_adm']}-{ASheet}!{AC[y]}${A['rt_rd']}-{ASheet}!$C${A['oth_rt']})"
                 f"*(1-{ASheet}!{AC[y]}${A['tax']})*(1-{ASheet}!{AC[y]}${A['mi_share']})")
            put(ws, r, 2 + YRS.index(y), f, 'f', PCT)
        rows['npm'] = r; r += 1
        put(ws, r, 1, '归母净利润', 't')
        for y in FCST:
            cl = YC[y]
            put(ws, r, 2 + YRS.index(y), f"={cl}{rows['rev']}*{cl}{rows['npm']}", 'f', NUM, bold=True)
        rows['np'] = r; r += 1
        put(ws, r, 1, 'FCFF (简化)', 't')
        for y in FCST:
            cl = YC[y]
            prev_ref = f"{ISs}!{YC[H0]}{I['rev']}" if y == F0 else f"{YC[y-1]}{rows['rev']}"
            f = (f"={cl}{rows['np']}+{PPEs}!{cl}{P['da']}-({cl}{rows['rev']}-{prev_ref})*{RHO}-{PPEs}!{cl}{P['capex']}")
            put(ws, r, 2 + YRS.index(y), f, 'f', NUM)
        rows['fcff'] = r; r += 1
        put(ws, r, 1, f'DCF每股价值({PER_SHARE_UNIT}, 简化法)', 't', bold=True)
        fcff = rows['fcff']
        terms = '+'.join(f"{YC[y]}{fcff}/(1+{WACC_REF})^{0.5 + j}" for j, y in enumerate(FCST))
        f = (f"=({terms}"
             f"+{YC[F1]}{fcff}*(1+{TG_REF})/({WACC_REF}-{TG_REF})/(1+{WACC_REF})^{NF-0.5}"
             f"+DCF!$D${D['netadj']})/{SH_REF}")
        put(ws, r, 2, f, 'f', PS, bold=True, fill=CHK_FILL)
        put(ws, r, NOTE_COL, '简化净利率法+年中折现; 熊/基/牛均使用相同FCFF、WACC/g及净现金桥口径', 'g', size=9)
        rows['ps'] = r; r += 1
        put(ws, r, 1, f'相对估值股价({PER_SHARE_UNIT})', 't', bold=True)
        put(ws, r, 2, f"={YC[F0]}{rows['np']}*{REL_MED_PE}/{SH_REF}", 'f', PS, bold=True, fill=CHK_FILL)
        put(ws, r, NOTE_COL, f'=情景{F0}E归母×计价可比中位{F0}E PE/稀释股数', 'g', size=9)
        rows['rel'] = r; r += 1
        scen_rows[skey] = rows
        r += 1

    sec(ws, r, '主模型桥接 (不参与熊/基/牛同引擎横向替代)'); r += 1
    put(ws, r, 1, f'主模型当前DCF ({PER_SHARE_UNIT})', 't', bold=True)
    put(ws, r, 2, f"={DCFs}!D{D['ps']}", 'f', PS, bold=True, fill=CHK_FILL)
    put(ws, r, NOTE_COL, '完整三表/FIN/FCFF主模型输出; 当前情景由Assumptions开关决定', 'g', size=9)
    SCEN_MAIN_PS_ROW = r; r += 1
    put(ws, r, 1, f'同引擎基准DCF ({PER_SHARE_UNIT})', 't')
    put(ws, r, 2, f"=B{scen_rows['base']['ps']}", 'f', PS)
    SCEN_BRIDGE_BASE_ROW = r; r += 1
    put(ws, r, 1, '主模型/同引擎基准 比值', 't', bold=True)
    put(ws, r, 2, f'=B{SCEN_MAIN_PS_ROW}/B{scen_rows["base"]["ps"]}', 'f', PCT, bold=True)
    put(ws, r, NOTE_COL, '桥接差异来自FIN计息、逐项营运资本与完整税务口径; 不覆盖三情景简化法基准', 'g', size=9)
    SCEN_BRIDGE_RATIO_ROW = r; r += 1
    put(ws, r, 1, '主模型/同引擎基准 控制状态', 't')
    put(ws, r, 2, f'=IF(ABS(B{SCEN_BRIDGE_RATIO_ROW}-1)>=0.3,"警告: 桥接差异≥30%, 复核简化假设","通过: 桥接差异<30%")',
        'f', bold=True, fill=CHK_FILL)
    SCEN_BRIDGE_STATUS_ROW = r; r += 2
    # 汇总对比
    sec(ws, r, '三情景汇总'); r += 1
    put(ws, r, 1, '指标', 't', bold=True)
    for i, t in enumerate(['熊市', '基准', '牛市', '现价', '熊vs现价', '基vs现价', '牛vs现价']):
        put(ws, r, 2 + i, t, 't', bold=True, align='center')
    r += 1
    put(ws, r, 1, f'DCF每股价值({PER_SHARE_UNIT})', 't')
    for i, k in enumerate(['bear', 'base', 'bull']):
        put(ws, r, 2 + i, f"=B{scen_rows[k]['ps']}", 'f', PS, bold=True)
    put(ws, r, 5, f"={ASheet}!$C${A['px']}", 'f', PS)
    for i, k in enumerate(['bear', 'base', 'bull']):
        put(ws, r, 6 + i, f"=B{scen_rows[k]['ps']}/{ASheet}!$C${A['px']}-1", 'f', PCT)
    SUM_PS = r
    r += 1
    put(ws, r, 1, f'相对估值股价({PER_SHARE_UNIT})', 't')
    for i, k in enumerate(['bear', 'base', 'bull']):
        put(ws, r, 2 + i, f"=B{scen_rows[k]['rel']}", 'f', PS)
    SUM_REL = r
    r += 1
    put(ws, r, 1, f'{F0}E归母(百万)', 't')
    for i, k in enumerate(['bear', 'base', 'bull']):
        put(ws, r, 2 + i, f"={YC[F0]}{scen_rows[k]['np']}", 'f', NUM)
    SCENs = 'Scenarios'
    print('DCF/RV/Sensitivity/Scenarios done')

    # ============================================================
    # Summary (football field 区间汇总)
    # ============================================================
    _sc_bear = cfg.raw['scenarios']['bear']; _sc_bull = cfg.raw['scenarios']['bull']
    ws = wb.create_sheet('Summary')
    title_bar(ws, f'{NAME} — Summary 估值结论汇总 (Football Field)',
              f'单位: {PER_SHARE_UNIT} | 熊/基/牛均为同一简化情景引擎; 完整主模型单列桥接; 自动外推不纳入经验证包络', 8)
    ws.column_dimensions['A'].width = 34
    for col, w in zip('BCDEFG', [30, 12, 12, 12, 12, 40]):
        ws.column_dimensions[col].width = w
    r = 4
    hdrs = ['估值方法', '口径', f'低({PS_DENOM})', f'高({PS_DENOM})', f'中点({PS_DENOM})', '中点vs现价', '备注']
    for i, h in enumerate(hdrs):
        c = put(ws, r, 1 + i, h, 't', bold=True, align='center')
        c.fill = TOT_FILL
    r += 1
    FF_FIRST = r
    # F5: _pe_lo_v 已在Relative_Val区规整为float标量, 此处直接复用
    if PRICE_ROWS:
        _pe_band = f'PE {_pe_lo_v:g}x ~ 可比中位'
        _pe_note = '目标PE带'
    else:
        _pe_band = f'PE {_pe_lo_v:g}x 单点（无正式可比）'
        _pe_note = '目标PE单点'
    # 最后一列控制是否进入主估值包络; 自动推导的“consensus”只展示、不混入经验证方法。
    ff_rows = [
        ('DCF完整主模型 — 当前开关', '完整三表/FIN', f"={DCFs}!D{D['ps']}", f"={DCFs}!D{D['ps']}",
         '完整FCFF+Gordon终值; 与简化基准的桥接见Scenarios/Checks', True),
        ('DCF情景 — 熊市', '同引擎简化法', f"={SCENs}!B{scen_rows['bear']['ps']}", f"={SCENs}!B{scen_rows['bear']['ps']}",
         f"分部增速{_sc_bear['rev_adj']*100:g}pct/净利率{_sc_bear['npm_adj']*100:g}pct", True),
        ('DCF情景 — 基准', '同引擎简化法', f"={SCENs}!B{scen_rows['base']['ps']}", f"={SCENs}!B{scen_rows['base']['ps']}",
         '与熊/牛完全相同的简化收入、净利率、FCFF及折现引擎', True),
        ('DCF情景 — 牛市', '同引擎简化法', f"={SCENs}!B{scen_rows['bull']['ps']}", f"={SCENs}!B{scen_rows['bull']['ps']}",
         f"分部增速+{_sc_bull['rev_adj']*100:g}pct/净利率+{_sc_bull['npm_adj']*100:g}pct", True),
        (f'相对估值 — 模型{F0}E归母', _pe_band, f"={RVs}!C{P_LO}", f"={RVs}!C{P_HI}",
         f'模型基准{F0}E归母×{_pe_note}', True),
        (f'相对估值 — {CONS_NP_LABEL}{F0}E归母', _pe_band, f"={RVs}!C{P_CONS_LO}", f"={RVs}!D{P_CONS_LO}",
         (f'{CONS_NP_LABEL}{cons_np_v[0]/100:.1f}亿×{_pe_note}; '
          + ('已验证来源, 纳入包络' if CONS_NP_VERIFIED else '自动推导, 仅展示不纳入包络')), CONS_NP_VERIFIED),
        ('FCFE股权现金流 — 基准', 'Ke折现(股权口径)', f"={FCFEs}!D{E['ps']}", f"={FCFEs}!D{E['ps']}",
         'FCFE=归母+D&A-ΔNWC-Capex+净新增借款; 差异≥30%触发下方控制警报', True),
    ]
    ff_output_rows = []
    for name, basis, flo, fhi, note, include_envelope in ff_rows:
        ff_output_rows.append((r, include_envelope))
        put(ws, r, 1, name, 't', bold=True)
        put(ws, r, 2, basis, 'g', align='center', size=9)
        put(ws, r, 3, flo, 'f', PS)
        put(ws, r, 4, fhi, 'f', PS)
        put(ws, r, 5, f"=(C{r}+D{r})/2", 'f', PS, bold=True)
        put(ws, r, 6, f"=E{r}/{ASheet}!$C${A['px']}-1", 'f', PCT)
        put(ws, r, 7, note, 'g', size=9, wrap=True)
        r += 1
    put(ws, r, 1, f'现价 ({VAL_DATE})', 't', bold=True)
    put(ws, r, 3, f"={ASheet}!$C${A['px']}", 'f', PS)
    put(ws, r, 4, f"={ASheet}!$C${A['px']}", 'f', PS)
    put(ws, r, 5, f"={ASheet}!$C${A['px']}", 'f', PS)
    PX_ROW = r
    r += 1
    put(ws, r, 1, '综合参考区间(经验证方法包络)', 't', bold=True)
    _eligible_rows = [rr for rr, include in ff_output_rows if include]
    put(ws, r, 3, '=MIN(' + ','.join(f'C{rr}' for rr in _eligible_rows) + ')', 'f', PS, bold=True, fill=CHK_FILL)
    put(ws, r, 4, '=MAX(' + ','.join(f'D{rr}' for rr in _eligible_rows) + ')', 'f', PS, bold=True, fill=CHK_FILL)
    put(ws, r, 5, f"=(C{r}+D{r})/2", 'f', PS, bold=True, fill=CHK_FILL)
    put(ws, r, 6, f"=E{r}/{ASheet}!$C${A['px']}-1", 'f', PCT, bold=True, fill=CHK_FILL)
    put(ws, r, 7, '仅纳入完整主模型、同引擎情景、模型相对估值、FCFE及已验证外部一致预期; 自动外推排除',
        'g', size=9, wrap=True)
    ENV_ROW = r
    r += 2
    # 简易文本条形(football field可视化)
    sec(ws, r, '区间可视化 (每个▮≈10元, 起点=区间低值)', 8); r += 1
    put(ws, r, 1, '方法', 't', bold=True)
    put(ws, r, 2, '区间条(低→高, 括号为现价相对位置)', 't', bold=True)
    r += 1
    for i, (name, basis, flo, fhi, note, include_envelope) in enumerate(ff_rows):
        rr = FF_FIRST + i
        bar = (f"=REPT(\"▮\",MAX(1,ROUND((D{rr}-C{rr})/10,0)))"
               f"&\"  [\"&TEXT(C{rr},\"0\")&\"—\"&TEXT(D{rr},\"0\")&\"]\"")
        put(ws, r, 1, name, 'g', size=9)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        put(ws, r, 2, bar, 'f', size=9)
        r += 1
    r += 1
    put(ws, r, 1, '关键假设提示', 't', bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    put(ws, r, 2, 'WACC采用值与可比beta联动(Assumptions第十节); 三情景主模型切换见Assumptions情景开关; 校验状态见Checks页', 'g', size=9, wrap=True)
    r += 1
    put(ws, r, 1, 'FCFE/FCFF差异控制 (阈值30%)', 't', bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    put(ws, r, 2, fcfe_control_formula(f'{FCFEs}!D{E["diff"]}'),
        'f', bold=True, fill=CHK_FILL)
    SUMMARY_FCFE_CONTROL_ROW = r; r += 1
    put(ws, r, 1, '主模型/同引擎基准桥接控制', 't', bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    put(ws, r, 2, f'=IF(ABS({SCENs}!B{SCEN_BRIDGE_RATIO_ROW}-1)>=0.3,"警告: 桥接差异≥30%, 复核简化引擎","通过: 桥接差异<30%")',
        'f', bold=True, fill=CHK_FILL)
    SUMMARY_BRIDGE_CONTROL_ROW = r
    SUMMARYs = 'Summary'

    # ============================================================
    # Checks 校验页
    # ============================================================
    # 关键驱动 named ranges (审计轨迹): WACC采用值/永续g/税率/股数/情景开关等15个
    NAMED_RANGES = {
        'SCEN_SWITCH': f"Assumptions!$C${A['sw']}",
        'WACC_USED': f"Assumptions!$C${A['wacc']}",
        'WACC_CALC': f"Assumptions!$C${A['wacc_calc']}",
        'TERM_G': f"Assumptions!$C${A['tg']}",
        'KE': f"Assumptions!$C${A['ke']}",
        'KD': f"Assumptions!$C${A['kd']}",
        'RF_RATE': f"Assumptions!$C${A['rf']}",
        'ERP': f"Assumptions!$C${A['erp']}",
        'BETA_U': f"Assumptions!$C${A['beta_u']}",
        'TAX_RATE_Y1': f"Assumptions!$C${A['tax']}",
        'PAYOUT_Y1': f"Assumptions!$C${A['payout']}",
        'MIN_CASH_PCT': f"Assumptions!$C${A['min_cash_pct']}",
        'SHARES_DIL': f"Equity_Roll!$C${Q['shares']}",
        'DCF_PS': f"DCF!$D${D['ps']}",
        'FCFE_PS': f"FCFE!$D${E['ps']}",
    }
    ws = wb.create_sheet('Checks')
    title_bar(ws, f'{NAME} — Checks 自动校验',
              '所有校验为公式, 应全部显示TRUE; 任一FALSE请检查模型; 容差0.01')
    ws.column_dimensions['A'].width = 40
    for y in YRS:
        ws.column_dimensions[YC[y]].width = 11
    ws.column_dimensions[get_column_letter(NOTE_COL)].width = 50
    put(ws, 3, 1, '校验项', 't', bold=True)
    for y in YRS:
        put(ws, 3, 2 + YRS.index(y), (f'{y}A' if y in HIST else f'{y}E'), 't', bold=True, align='center').fill = TOT_FILL
    put(ws, 3, NOTE_COL, '说明', 't', bold=True)

    r = 4
    CHK_SUM = []
    def chkrow(label, fn, years, note=''):
        nonlocal r
        put(ws, r, 1, label, 't')
        for y in years:
            put(ws, r, 2 + YRS.index(y), fn(y), 'f', 'General', align='center')
        cl0 = YC[years[0]]; cl1 = YC[years[-1]]
        put(ws, r, NOTE_COL, note, 'g', size=9)
        CHK_SUM.append(f"{cl0}{r}:{cl1}{r}")
        r += 1

    def chkcell(label, formula, note=''):
        nonlocal r
        put(ws, r, 1, label, 't')
        put(ws, r, 2, formula, 'f', align='center')
        put(ws, r, NOTE_COL, note, 'g', size=9)
        CHK_SUM.append(f"B{r}")
        r += 1

    chkrow('1. BS配平: 资产=负债+权益 (差额<0.01)', lambda y: f"=ABS({BSs}!{YC[y]}{B['chk']})<0.01", YRS,
           '差额检查行见BS底部; 历史尾差已清零')
    chkrow('2. CF期末现金=BS货币资金 (<0.01)', lambda y: f"=ABS({CFs}!{YC[y]}{C['cash1']}-{BSs}!{YC[y]}{B['cash']})<0.01", FCST,
           '预测期逐年勾稽; BS现金=FIN期末现金=最低现金')
    chkrow('3. 未分配利润: 期末=期初+归母-分红-盈余公积计提', lambda y: f"=ABS({BSs}!{YC[y]}{B['re']}-{BSs}!{YC[y-1]}{B['re']}-{ISs}!{YC[y]}{I['np_p']}+{SCH}!{YC[y]}{S['div']}+{EQs}!{YC[y]}{Q['sr_acc']})<0.01", FCST,
           'Equity_Roll滚动勾稽')
    chkrow('4. IS收入=Revenue_Segments合计', lambda y: f"=ABS({ISs}!{YC[y]}{I['rev']}-{RS}!{YC[y]}{REV_TOT})<0.01", YRS,
           '分部合计自动汇总进IS')
    # E5: 毛利率带宽改为"信息行", 不进CHK_SUM布尔汇总 — 高/低毛利行业越界不应闸门全簿;
    #     区间从 checks.gm_band 读取(默认[0,0.6], 与原硬编码等值)
    _gm_band = cfg.get('checks.gm_band') or [0.0, 0.6]
    put(ws, r, 1, f'5. 信息项(不闸门): 毛利率处于({_gm_band[0]:.0%},{_gm_band[1]:.0%})区间', 't')
    for y in YRS:
        put(ws, r, 2 + YRS.index(y),
            f"=AND({ISs}!{YC[y]}{I['gm']}>{_gm_band[0]:g},{ISs}!{YC[y]}{I['gm']}<{_gm_band[1]:g})",
            'f', 'General', align='center')
    put(ws, r, NOTE_COL, '信息行, 不参与布尔汇总; 越界仅提示复核毛利率假设; 区间可经checks.gm_band配置', 'g', size=9)
    r += 1
    chkcell('6. DCF每股价值>0', f"={DCFs}!D{D['ps']}>0", 'DCF页引用校验')
    chkcell('7. Sensitivity矩阵一中心格=DCF每股价值', f"=ABS({SENs}!{M1_CENTER}-{DCFs}!D{D['ps']})<0.01",
            '矩阵轴自动对中WACC采用值/g, 中心格恒等于DCF输出')
    chkcell('8. WACC采用值=计算值 或 override非空',
            f"=OR(ABS({ASheet}!C{A['wacc']}-{ASheet}!C{A['wacc_calc']})<0.000001,{ASheet}!C{A['wacc_ovr']}<>\"\")",
            '采用值=IF(override="",计算值,override)的逻辑校验')
    chkrow('9. 货币资金>0', lambda y: f"={BSs}!{YC[y]}{B['cash']}>0.01", FCST,
           'FIN调度下现金=最低现金(收入×最低现金占比), 恒正')
    chkrow('10. 四档债务余额≥0', lambda y: f"=MIN({FINs}!{YC[y]}{F['st_t1']},{FINs}!{YC[y]}{F['cur_t1']},{FINs}!{YC[y]}{F['lt_t1']},{FINs}!{YC[y]}{F['le_t1']})>=-0.01", FCST,
           'sweep/还款均以MAX(0,·)兜底')
    chkrow('11. FIN迭代残差<0.01 (迭代展开收敛)', lambda y: f"={FINs}!{YC[y]}{F['resid_max']}<0.01", FCST,
           '第4轮采用值 vs 按期末余额重算隐含值; 全簿无循环引用, 无需开启迭代计算')
    chkcell('12. Named ranges存在性 (审计轨迹, 15个)',
            "=AND(" + ",".join(f'ISNUMBER(INDIRECT("{n}"))' for n in NAMED_RANGES) + ")",
            'WACC采用值/永续g/Ke/Kd/rf/ERP/βu/税率/股利/最低现金/情景开关/稀释股数/DCF&FCFE每股; 任一named range被删则FALSE')
    SCEN_MONO_CHECK_ROW = r
    chkcell('13. 三情景同引擎估值单调性',
            f"=AND({SCENs}!B{scen_rows['bear']['ps']}<={SCENs}!B{scen_rows['base']['ps']},"
            f"{SCENs}!B{scen_rows['base']['ps']}<={SCENs}!B{scen_rows['bull']['ps']},"
            f"{SCENs}!B{scen_rows['bear']['rel']}<={SCENs}!B{scen_rows['base']['rel']},"
            f"{SCENs}!B{scen_rows['base']['rel']}<={SCENs}!B{scen_rows['bull']['rel']})",
            '配置层已强制熊≤基≤牛调整; 此处确保同一简化引擎的DCF与相对估值输出保持单调')
    r += 1
    put(ws, r, 1, '信息项: 隐含折旧率(折旧/期初固资净值)', 't')
    for y in FCST:
        put(ws, r, 2 + YRS.index(y), f"={PPEs}!{YC[y]}{P['dep_rt']}", 'f', PCT, align='center')
    put(ws, r, NOTE_COL, '信息行: 合理带宽8%-18%', 'g', size=9); r += 1
    put(ws, r, 1, f'信息项: {F0}E模型收入/{CONS_REV_LABEL}', 't')
    put(ws, r, 2, f"={ISs}!{YC[F0]}{I['rev']}/{ASheet}!$C${A[f'c_rev{_cy[0]}']}", 'f', PCT)
    put(ws, r, NOTE_COL, f'接近100%即基准贴合{CONS_REV_LABEL}', 'g', size=9); r += 1
    put(ws, r, 1, f'信息项: {F0}E模型归母/{CONS_NP_LABEL}', 't')
    put(ws, r, 2, f"={ISs}!{YC[F0]}{I['np_p']}/{ASheet}!$C${A[f'c_np{_cy[0]}']}", 'f', PCT); r += 1
    put(ws, r, 1, '信息项: WACC计算值/采用值', 't')
    put(ws, r, 2, f"={ASheet}!C{A['wacc_calc']}", 'f', PCT)
    put(ws, r, 3, f"={ASheet}!C{A['wacc']}", 'f', PCT)
    put(ws, r, NOTE_COL, '左=计算值, 右=采用值(override为空时两者相等)', 'g', size=9); r += 1
    put(ws, r, 1, '信息项: FCFE/FCFF每股价值比', 't')
    put(ws, r, 2, f"={FCFEs}!D{E['ps']}/{DCFs}!D{D['ps']}", 'f', PCT)
    put(ws, r, 3, f"={FCFEs}!D{E['ps']}", 'f', PS)
    put(ws, r, NOTE_COL, f'左=FCFE/FCFF比值, 右=FCFE每股({PER_SHARE_UNIT}); 两法差异原因见FCFE页底部注记', 'g', size=9); r += 1
    put(ws, r, 1, '信息项: FCFE/FCFF差异控制 (阈值30%)', 't')
    put(ws, r, 2, fcfe_control_formula(f'{FCFEs}!D{E["diff"]}'),
        'f', bold=True, fill=CHK_FILL)
    put(ws, r, NOTE_COL, '信息警报, 不加入布尔汇总; 避免把方法论差异误判为机械勾稽失败', 'g', size=9)
    CHECKS_FCFE_CONTROL_ROW = r; r += 1
    put(ws, r, 1, '信息项: 完整主模型 / 同引擎基准桥接', 't')
    put(ws, r, 2, f"={SCENs}!B{SCEN_MAIN_PS_ROW}/{SCENs}!B{SCEN_BRIDGE_BASE_ROW}", 'f', PCT)
    put(ws, r, 3, f"={SCENs}!B{SCEN_MAIN_PS_ROW}", 'f', PS)
    put(ws, r, NOTE_COL, f'左=完整/简化基准比值, 右=完整主模型每股({PER_SHARE_UNIT}); 差异说明见Scenarios桥接区', 'g', size=9); r += 1
    put(ws, r, 1, '信息项: 主模型/同引擎基准控制 (阈值30%)', 't')
    put(ws, r, 2, f'=IF(ABS({SCENs}!B{SCEN_BRIDGE_RATIO_ROW}-1)>=0.3,"警告: 桥接差异≥30%, 复核简化假设","通过: 桥接差异<30%")',
        'f', bold=True, fill=CHK_FILL)
    put(ws, r, NOTE_COL, '信息警报, 不加入布尔汇总', 'g', size=9)
    CHECKS_BRIDGE_CONTROL_ROW = r; r += 1
    if INTERIM_STATUS == 'error':
        put(ws, r, 1, f'⚠ 中报抓取失败: {INTERIM_ERROR}', 'w', bold=True)
        put(ws, r, 2, 'ERROR', 'w', bold=True, fill=CHK_FILL, align='center')
        put(ws, r, NOTE_COL, '取数失败不等于not_available(无可用披露); 本行不加入机械布尔汇总, 但交付前必须复核数据时效性',
            'w', size=9, wrap=True)
        CHECKS_INTERIM_ERROR_ROW = r; r += 1
    else:
        CHECKS_INTERIM_ERROR_ROW = None
    if TP_ROW:
        put(ws, r, 1, '信息项: 一致预期目标价 vs DCF基准', 't')
        put(ws, r, 2, f"={RVs}!C{TP_ROW}", 'f', PS)
        put(ws, r, 3, f"={RVs}!C{TP_ROW}/{DCFs}!D{D['ps']}-1", 'f', PCT)
        put(ws, r, NOTE_COL, f'左=目标价({PS_DENOM}), 右=目标价较模型DCF溢价; 来源见Relative_Val目标价行', 'g', size=9); r += 1
    put(ws, r, 1, f'信息项: {F1}E现金 vs 最低现金', 't')
    put(ws, r, 2, f"={BSs}!{YC[F1]}{B['cash']}", 'f', NUM)
    put(ws, r, 3, f"={FINs}!{YC[F1]}{F['mincash']}", 'f', NUM)
    put(ws, r, NOTE_COL, '两者应相等(sweep后现金=最低现金, 溢出去理财)', 'g', size=9); r += 1
    # X1: 中报锚点信息行 — 仅cfg含interim段且报告期年份==首预测年F0且rev/months可年化时生成;
    #     追加在既有信息行之后、不加入CHK_SUM, 故不影响布尔闸门与汇总单元格(其行号按r变量顺延);
    #     无interim段(存量配置)时本块零输出, 工作簿与既有版本逐单元格等价
    _itm = cfg.raw.get('interim') or {}

    def _itm_num(v):
        # interim字段规整为float标量; None/缺失/非数值→None (字段允许 {value,basis} 包装)
        if isinstance(v, dict):
            v = _norm_entry(v)['values']
        return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

    _itm_lab = str(_itm.get('label') or '最新报告期')
    _itm_rev, _itm_np = _itm_num(_itm.get('rev')), _itm_num(_itm.get('np_p'))
    _itm_rev_pr = _itm_num(_itm.get('rev_prior'))
    _itm_basis = str(_itm.get('basis') or '')
    _itm_yr = None   # 报告期年份: 优先label前4位(如'2026Q1'), 兜底date前4位
    for _s in (_itm.get('label'), _itm.get('date')):
        _s = str(_s or '')
        if len(_s) >= 4 and _s[:4].isdigit():
            _itm_yr = int(_s[:4]); break
    _itm_mo = _itm.get('months')
    _itm_mo = int(_itm_mo) if isinstance(_itm_mo, (int, float)) and not isinstance(_itm_mo, bool) and int(_itm_mo) in (3, 6, 9) else None
    ITM_RATIO_CELL = None
    if _itm and _itm_yr == F0 and _itm_rev and _itm_mo:
        _itm_ann = _itm_rev * 12.0 / _itm_mo
        put(ws, r, 1, f'信息项: 首年营收预测 / {_itm_lab}年化', 't')
        put(ws, r, 2, f"={ISs}!{YC[F0]}{I['rev']}/{_itm_ann:.10g}", 'f', PCT, align='center')
        put(ws, r, NOTE_COL,
            f'信息行, 不参与布尔汇总; 除数为硬数, 口径: {_itm_lab}营收{_itm_rev:,.1f}×12/{_itm_mo}={_itm_ann:,.1f}百万; {_itm_basis}',
            'g', size=9)
        ITM_RATIO_CELL = f'B{r}'; r += 1
        put(ws, r, 1, '信息项: 偏离提示', 't')
        put(ws, r, 2,
            f"=IF(AND({ITM_RATIO_CELL}>=0.7,{ITM_RATIO_CELL}<=1.3),\"±30%内\",\"偏离超±30%, 请核对首年假设\")",
            'f', align='center')
        put(ws, r, NOTE_COL, f'上行比值0.7~1.3之外提示复核首年假设; {_itm_basis}', 'g', size=9); r += 1
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    put(ws, r, 1, '汇总: 全部通过 = ', 't', bold=True, size=14, align='right')
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cc = ws.cell(row=r, column=6, value=f"=IF(AND({','.join(CHK_SUM)}),\"全部通过=TRUE\",\"存在未通过=FALSE\")")
    cc.font = Font(name=FONT, size=16, bold=True, color=C_RED)
    cc.fill = CHK_FILL
    cc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 30
    SUM_CELL = f"F{r}"
    r += 1
    put(ws, r, 1, '布尔汇总', 't')
    put(ws, r, 6, f"=AND({','.join(CHK_SUM)})", 'f', bold=True, align='center')
    BOOL_CELL = f"F{r}"
    CHKs = 'Checks'

    # ============================================================
    # Cover
    # ============================================================
    _cover = cfg.get('cover') or {}
    ws = wb.create_sheet('Cover', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 60
    for col in 'DEFGH':
        ws.column_dimensions[col].width = 13
    r = 2
    ws.merge_cells('B2:H2')
    c = ws.cell(row=2, column=2, value=f'{NAME} ({CODE_FULL}) — 机构级估值模型')
    c.font = Font(name=FONT, size=20, bold=True, color='FF1F3864')
    ws.merge_cells('B3:H3')
    c = ws.cell(row=3, column=2, value=_cover.get('subtitle', ''))
    c.font = Font(name=FONT, size=11, color=C_GREY)
    rows_cover = [
        ('模型构建日期', f'{BUILD_DATE} (行情基准日: {VAL_DATE}收盘; FIN调度/WACC做实/年中折现; FIN迭代展开消除循环引用)'),
        ('经营货币 / 单位', cfg.get('company.currency_note', f'{CURRENCY_CODE} / 百万元 (每股数据为{PS_DENOM})')),
        ('颜色规范', '蓝色=输入/假设 | 黑色=公式 | 绿色=外部数据/链接 | 红色=警告'),
        ('✓ 迭代展开(无循环引用)', '原循环链(利息=平均债务余额×利率→净利→现金流→sweep→债务余额→利息)已在FIN页表内显式展开4轮(第1轮期初余额计息, 第k轮按(期初+上轮期末)/2×利率重算), 末轮输出供IS/CF/BS引用。全簿无任何循环引用, 无需开启迭代计算; 收敛残差<0.01(见FIN页残差行/Checks第11项), Excel/WPS/LibreOffice重算结果一致'),
        ('', ''),
        ('—— 关键市场数据 ——', ''),
        (f'现价 ({PS_DENOM})', f"={ASheet}!C{A['px']}"),
        ('总股本 (百万股)', f"={ASheet}!C{A['sh']}"),
        ('总市值 (百万元)', f"={ASheet}!C{A['mcap']}"),
        (f'{CONS_REV_LABEL}/{CONS_NP_LABEL} {F0}E 营收/归母 (百万)', f"=\"营收\"&TEXT({ASheet}!C{A[f'c_rev{_cy[0]}']},\"#,##0\")&\" / 归母\"&TEXT({ASheet}!C{A[f'c_np{_cy[0]}']},\"#,##0\")&\" ({cons_rev['basis']})\""),
        ('', ''),
        ('—— 模型输出 (随情景开关联动) ——', ''),
        ('当前情景', f"={SCENs}!C3"),
        (f'模型{F0}E营收 (百万)', f"={ISs}!{YC[F0]}{I['rev']}"),
        (f'模型{F0}E归母 (百万)', f"={ISs}!{YC[F0]}{I['np_p']}"),
        ('WACC 计算值/采用值', f"=TEXT({ASheet}!C{A['wacc_calc']},\"0.00%\")&\" / \"&TEXT({ASheet}!C{A['wacc']},\"0.00%\")"),
        (f'DCF每股价值-完整主模型 ({PS_DENOM})', f"={DCFs}!D{D['ps']}"),
        (f'DCF每股-熊/基/牛同引擎 ({PS_DENOM})', f"=TEXT({SCENs}!B{scen_rows['bear']['ps']},\"0.00\")&\" / \"&TEXT({SCENs}!B{scen_rows['base']['ps']},\"0.00\")&\" / \"&TEXT({SCENs}!B{scen_rows['bull']['ps']},\"0.00\")"),
        (f'相对估值区间 ({PS_DENOM})', f"=TEXT({RVs}!C{P_LO},\"0.00\")&\" ~ \"&TEXT({RVs}!C{P_HI},\"0.00\")"),
        (f'综合参考区间 ({PS_DENOM}, Summary页)', f"=TEXT({SUMMARYs}!C{ENV_ROW},\"0.00\")&\" ~ \"&TEXT({SUMMARYs}!D{ENV_ROW},\"0.00\")"),
        ('Checks校验', f"={CHKs}!{SUM_CELL}"),
    ]
    if IS_HK:
        rows_cover.insert(2, ('市场口径 (港股)',
                              '无单日涨跌停限制(A股为±10%/20%, 口径差异不影响模型公式); 财务与估值为财报币种口径(见上行), '
                              '港元现价/总市值按配置fx折算, PE-TTM为港元行情口径仅供对照; 历史三表=东财HKF10(IFRS科目映射, 现金流量表按隐含汇率折算回财报币种)'))
    if INTERIM_STATUS == 'error':
        rows_cover.insert(3, (f'⚠ 中报抓取失败: {INTERIM_ERROR}',
                              '该状态与not_available(确认无可用披露)不同; 当前模型可能缺少最新中期锚点, 交付前须人工复核'))
    r = 5
    for k, v in rows_cover:
        if k.startswith('——'):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            put(ws, r, 2, k, 't', bold=True, size=12)
        else:
            put(ws, r, 2, k, 't', bold=True)
            if k.startswith(('⚠', '✓')):
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
                put(ws, r, 3, v, 'w', size=9, wrap=True)
                ws.row_dimensions[r].height = 30
            else:
                put(ws, r, 3, v, 'f' if isinstance(v, str) and v.startswith('=') else 't')
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    put(ws, r, 2, '—— 关键建模处理 (分析师注记) ——', 't', bold=True, size=12); r += 1
    _default_notes = [
        f'1. FIN页债务调度: 资金缺口=MAX(0,期初现金+经营CF+投资CF+计划分红+计划还款+利息-最低现金)→revolver新增借款; 超额现金先sweep偿还短借/长借, 债务清零后溢出购买理财(交易性金融资产); 货币资金=最低现金(收入×{a1("financing.min_cash_pct"):.0%}), 杜绝"高现金+高借款"并存;',
        '2. 利息=四档债务各自(期初+期末)/2平均余额×分档利率; 该循环依赖已在FIN页"迭代展开"区表内展开4轮(与外部不动点法同构且实测6轮内收敛至1e-8, 4轮充足), 末轮输出供IS/CF/BS引用; 全簿无循环引用, 无需开启迭代计算, LibreOffice headless可直接重算验收;',
        '3. BS配平项由货币资金改为FIN页revolver新增借款/现金sweep; 历史尾差已并入"其他权益项目(轧差)"清零, 配平差额全期恒=0(容差0.01);',
        '4. WACC做实: 可比公司βl按各自D/E与税率unlever取中位数(Relative_Val页L-O列), 按公司市值口径目标结构relever; Wd=最近年报末有息负债/(有息负债+总市值); 采用值=IF(override="",计算值,override), 下游全部引用采用值;',
        f'5. DCF年中折现: 估值基准日{VAL_DATE}, t=0.5/1.5/.../{NF-0.5}; EV→Equity桥含货币资金/交易性金融资产/其他权益工具投资/有息负债/少数股东权益;',
        '6. CF口径: 利息支出经营活动加回、筹资活动列"偿付利息"; 投资活动含"理财净变动"行;',
        '7. 权益滚动: Equity_Roll页逐项滚动, 盈余公积按归母计提; 稀释股数=总股本+稀释工具占位; Summary页为football field区间汇总;',
        '8. 预测期资产减值/投资收益等非经常项简化为固定小额(Assumptions"其他经营损益"), 以保证三表严格配平;',
        f'9. 分产品历史收入=产品占比×总收入; 预测按"量×价"驱动并受情景开关调整; 预测期({F0}E-{F1}E)所有单元格均为公式, 硬编码仅限历史实际值(绿色)与蓝色假设输入;',
    ]
    for n in (_cover.get('notes') or _default_notes):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        put(ws, r, 2, n, 'g', size=10, wrap=True)
        ws.row_dimensions[r].height = 40 if len(n) > 80 else 26
        r += 1
    # X2: 中报注记 — 仅cfg含interim段时在建模注记区末追加一行(样式与上方注记行一致);
    #     数字取自interim字段(百万→亿), prior缺失省略同比; anchor非true改写"未参与校准仅供对照"
    if _itm:
        _p = []
        if _itm_rev is not None:
            _p.append(f'营收 {_itm_rev / 100:.1f} 亿')
        if _itm_np is not None:
            _p.append(f'归母 {_itm_np / 100:.1f} 亿')
        _itm_note = f'最新报告期 {_itm_lab}: ' + ' / '.join(_p)
        if _itm_rev is not None and _itm_rev_pr:
            _itm_note += f'(同比: 营收 {_itm_rev / _itm_rev_pr - 1:+.1%})'
        _itm_note += ', 兜底预测已按其年化校准' if _itm.get('anchor') is True else ', 未参与校准仅供对照'
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        put(ws, r, 2, _itm_note, 'g', size=10, wrap=True)
        ws.row_dimensions[r].height = 40 if len(_itm_note) > 80 else 26
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    put(ws, r, 2, '—— 数据来源 ——', 't', bold=True, size=12); r += 1
    _default_sources = [
        f'历史三表: 东方财富F10 ({HIST[0]}A-{H0}A年报); 分产品结构: 东财F10主营构成/招股书/券商深度报告(以配置"依据"字段为准);',
        f'{CONS_REV_LABEL}: {cons_rev["basis"]}; 行情与可比: 腾讯行情({VAL_DATE});',
        '可比公司βl/D/E: 分析师输入(参考行情终端β区间与各公司最新年报杠杆), 可按终端数据更新.',
    ]
    for n in (_cover.get('sources') or _default_sources):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        put(ws, r, 2, n, 'x', size=9, wrap=True)
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=8)
    put(ws, r, 2, '免责声明: 本模型仅为研究学习用途, 所有预测基于公开信息与主观假设, 不构成任何投资建议或证券买卖要约。历史数据虽经核对来源, 仍可能存在口径差异或错误; 预测存在重大不确定性。使用者应自行判断并承担风险。',
        'g', size=9, wrap=True)

    # ============================================================
    # 研究层产出页 (DR研究 / 研究摘要; 仅 --dr/--consensus/--announcements/--llm 激活时生成)
    # ============================================================
    EXTRA_SHEETS = []
    if research:
        from research.sheets import build_dr_sheet, build_summary_sheet
        if research.get('dr'):
            build_dr_sheet(wb, research['dr'], put, title_bar, sec)
            EXTRA_SHEETS.append('DR研究')

        trows = []   # (假设项, 采用值, 来源, 置信度, 备注)
        trows.append(('WACC 采用值', ('f', f'=TEXT({ASheet}!C{A["wacc"]},"0.00%")&" (计算值 "&TEXT({ASheet}!C{A["wacc_calc"]},"0.00%")&")"'),
                      '可比公司β unlever/relever实算(Relative_Val页L-O列→Assumptions第十节)', '中(实算)', '采用值=IF(override="",计算值,override)'))
        trows.append(('永续增长率 g', f"{a1('wacc.tg'):.1%}", ab('wacc.tg'), '中(分析师假设)', 'Sensitivity页做WACC×g双维敏感性'))
        trows.append((f'有效税率 ({F0}E)', f"{a5('opex.tax_rate')[0]:.1%}", ab('opex.tax_rate'), '中(分析师假设)', ''))
        trows.append(('Capex占收入比路径', ' → '.join(f"{v:.0%}" for v in a5('capex.capex_rate')),
                      ab('capex.capex_rate'), '中(分析师假设)', ''))
        trows.append(('DSO/DIO/DPO路径 (天)',
                      ' / '.join(' → '.join(f"{v:.0f}" for v in a5(f'working_capital.{k}')) for k in ('dso', 'dio', 'dpo')),
                      ab('working_capital.dso'), '中(分析师假设)', 'DPO单维敏感性见Sensitivity矩阵三'))
        trows.append(('股利支付率', f"{a5('dividend.payout')[0]:.0%}", ab('dividend.payout'), '中(分析师假设)', ''))
        _scb, _scu = cfg.raw['scenarios']['bear'], cfg.raw['scenarios']['bull']
        trows.append(('情景调整 (熊/牛)', f"增速{_scb['rev_adj']:+.0%}/{_scu['rev_adj']:+.0%}, 净利率{_scb['npm_adj']:+.1%}/{_scu['npm_adj']:+.1%}",
                      f"熊: {_scb.get('logic', '')}; 牛: {_scu.get('logic', '')}", '低(情景假设)', '基准=主模型输出'))

        if research.get('dr'):
            dr = research['dr']
            for f in dr['findings'][:12]:
                trows.append((f"档案§{f['section']} {f['title']}", f['text'],
                              f"dr档案({os.path.basename(dr['path'])})", '高(研究档案)',
                              '已按关键词回填相关假设的依据列(标注dr档案§章节号)'))

        if research.get('consensus'):
            cc = research['consensus']
            csrc = f"{cc['source']}{' ' + cc['date'] if cc['date'] else ''} (--consensus文件)"
            for i, y in enumerate(FCST[:3]):
                if i < len(cons_rev_v):
                    _rev_verified = CONSENSUS_VERIFIED['rev'][y]
                    trows.append((f'{CONSENSUS_LABELS["rev"][y]}{y}E营收', f'{cons_rev_v[i]:,.0f}百万',
                                  csrc if _rev_verified else cons_rev['basis'],
                                  '中高(卖方一致预期)' if _rev_verified else '低(未由文件覆盖)',
                                  ('f', f'="模型 "&TEXT({ISs}!{YC[y]}{I["rev"]},"#,##0")&" 百万, 占一致预期 "'
                                        f'&TEXT({ISs}!{YC[y]}{I["rev"]}/{cons_rev_v[i]},"0.0%")')))
                if i < len(cons_np_v):
                    _np_verified = CONSENSUS_VERIFIED['np'][y]
                    trows.append((f'{CONSENSUS_LABELS["np"][y]}{y}E归母', f'{cons_np_v[i]:,.0f}百万',
                                  csrc if _np_verified else cons_np['basis'],
                                  '中高(卖方一致预期)' if _np_verified else '低(未由文件覆盖)',
                                  ('f', f'="模型 "&TEXT({ISs}!{YC[y]}{I["np_p"]},"#,##0")&" 百万, 占一致预期 "'
                                        f'&TEXT({ISs}!{YC[y]}{I["np_p"]}/{cons_np_v[i]},"0.0%")')))
            if cc.get('target_price'):
                trows.append(('一致预期目标价', f"{float(cc['target_price']):.2f}{PS_DENOM}", csrc, '中高(卖方一致预期)',
                              ('f', f'="模型DCF "&TEXT({DCFs}!D{D["ps"]},"0.00")&" {PS_DENOM}, 目标价较模型 "'
                                    f'&TEXT({cc["target_price"]}/{DCFs}!D{D["ps"]}-1,"+0.0%;-0.0%")')))

        for it in (research.get('announcements') or {}).get('items') or []:
            trows.append(('公告要点', it, '东财业绩预告/业绩快报(系统curl)', '高(公司公告事实)', ''))

        build_summary_sheet(wb, research, trows, research.get('memo'), put, title_bar, sec)
        EXTRA_SHEETS.append('研究摘要')

    # ---------- 计算设置 ----------
    # H1: 不开启Excel迭代计算 — 全簿无循环引用, 保持默认(不迭代)可避免掩盖未来意外引入的循环引用;
    #     仅保留打开时全量重算
    wb.calculation.fullCalcOnLoad = True

    # ---------- Named ranges (关键驱动审计轨迹; FCFE/Checks公式已引用) ----------
    from openpyxl.workbook.defined_name import DefinedName
    for _n, _ref in NAMED_RANGES.items():
        wb.defined_names[_n] = DefinedName(_n, attr_text=_ref)

    # ---------- 工作表顺序 (公开API move_sheet 重排) ----------
    ORDER = ['Cover', 'Summary', 'Assumptions', 'Revenue_Segments', 'IS', 'BS', 'CF',
             'Schedules', 'PPE', 'FIN', 'Equity_Roll', 'DCF', 'FCFE', 'Relative_Val',
             'Sensitivity', 'Scenarios', 'Checks'] + EXTRA_SHEETS
    for _i, _n in enumerate(ORDER):
        wb.move_sheet(_n, offset=_i - wb.sheetnames.index(_n))

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print('SAVED', _public_source_reference(out_path)[0])

    _config_ref, _config_scope = _public_source_reference(getattr(cfg, 'source_path', None))
    addr = {
        'meta': {'code': co.get('code'), 'code_full': CODE_FULL, 'name': NAME,
                 'hist_years': HIST, 'fcst_years': FCST, 'valuation_date': VAL_DATE,
                 'research_sheets': EXTRA_SHEETS,
                 'config_path': _config_ref,
                 'config_path_scope': _config_scope,
                 'config_sha256': getattr(cfg, 'config_sha256', None),
                 'is_default_config': bool(getattr(cfg, 'is_default_config', False)),
                 'currency_code': CURRENCY_CODE, 'per_share_unit': PER_SHARE_UNIT,
                 'historical_plugs': [
                     {k: plug[k] for k in ('year', 'diff', 'assets', 'ratio')}
                     for plug in HISTORICAL_PLUGS
                 ],
                 'scenario_engine': 'simplified_same_engine_v1',
                 'consensus_rev_verified': CONSENSUS_VERIFIED['rev'][F0],
                 'consensus_np_verified': CONS_NP_VERIFIED,
                 'consensus_verified_by_year': {
                     key: {str(y): verified for y, verified in by_year.items()}
                     for key, by_year in CONSENSUS_VERIFIED.items()
                 },
                 'fcfe_divergence_waiver': FCFE_DIVERGENCE_WAIVER,
                 'interim_status': INTERIM_STATUS},
        'dcf_ps': f"DCF!D{D['ps']}", 'dcf_upside': f"DCF!D{D['upside']}",
        'dcf_ev': f"DCF!D{D['ev']}", 'dcf_eq': f"DCF!D{D['eq']}",
        'dcf_tidx': f"DCF!{YC[F0]}{D['t_idx']}",   # E3: 折现年序行首个预测年单元格(verify端读行号)
        'fcfe_ps': f"FCFE!D{E['ps']}", 'fcfe_eq': f"FCFE!D{E['eq']}",
        'fcfe_fcff_ps': f"FCFE!D{E['ps_fcff']}", 'fcfe_diff': f"FCFE!D{E['diff']}",
        'named_ranges': NAMED_RANGES,
        'bs_chk_row': B['chk'], 'checks_sum': f"Checks!{SUM_CELL}", 'checks_bool': f"Checks!{BOOL_CELL}",
        'scen_bear_ps': f"Scenarios!B{scen_rows['bear']['ps']}",
        'scen_base_ps': f"Scenarios!B{scen_rows['base']['ps']}",
        'scen_bull_ps': f"Scenarios!B{scen_rows['bull']['ps']}",
        'scen_bear_rel': f"Scenarios!B{scen_rows['bear']['rel']}",
        'scen_base_rel': f"Scenarios!B{scen_rows['base']['rel']}",
        'scen_bull_rel': f"Scenarios!B{scen_rows['bull']['rel']}",
        'scen_main_ps': f"Scenarios!B{SCEN_MAIN_PS_ROW}",
        'scen_full_simplified_ratio': f"Scenarios!B{SCEN_BRIDGE_RATIO_ROW}",
        'scen_full_simplified_status': f"Scenarios!B{SCEN_BRIDGE_STATUS_ROW}",
        'scen_monotonic_check': f"Checks!B{SCEN_MONO_CHECK_ROW}",
        'is_rev': {y: f"IS!{YC[y]}{I['rev']}" for y in YRS},
        'is_np': {y: f"IS!{YC[y]}{I['np_p']}" for y in YRS},
        'rel_lo': f"Relative_Val!C{P_LO}", 'rel_hi': f"Relative_Val!C{P_HI}",
        'rel_med_pe': REL_MED_PE,
        'relative_price_rows': [f"Relative_Val!F{rr}" for rr in PRICE_ROWS],
        'rel_median_cell': f"Relative_Val!$F${MED_ROW}",
        'relative_f1_rows': [f"Relative_Val!H{rr}" for rr in F1_ROWS],
        'rel_f1_median_cell': f"Relative_Val!$H${MED_ROW}",
        'relative_peg_rows': [f"Relative_Val!J{rr}" for rr in PEG_ROWS],
        'rel_peg_median_cell': f"Relative_Val!$J${MED_ROW}",
        'sens_center': f"Sensitivity!{M1_CENTER}",
        'wacc_calc': f"Assumptions!C{A['wacc_calc']}", 'wacc_used': f"Assumptions!C{A['wacc']}",
        'wacc_ovr': f"Assumptions!C{A['wacc_ovr']}", 'ke': f"Assumptions!C{A['ke']}",
        'beta_u': f"Assumptions!C{A['beta_u']}", 'beta_l': f"Assumptions!C{A['beta']}",
        'wd': f"Assumptions!C{A['wd']}",
        'rv_beta_rows': [f"Relative_Val!L{rr}" for rr in BETA_ROWS],
        'fin_cash1_row': F['cash1'], 'fin_mincash_row': F['mincash'],
        'fin_intexp_row': F['intexp'], 'fin_dtot1_row': F['dtot1'],
        'fin_st_t1': F['st_t1'], 'fin_lt_t1': F['lt_t1'], 'fin_tfa1': F['tfa1'],
        'bs_cash_row': B['cash'], 'bs_tfa_row': B['tfa'],
        'eq_shares': f"Equity_Roll!C{Q['shares']}",
        'summary_env': f"Summary!C{ENV_ROW}", 'summary_env_hi': f"Summary!D{ENV_ROW}",
        'dpo_center': f"Sensitivity!{DPO_CENTER}",
        'is_fin_row': I['fin'], 'fin_finnet_row': F['finnet'], 'fin_intinc_row': F['intinc'],
        'cf_intadd_row': C['intadd'], 'fin_intexp_row2': F['intexp'], 'fin_intpay_row': F['intpay'],
        'fin_resid_max_row': F['resid_max'], 'fin_resid_int_row': F['resid_int'],
        'fin_resid_inc_row': F['resid_inc'], 'fin_resid_ok_row': F['resid_ok'],
        'fin_iter_resid_rows': [IT[k]['resid'] for k in sorted(IT)],
        'cf_cash1_row': C['cash1'],
        'fcfe_control': f"FCFE!D{E['control']}",
        'checks_fcfe_control': f"Checks!B{CHECKS_FCFE_CONTROL_ROW}",
        'summary_fcfe_control': f"Summary!B{SUMMARY_FCFE_CONTROL_ROW}",
    }
    # X3: 可选键 — 仅interim信息行生成时写入(指向首年预测/中报年化比值单元格, 供人工查阅, verify不依赖)
    if ITM_RATIO_CELL:
        addr['chk_interim_ratio'] = f"Checks!{ITM_RATIO_CELL}"
    with open(addr_path, 'w', encoding='utf-8') as f:
        json.dump(addr, f, ensure_ascii=False, indent=1)
    print('ADDR', _public_source_reference(addr_path)[0])
    return addr


def _argument_parser():
    ap = argparse.ArgumentParser(description='A股机构级估值模型生成器 (DCF+FCFE+三表+FIN调度, 17 sheets)')
    ap.add_argument('--code', help='A股代码, 如 300476; 读取 configs/<code>.yaml')
    ap.add_argument('--config', help='直接指定YAML配置路径')
    ap.add_argument('--out', help='输出xlsx路径 (默认 out/<code>_<名称>_估值模型.xlsx)')
    ap.add_argument('--addr', help='单元格地址json输出路径 (默认随xlsx同名)')
    ap.add_argument('--dr', help='研究层: 导入深度研究档案(md), 量化结论回填假设依据 + DR研究/研究摘要两页')
    ap.add_argument('--consensus', help='研究层: 聚源/gildata一致预期文件(json/csv, 含目标价)')
    ap.add_argument('--announcements', action='store_true',
                    help='研究层: 东财业绩预告/快报(系统curl)最新一期要点进研究摘要')
    ap.add_argument('--llm', nargs='?', const='auto', default='off',
                    choices=['auto', 'codex', 'claude', 'kimi', 'off'],
                    help='研究层: 本机LLM CLI生成研究备忘录; 裸--llm等于auto, 默认off')
    ap.add_argument('--refresh', action='store_true',
                    help='忽略已存配置，在内存中重新抓取并构建公开数据兜底模型(手工假设不会沿用)')
    ap.add_argument('--as-of', help='数据截止日 YYYY-MM-DD；拒绝估值日晚于该日的前视配置')
    ap.add_argument('--stale-after-days', type=int, default=30,
                    help='估值日距as-of超过该天数时告警(默认30)')
    ap.add_argument('--fail-on-stale', action='store_true', help='把配置过期告警升级为硬失败')
    ap.add_argument('--require-interim', action='store_true',
                    help='要求存在成功且启用的最新季报/中报锚点，否则硬失败')
    return ap


def main():
    ap = _argument_parser()
    args = ap.parse_args()
    if not args.code and not args.config:
        ap.error('需提供 --code 或 --config')
    cfg = load_config(
        code=args.code, config_path=args.config, refresh=args.refresh, as_of=args.as_of,
        stale_after_days=args.stale_after_days, fail_on_stale=args.fail_on_stale,
        require_interim=args.require_interim,
    )
    research = None
    if args.dr or args.consensus or args.announcements or args.llm != 'off':
        from research import collect_research
        research = collect_research(cfg, dr_path=args.dr, consensus_path=args.consensus,
                                    want_announcements=args.announcements, llm=args.llm)
    code = args.code or cfg.get('company.code')
    name = cfg.get('company.name', code)
    out_path = args.out or os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                        'out', f'{code}_{name}_估值模型.xlsx')
    addr_path = args.addr or os.path.splitext(out_path)[0] + '.addr.json'
    build(cfg, out_path, addr_path, research=research)


if __name__ == '__main__':
    main()
