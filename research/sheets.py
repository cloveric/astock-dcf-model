# -*- coding: utf-8 -*-
"""研究层产出页: DR研究(档案全文存档) / 研究摘要(假设×采用值×来源×置信度)
由 build_model.py 传入 put/title_bar/sec 样式助手, 避免循环导入。"""

DR_SHEET = 'DR研究'
SUM_SHEET = '研究摘要'


def build_dr_sheet(wb, dr, put, title_bar, sec, max_rows=400):
    """DR研究: 深度研究档案全文存档 (逐行落表, 可检索可追溯)"""
    ws = wb.create_sheet(DR_SHEET)
    title_bar(ws, 'DR研究 — 深度研究档案全文存档',
              f"来源文件: {dr['path']} | 共{len(dr['sections'])}节, 提取量化结论{len(dr['findings'])}条; "
              '量化结论已回填Assumptions依据列(标注dr档案§章节号)', 8)
    ws.column_dimensions['A'].width = 130
    r = 4
    for s in dr['sections']:
        sec(ws, r, f"§{s['no']} {s['title']}", 8)
        r += 1
        for ln in s['lines']:
            if r - 4 > max_rows:
                put(ws, r, 1, f'... (存档截断于{max_rows}行, 全文见来源文件)', 'g', size=9)
                return ws
            put(ws, r, 1, ln, 't', size=9, wrap=True)
            ws.row_dimensions[r].height = 14 if len(ln) < 90 else 26
            r += 1
    return ws


def build_summary_sheet(wb, ctx, table_rows, memo, put, title_bar, sec):
    """研究摘要: 假设项×采用值×来源×置信度 对照表 + LLM备忘录页眉 + 公告要点
    table_rows: [(假设项, 采用值, 来源, 置信度, 备注)]"""
    ws = wb.create_sheet(SUM_SHEET)
    title_bar(ws, '研究摘要 — 假设 grounding 对照表',
              '假设项 × 采用值 × 来源 × 置信度 | 来源优先级: dr档案 > 聚源/gildata一致预期 > 公告要点 > 配置依据; '
              '本页仅为研究留痕, 不参与模型计算', 10)
    for col, w in zip('ABCDE', [30, 34, 44, 12, 60]):
        ws.column_dimensions[col].width = w
    r = 4

    if memo:
        sec(ws, r, '研究备忘录 (LLM生成, 仅供分析师参考, 不进入模型计算)', 10)
        r += 1
        for para in memo.split('\n'):
            if not para.strip():
                continue
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            put(ws, r, 1, para.strip(), 'g', size=9, wrap=True)
            ws.row_dimensions[r].height = 14 if len(para) < 100 else 26
            r += 1
        r += 1

    hdr = ['假设项', '采用值', '来源', '置信度', '备注']
    for i, h in enumerate(hdr):
        c = put(ws, r, 1 + i, h, 't', bold=True, align='center')
        from openpyxl.styles import PatternFill
        c.fill = PatternFill('solid', fgColor='FFF2F2F2')
    r += 1
    for item, val, src, conf, note in table_rows:
        put(ws, r, 1, item, 't', size=9)
        for ci, v in ((2, val), (3, src), (5, note)):
            if isinstance(v, tuple):        # ('f', 公式) → 活公式单元格
                put(ws, r, ci, v[1], 'f', size=9, wrap=True)
            else:
                put(ws, r, ci, v, 'g' if ci != 2 else 't', size=9, wrap=True)
        put(ws, r, 4, conf, 't', size=9, align='center')
        ws.row_dimensions[r].height = 24 if max(len(str(val)), len(str(src))) > 40 else 14
        r += 1

    if ctx.get('notes'):
        r += 1
        sec(ws, r, '降级记录 (来源不可用时的降级说明)', 10)
        r += 1
        for n in ctx['notes']:
            put(ws, r, 1, n, 'g', size=9)
            r += 1
    return ws
